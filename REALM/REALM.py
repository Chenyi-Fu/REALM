from __future__ import print_function
import sys

import cplex
from cplex.exceptions import CplexError
import random
import time
from math import log, exp, fabs
from scipy import stats
from numpy import array, std, var
# import Queue
import numpy
# import Datasetnew
from itertools import product
import os
import copy
import xlwt
import openpyxl
import gurobipy as gp
from gurobipy import GRB

def get_keys(d, value):
    return [k for k,v in d.items() if v == value]

class REALM():
    def __init__(self, name=None):
        self.__M = 10000
        self.__Mset = [i for i in range(10000)]
        self.compartment = 0
        self.compartment_name = {}
        self.compartment_info = {}
        self.group = None
        self.Time = None

        self.solver = 'cplex'

        self.gap = 5*10**(-5)
        self.Lq = None
        self.Lp = None
        self.vunder = None
        self.vover = None
        self.uunder = None
        self.uover = None

        self.label = None

        self.name = name
        self.model = None
        self.__approximation = 'ME'
        self.types = 'Expectation'
        self.theta = 50000
        self.objtype = 0

        self.objective_type = 'min'
        self.expobj = []
        self.robustobj = []
        self.constraint = {}
        self.quadraticconstraint = {}
        self.robustconstraint = {}
        self.objectiveterm = []
        self.custom_var = {}
        self.con_var_para = []
        self.con_var_para_qp = []
        self.__initialsol = {}

        self.latin_params1 = {}
        self.latin_params2 = {}
        self.latin_params3 = {}
        self.x = None
        self.__log_stream = 0
        self.__log_file = 'model_log_stream'

    # def set_data(self, data):
    #     for key, value in data.items():
    #         if hasattr(self, key) and key != 'solver':
    #             setattr(self, key, value)
    '''=============== Step 1. Define model and model size ====================='''
    def set_solver(self, solver='cplex'):
        if solver == 'cplex':
            setattr(self, 'solver', 'cplex')
            setattr(self, 'model', cplex.Cplex())
        elif solver == 'gurobi':
            setattr(self, 'solver', 'gurobi')
            setattr(self, 'model', gp.Model())

    def set_types(self, types = None):
        if types == 'Robustness' or types == 'Expectation':
            setattr(self, 'types', types)
        else:
            raise ValueError('Value of types should be Robustness or Expectation')

    def set_group(self, group=None):
        '''Set the number of groups in all compartments'''
        if isinstance(group, int):
            setattr(self, 'group', group)
        else:
            raise ValueError('The number of compartments should be integer')

    def set_time(self, time = None):
        '''Set the number of time periods'''
        if isinstance(time, int):
            setattr(self, 'Time', time)
        else:
            raise ValueError('The number of time horizons should be integer')


    '''=============== Step 2. Define name of compartment ====================='''
    def set_all_compartment(self,name=None):
        '''Define all new compartments'''
        '''name: a list and each element should be a string'''
        for i in name:
            self.set_compartment(compartment_name=i)

    def set_compartment(self, compartment_name= None, opt = None):
        '''Define new compartment'''
        '''opt = rename or delete the key'''
        if isinstance(compartment_name, str):
            temp = self.compartment_name.values()
            temp = list(temp)
            if compartment_name not in temp:
                self.compartment_name.update({self.compartment: compartment_name})
                self.compartment_info.update({self.compartment: {'local': {'name': compartment_name,
                                                                           'population': None,
                                                                            'p_idp': None,
                                                                            'p_isdp': [],
                                                                            'p_dp': {}},
                                                                 'tocpm': {},
                                                                 'fromcpm': {}
                                                                 }
                                              })
                self.__Mset.remove(self.compartment)
                self.compartment = min(self.__Mset)
            else:

                if isinstance(opt, list) and opt[0] == 'rename':
                    index = temp.index(compartment_name)
                    temp1 = self.compartment_name.keys()
                    temp1 = list(temp1)

                    self.compartment_name[temp1[index]] = opt[1]
                    self.compartment_info[temp1[index]]['local']['name'] = opt[1]
                elif opt == 'delete':
                    index = temp.index(compartment_name)
                    del self.compartment_name[index]
                    del self.compartment_info[index]
                    self.__Mset.append(index)
                else:
                    raise ValueError('This name of compartment exists, please rename it')

    '''=============== Step 3.1. Define transition between compartment and compartment  ========='''
    def set_transition_compartment(self, n=None, m=None, prob=None, opt = None):
        '''Define decision-independent transition probability from compartments n to m'''
        if isinstance(n, str):
            n = get_keys(self.compartment_name, n)[0]
        if isinstance(m, str):
            m = get_keys(self.compartment_name, m)[0]
        temp = self.compartment_info.keys()
        if n in temp and m in temp:
            if isinstance(prob, (int, float)):
               pp = [[prob for t in range(self.Time)] for j in range(self.group)]
            elif len(prob) == self.group:
                if all(isinstance(prob[j], (int, float)) for j in range(self.group)):
                    pp = [[prob[j] for t in range(self.Time)] for j in range(self.group)]
                elif all( len(prob[j]) >= self.Time for j in range(self.group)):
                    pp = [[prob[j][t] for t in range(len(prob[j]))] for j in range(self.group)]
                else:
                    raise ValueError('Domain of prob is incorrect')
            elif len(prob) == self.Time:
                pp = [[prob[t] for t in range(self.Time)] for j in range(self.group)]
            else:
                raise ValueError('Domain of prob is incorrect')
            if m in self.compartment_info[n]['tocpm'].keys():
                if opt == 'delete':
                    del self.compartment_info[n]['tocpm'][m]
                else:
                    self.compartment_info[n]['tocpm'][m]['q_idp'] = pp
            elif m not in self.compartment_info[n]['tocpm'].keys():
                self.compartment_info[n]['tocpm'].update({ m: {'x': [],
                                                                'xupper': [[0 for t in range(self.Time)]
                                                                          for j in range(self.group)],
                                                                'xlower': [[0 for t in range(self.Time)]
                                                                           for j in range(self.group)],
                                                                'q_idp': pp,
                                                                'q_isdp': [],
                                                                'q_norm': {},
                                                                'q_dp': {},
                                                                'q_lb': {},
                                                                'q_ub': {} }})
            if n in self.compartment_info[m]['fromcpm'].keys():
                if opt == 'delete':
                    del self.compartment_info[m]['fromcpm'][n]
                else:
                    self.compartment_info[m]['fromcpm'][n]['q_idp'] = pp
            elif n not in self.compartment_info[m]['fromcpm'].keys():
                self.compartment_info[m]['fromcpm'].update({n: { 'x':[],
                                                              'xupper': [[0 for t in range(self.Time)]
                                                                          for j in range(self.group)],
                                                              'xlower': [[0 for t in range(self.Time)]
                                                                           for j in range(self.group)],
                                                              'q_idp': pp,
                                                              'q_isdp': [],
                                                              'q_norm': {},
                                                              'q_dp': {},
                                                              'q_lb': {},
                                                              'q_ub': {}}})
        else:
            raise ValueError('Compartment is not defined')

    def set_transition_compartment_dp(self, flow=None, beta=None, gamma=None,
                                        alpha=None, betadeno=None, gammadeno=None, alphadeno=1, lb=-1e+20, ub=1e+20, opt=None):
        '''Define the coefficient of decision-dependent transition probability
        beta^{m'}_{k,j,t} from compartment n to m'''
        '''beta : m' * k'''
        if opt == None:
            if isinstance(flow[0], str):
                n = get_keys(self.compartment_name, flow[0])[0]
                flow = (n, flow[1], flow[2],flow[3])
            if isinstance(flow[1], str):
                m = get_keys(self.compartment_name, flow[1])[0]
                flow = (flow[0],m,flow[2],flow[3])
            valuenorm = [[0 for j in range(self.group)] for n in range(self.compartment)]
            if beta == None:
                beta = valuenorm
            if gamma == None:
                gamma = valuenorm
            if alpha == None:
                alpha = 0
            if betadeno == None:
                betadeno = valuenorm
            if gammadeno == None:
                gammadeno = valuenorm
            if alphadeno == None:
                alphadeno = 1
            self.set_transition_compartment(n=flow[0], m=flow[1],
                                            prob=[[0 for t in range(self.Time)] for j in range(self.group)], opt = None)
            self.compartment_info[n]['tocpm'][m]['q_ub'][(flow[2],flow[3])] = ub
            self.compartment_info[n]['tocpm'][m]['q_lb'][(flow[2],flow[3])] = lb
            self.compartment_info[m]['fromcpm'][n]['q_lb'][(flow[2],flow[3])] = lb
            self.compartment_info[m]['fromcpm'][n]['q_ub'][(flow[2],flow[3])] = ub
            self.set_transition_compartment_norm(flow = flow, betabar= valuenorm, gammabar= betadeno, phibar= valuenorm,
                                                 psibar= gammadeno, alphabar=alphadeno, opt = opt)
            self.set_transition_compartment_beta(flow=flow, beta = valuenorm, gamma= beta, phi= valuenorm, psi= gamma,
                                            alpha=alpha, opt = opt)
        elif opt == 'delete':
            if isinstance(flow[0], str):
                n = get_keys(self.compartment_name, flow[0])[0]
                flow = (n, flow[1])
            if isinstance(flow[1], str):
                m = get_keys(self.compartment_name, flow[1])[0]
                flow = (flow[0], m)
            self.set_transition_compartment_norm(flow=flow,opt=opt)
            self.set_transition_compartment_beta(flow=flow, opt=opt)
            self.compartment_info[flow[0]]['tocpm'][flow[1]]['q_ub'] = {}
            self.compartment_info[flow[0]]['tocpm'][flow[1]]['q_lb'] = {}
            self.compartment_info[flow[1]]['fromcpm'][flow[0]]['q_ub'] = {}
            self.compartment_info[flow[1]]['fromcpm'][flow[0]]['q_lb'] = {}

    def set_transition_compartment_self(self):
        for m in self.compartment_name.keys():
            prob = [[1 for t in range(self.Time)] for j in range(self.group)]
            term = list(self.compartment_info[m]['tocpm'].keys())
            try:
                term.remove(m)
            except:
                pass
            label = [[0 for t in range(self.Time)] for j in range(self.group)]
            for j in range(self.group):
                for t in range(self.Time):
                    betabar = []
                    gammabar = []
                    phibar = []
                    psibar = []
                    alphabar = []
                    beta0 = []
                    gamma0 = []
                    phi0 = []
                    psi0 = []
                    alpha0 = []
                    for n in term:
                        if (j, t) not in self.compartment_info[m]['tocpm'][n]['q_norm'].keys():
                            prob[j][t] -= self.compartment_info[m]['tocpm'][n]['q_idp'][j][t]
                        else:
                            label[j][t] = 1
                            betabar.append(self.compartment_info[m]['tocpm'][n]['q_norm'][(j, t)][0])
                            gammabar.append(self.compartment_info[m]['tocpm'][n]['q_norm'][(j, t)][1])
                            phibar.append(self.compartment_info[m]['tocpm'][n]['q_norm'][(j, t)][2])
                            psibar.append(self.compartment_info[m]['tocpm'][n]['q_norm'][(j, t)][3])
                            alphabar.append(self.compartment_info[m]['tocpm'][n]['q_norm'][(j, t)][4])
                            beta0.append(self.compartment_info[m]['tocpm'][n]['q_dp'][(j, t)][0])
                            gamma0.append(self.compartment_info[m]['tocpm'][n]['q_dp'][(j, t)][1])
                            phi0.append(self.compartment_info[m]['tocpm'][n]['q_dp'][(j, t)][2])
                            psi0.append(self.compartment_info[m]['tocpm'][n]['q_dp'][(j, t)][3])
                            alpha0.append(self.compartment_info[m]['tocpm'][n]['q_dp'][(j, t)][4])
                    if label[j][t] == 1:
                        Beta0 = [[sum([beta0[i][mm][k] for i in range(len(beta0))]) for k in range(self.group)]
                                 for mm in range(max(self.compartment_name.keys()) + 1)]
                        Gamma0 = [[sum([gamma0[i][mm][k] for i in range(len(gamma0))]) for k in range(self.group)]
                                  for mm in range(max(self.compartment_name.keys()) + 1)]
                        Phi0 = [[sum([phi0[i][mm][k] for i in range(len(phi0))]) for k in range(self.group)]
                                for mm in range(max(self.compartment_name.keys()) + 1)]
                        Psi0 = [[sum([psi0[i][mm][k] for i in range(len(psi0))]) for k in range(self.group)]
                                for mm in range(max(self.compartment_name.keys()) + 1)]
                        Alpha0 = sum(alpha0)

                        beta3 = [[betabar[0][mm][k] * prob[j][t] - Beta0[mm][k] for k in range(self.group)] for mm in
                                 range(max(self.compartment_name.keys()) + 1)]
                        gamma3 = [[gammabar[0][mm][k] * prob[j][t] - Gamma0[mm][k] for k in range(self.group)] for mm in
                                  range(max(self.compartment_name.keys()) + 1)]
                        phi3 = [[phibar[0][mm][k] * prob[j][t] - Phi0[mm][k] for k in range(self.group)] for mm in
                                range(max(self.compartment_name.keys()) + 1)]
                        psi3 = [[psibar[0][mm][k] * prob[j][t] - Psi0[mm][k] for k in range(self.group)] for mm in
                                range(max(self.compartment_name.keys()) + 1)]
                        alpha3 = alphabar[0] * prob[j][t] - Alpha0

                        self.set_transition_compartment(n=m, m=m,
                                                        prob=[[0 for tt in range(self.Time)] for jj in range(self.group)],
                                                        opt=None)
                        self.set_transition_compartment_norm(flow=(m, m, j, t),
                                                             betabar=betabar[0], gammabar=gammabar[0], phibar=phibar[0],
                                                             psibar=psibar[0], alphabar=alphabar[0], opt=None)
                        self.set_transition_compartment_beta(flow=(m, m, j, t),
                                                             beta=beta3, gamma=gamma3, phi=phi3, psi=psi3,
                                                             alpha=alpha3, opt=None)
                        # print(self.compartment_info[m]['tocpm'][m])
                    else:
                        self.set_transition_compartment(n=m, m=m, prob=prob, opt=None)
                        # print(n, self.compartment_info[n]['fromcpm'].keys())

    def set_transition_compartment_norm(self, flow = None, betabar= None, gammabar= None, phibar= None, psibar= None,
                                        alphabar=None, opt = None):
        '''define the coefficient of N of decision-dependent transition probability \bar{beta}^{m'}_{k,j,t}
        in step function'''
        if len(flow) == 4:
            (n, m, j, t) = flow
            temp = self.compartment_name.keys()
            if n in temp and m in temp:
                # print(self.compartment_info[n]['tocpm'][m]['q_norm'])
                if opt == None:
                    if (j,t) in self.compartment_info[n]['tocpm'][m]['q_norm'].keys():
                        self.compartment_info[n]['tocpm'][m]['q_norm'][(j,t)] \
                            = [betabar, gammabar, phibar, psibar, alphabar]
                        self.compartment_info[m]['fromcpm'][n]['q_norm'][(j,t)] \
                            = [betabar, gammabar, phibar, psibar, alphabar]
                    else:
                        self.compartment_info[n]['tocpm'][m]['q_norm'].update(
                            {(j, t): [betabar, gammabar, phibar, psibar, alphabar]})
                        self.compartment_info[m]['fromcpm'][n]['q_norm'].update(
                            {(j, t): [betabar, gammabar, phibar, psibar, alphabar]})
                elif opt == 'delete':
                    del self.compartment_info[n]['tocpm'][m]['q_norm'][(j,t)]
                    del self.compartment_info[m]['fromcpm'][n]['q_norm'][(j,t)]
            else:
                raise  ValueError('Compartment is not defined')
        elif len(flow) == 2:
            (n, m) = flow
            if opt == 'delete':
                self.compartment_info[n]['tocpm'][m]['q_norm'] = {}
                self.compartment_info[m]['fromcpm'][n]['q_norm'] = {}
        else:
            raise ValueError('Domain of flow is incorrect')

    def set_transition_compartment_beta(self, flow=None, beta = None, gamma= None, phi= None, psi= None,
                                        alpha=None, opt = None):
        '''Define the coefficient of decision-dependent transition probability
        beta^{m'}_{k,j,t} from compartment n to m'''
        '''beta : m' * k'''
        temp = self.compartment_name.keys()
        if len(flow) == 4:
            (n, m, j, t) = flow
            if n in temp and m in temp:
                if opt == None:
                    if (j,t) in self.compartment_info[n]['tocpm'][m]['q_dp'].keys():
                        self.compartment_info[n]['tocpm'][m]['q_dp'][(j,t)] = [beta, gamma, phi, psi, alpha]
                        self.compartment_info[m]['fromcpm'][n]['q_dp'][(j,t)] = [beta, gamma, phi, psi, alpha]
                    else:
                        self.compartment_info[n]['tocpm'][m]['q_dp'].update({(j,t): [beta, gamma, phi, psi, alpha]})
                        self.compartment_info[m]['fromcpm'][n]['q_dp'].update({(j,t): [beta, gamma, phi, psi, alpha]})
                elif opt == 'delete':
                    del self.compartment_info[n]['tocpm'][m]['q_dp'][(j,t)]
                    del self.compartment_info[m]['fromcpm'][n]['q_dp'][(j,t)]
            else:
                raise  ValueError('Compartment is not defined')
        elif len(flow) == 2:
            (n, m) = flow
            if opt == 'delete':
                self.compartment_info[n]['tocpm'][m]['q_dp'] = {}
                self.compartment_info[m]['fromcpm'][n]['q_dp'] = {}
        else:
            raise ValueError('Domain of flow is incorrect')


    '''=============== Step 3.2. Define transition between group and group  ====================='''

    def set_transition_group(self, compartment=None, prob=None):
        '''Define the decision-independent in-compartment transition probability'''
        '''compartment should be an integer'''
        '''For each compartment n, prob = p[j][k] or p[j][k][t]'''
        if isinstance(compartment, str):
            compartment = get_keys(self.compartment_name, compartment)[0]
        if compartment in self.compartment_info.keys():
            n = compartment
            temp = numpy.array(prob).shape
            if len(temp) == 2 and temp == (self.group, self.group):
                self.compartment_info[n]['local']['p_idp'] = [[[prob[j][k] for t in range(self.Time)]
                                                               for k in range(self.group)]
                                                              for j in range(self.group)]
            elif len(temp) == 3 and temp == (self.group, self.group, self.Time):
                self.compartment_info[n]['local']['p_idp'] = [[[prob[j][k][t] for t in range(self.Time)]
                                                               for k in range(self.group)]
                                                              for j in range(self.group)]
            else:
                ValueError('Domain of parameter (prob) is incorrect')
        else:
            raise ValueError('Out of range')


    def set_transition_group_affine(self, state=None, phatX=None, phatx= None, opt= None):
        '''Define the coefficients of the decision-dependent in-compartment transition probability'''
        '''state should be a tuple (compartment id, group id, group id, time)'''
        '''For each compartment n, phaX = p[n'][j'][tau], phax = p[n'][n''][j'][tau]'''
        if len(state) == 4:
            (n,j,k,t) = state
            if n in self.compartment_info.keys():
                if opt == None:
                    if (j,k,t) in self.compartment_info[n]['local']['p_dp'].keys():
                        self.compartment_info[n]['local']['p_dp'][(j,k,t)] = [phatX, phatx]
                    else:
                        self.compartment_info[n]['local']['p_dp'].update({(j,k,t): [phatX, phatx]})
                elif opt == 'delete':
                    del self.compartment_info[n]['local']['p_dp'][(j,k,t)]
            else:
                raise ValueError('Compartment is not defined')
        elif isinstance(state, int) and state in self.compartment_name.keys():
            if opt == 'delete':
                self.compartment_info[state]['local']['p_dp'] = {}
        else:
            raise ValueError('Domain of state is incorrect')

    '''=============== Step 4. Define decision variable  ====================='''
    def set_flow_variable(self, n=None, m=None, xupper=None, xlower=None):
        '''Define the range of flow decision variables'''
        '''xupper is xupper[j][t] or xupper[j] or xupper[t] or None, and similar to xlower'''
        if isinstance(n, str):
            n = get_keys(self.compartment_name, n)[0]
        elif n == None:
            raise ValueError('Please input n')
        if isinstance(m, str):
            m = get_keys(self.compartment_name, m)[0]
        elif m == None:
            raise ValueError('Please input m')
        self.set_transition_compartment(n=n, m=m, prob=[[0 for t in range(self.Time)] for j in range(self.group)], opt = None)
        temp = self.compartment_name.keys()
        if n in temp and m in temp:
            if xupper == None:
                xu = [[1e+20 for t in range(self.Time)] for j in range(self.group)]
            elif isinstance(xupper, (int, float)):
                xu = [[xupper for t in range(self.Time)] for j in range(self.group)]
            elif isinstance(xupper,list):
                temp1 = numpy.array(xupper).shape
                if len(temp1) == 2 and temp1[0] == self.group and temp1[1] >= self.Time:
                    xu = [[xupper[j][t] for t in range(temp1[1])] for j in range(temp1[0])]
                elif len(temp1) == 1 and temp1 == self.group:
                    xu = [[xupper[j] for t in range(self.Time)] for j in range(self.group)]
                elif len(temp1) == 1 and temp1 == self.Time:
                    xu = [[xupper[t] for t in range(self.Time)] for j in range(self.group)]
                else:
                    raise ValueError('Domain of xupper is incorrect')
            else:
                raise ValueError('Domain of xupper is incorrect')

            if xlower == None:
                xl = [[0 for t in range(self.Time)] for j in range(self.group)]
            elif isinstance(xlower, (int, float)):
                xl = [[xlower for t in range(self.Time)] for j in range(self.group)]
            elif isinstance(xlower, list):
                temp1 = numpy.array(xlower).shape
                if len(temp1) == 2 and temp1[0] == self.group and temp1[1] >= self.Time:
                    xl = [[xlower[j][t] for t in range(temp1[1])] for j in range(self.group)]
                elif len(temp1) == 1 and temp1 == self.group:
                    xl = [[xlower[j] for t in range(self.Time)] for j in range(self.group)]
                elif len(temp1) == 1 and temp1 == self.Time:
                    xl = [[xlower[t] for t in range(self.Time)] for j in range(self.group)]
                else:
                    raise ValueError('Domain of xupper is incorrect')
            else:
                raise ValueError('Domain of xupper is incorrect')
            self.compartment_info[n]['tocpm'][m]['xupper'] = xu
            self.compartment_info[m]['fromcpm'][n]['xupper'] = xu
            self.compartment_info[n]['tocpm'][m]['xlower'] = xl
            self.compartment_info[m]['fromcpm'][n]['xlower'] = xl
            for j in range(self.group):
                for t in range(self.Time):
                    if xu[j][t] >= xl[j][t]:
                        self.compartment_info[n]['tocpm'][m]['x'].append((j,t))
                        self.compartment_info[m]['fromcpm'][n]['x'].append((j,t))
        else:
            raise ValueError('Compartment is not defined')

    def set_custom_variable(self, name=None, xupper=None, xlower=None, types='C'):
        '''Define the range of flow decision variables'''
        '''xupper is xupper[j][t] or xupper[j] or xupper[t] or None, and similar to xlower'''
        if isinstance(name,str):
            name = [name]
            if isinstance(xupper, (int, float)):
                xupper = [xupper]
            elif xupper == None:
                xupper = [1e+20]
            if isinstance(xlower, (int, float)):
                xlower = [xlower]
            elif xlower == None:
                xlower = [0]
            if types in ['C','B','I']:
                types = [types]
            else:
                raise ValueError('types should be C, B or I')
        else:
            raise ValueError('name should be a string')

        # print(name, xupper, xlower)
        if len(name) == len(xupper) == len(xlower):
            for i in range(len(xupper)):
                if xlower[i] == None:
                    xlower[i] = 0
                if xupper[i] == None:
                    xupper[i] = 1e+20
                if types[i] == None:
                    types[i] = 'C'
                self.custom_var.update({name[i]: [xlower[i], xupper[i], types[i]]})
        else:
            raise ValueError('Dimension is inconsistent')

    '''=============== Step 5. Define initial population  ====================='''
    def set_population(self, compartment=None, population=None):
        '''Define population of a compartment'''
        '''population[j]'''
        if isinstance(compartment, str):
            compartment = get_keys(self.compartment_name, compartment)[0]
        temp = self.compartment_name.keys()
        if compartment in temp and len(population)==self.group:
            self.compartment_info[compartment]['local']['population'] = population
        else:
            raise ValueError('Domain of compartment is incorrect or compartment is not defined')

    def __set_dvar_bound(self):
        xupper = [[[[min(100000, self.compartment_info[n]['tocpm'][m]['xupper'][j][t])
                     if m in self.compartment_info[n]['tocpm'].keys()
                        and (j, t) in self.compartment_info[n]['tocpm'][m]['x']
                     else 100000 for t in range(self.Time)] for j in range(self.group)]
                   for m in range(max(self.compartment_name.keys()) + 1)]
                  for n in range(max(self.compartment_name.keys()) + 1)]
        xlower = [[[[max(0, self.compartment_info[n]['tocpm'][m]['xlower'][j][t])
                     if m in self.compartment_info[n]['tocpm'].keys() and (j, t) in
                                                                          self.compartment_info[n]['tocpm'][m]['x']
                     else 0 for t in range(self.Time)] for j in range(self.group)]
                   for m in range(max(self.compartment_name.keys()) + 1)] for n in
                  range(max(self.compartment_name.keys()) + 1)]
        for i in self.constraint.keys():
            dvar = self.constraint[i][0]  # var
            dcoef = self.constraint[i][1]  # coef
            sense = self.constraint[i][2]
            rhs = self.constraint[i][3]
            label = [dvar[i][0] for i in range(len(dvar))]
            for j in range(len(label)):
                if label[j] == 'x':
                    temp = dvar[j].split('.')[1:]
                    temp = [int(temp[ii]) for ii in range(len(temp))]
                    if ((sense in ['L', 'E'] and dcoef[j] > 0) or (sense in ['G', 'E'] and dcoef[j] < 0)) and \
                                            0 < rhs / dcoef[j] < xupper[temp[0]][temp[1]][temp[2]][temp[3]]:
                        xupper[temp[0]][temp[1]][temp[2]][temp[3]] = rhs / dcoef[j]

        xupper = [[[[0 if xupper[n][m][j][t] == 100000 else xupper[n][m][j][t] for t in range(self.Time)]
                    for j in range(self.group)] for m in range(max(self.compartment_name.keys()) + 1)]
                  for n in range(max(self.compartment_name.keys()) + 1)]
        return xupper, xlower

    def set_step(self):
        xupper, xlower = self.__set_dvar_bound()
        self.set_x(x=xupper)
        X1, p1, q1 = self.DeterPrediction(types='stair')
        self.set_x(x=xlower)
        X2, p2, q2 = self.DeterPrediction(types='stair')

        pgap = [[[[fabs(p2[n][j][k][t] - p1[n][j][k][t]) for t in range(self.Time)] for k in range(self.group)]
                 for j in range(self.group)] for n in range(max(self.compartment_name.keys()) + 1)]

        qgap = [[[[fabs(q2[n][m][j][t] - q1[n][m][j][t]) for t in range(self.Time)] for j in range(self.group)]
                 for m in range(max(self.compartment_name.keys()) + 1)] for n in
                range(max(self.compartment_name.keys()) + 1)]

        Lp = [[[[int(round(pgap[n][j][k][t] / self.gap, 1) + 1) for t in range(self.Time)] for k in range(self.group)]
               for j in range(self.group)] for n in range(max(self.compartment_name.keys()) + 1)]

        Lq = [[[[int(round(qgap[n][m][j][t] / self.gap, 1) + 1) for t in range(self.Time)] for j in range(self.group)]
               for m in range(max(self.compartment_name.keys()) + 1)] for n in
              range(max(self.compartment_name.keys()) + 1)]

        self.set_L(L=Lp, ub=p2, lb=p1, label='p')
        self.set_L(L=Lq, ub=q2, lb=q1, label='q')
        
    def set_L(self, L=None, ub= None, lb=None, label=None):
        if label == 'p':
            setattr(self, 'Lp', L)
            Ub = [[[[max(ub[n][j][k][t], lb[n][j][k][t]) for t in range(self.Time)] for k in range(self.group)]
                   for j in range(self.group)] for n in range(max(self.compartment_name.keys())+1)]
            Lb = [[[[min(ub[n][j][k][t], lb[n][j][k][t]) for t in range(self.Time)] for k in range(self.group)]
                   for j in range(self.group)] for n in range(max(self.compartment_name.keys()) + 1)]
            setattr(self, 'uover', [[[[[round(Lb[n][j][k][t] + (l-1)*self.gap,6) if j!=k
                                        else round(Lb[n][j][k][t]+(l+0)*self.gap,6)
                                        for l in range(self.Lp[n][j][k][t]+1)]
                                       for t in range(self.Time)] for k in range(self.group)] for j in range(self.group)]
                                    for n in range(max(self.compartment_name.keys())+1)])
            setattr(self, 'uunder', [[[[[round(Lb[n][j][k][t] + (l+0) * self.gap, 6) if j != k
                                         else round(Lb[n][j][k][t]+(l+0)*self.gap, 6)
                                         for l in range(self.Lp[n][j][k][t])] for t in range(self.Time)]
                                      for k in range(self.group)] for j in range(self.group)]
                                    for n in range(max(self.compartment_name.keys()) + 1)])

        elif label == 'q':
            setattr(self, 'Lq', L)
            Ub = [[[[max(ub[n][m][j][t], lb[n][m][j][t]) for t in range(self.Time)] for j in range(self.group)]
                   for m in range(max(self.compartment_name.keys()) + 1)] for n in range(max(self.compartment_name.keys()) + 1)]
            Lb = [[[[min(ub[n][m][j][t], lb[n][m][j][t]) for t in range(self.Time)] for j in range(self.group)]
                   for m in range(max(self.compartment_name.keys()) + 1)] for n in range(max(self.compartment_name.keys()) + 1)]
            setattr(self, 'vover', [[[[[round(Lb[n][m][k][t] + (l-1) * self.gap, 6) if n != m
                                        else round(Lb[n][m][k][t] + (l+0) * self.gap, 6)
                                        for l in range(self.Lq[n][m][k][t]+1)]
                                       for t in range(self.Time)] for k in range(self.group)]
                                     for m in range(max(self.compartment_name.keys()) + 1)]
                                    for n in range(max(self.compartment_name.keys())+1)])
            setattr(self, 'vunder', [[[[[round(Lb[n][m][k][t] + (l+0) * self.gap, 6) if n != m
                                         else round(Lb[n][m][k][t] + (l+0) * self.gap, 6)
                                         for l in range(self.Lq[n][m][k][t])] for t in range(self.Time)]
                                      for k in range(self.group)] for m in range(max(self.compartment_name.keys()) + 1)]
                                    for n in range(max(self.compartment_name.keys()) + 1)])
        else:
            raise ValueError('Label should be p or q')

    '''=============== Step 6. Define range of variables  ====================='''
    def dvar_conpartment(self):
        if self.solver == 'cplex':
            temp = self.compartment_name.keys()
            self.model.variables.add(obj=[0 for n in temp for j in range(self.group) for t in range(self.Time)],
                            lb=[0 if t >0 else self.compartment_info[n]['local']['population'][j]
                                for n in temp for j in range(self.group) for t in range(self.Time)],
                            ub=[cplex.infinity if t >0 else self.compartment_info[n]['local']['population'][j]
                                for n in temp for j in range(self.group) for t in range(self.Time)],
                            names=["X." + str(n) + "." + str(j) + "." + str(t) for n in temp
                                   for j in range(self.group) for t in range(self.Time)])

            self.model.variables.add(obj=[0 for n in temp for m in self.compartment_info[n]['tocpm'].keys()
                                          for j in range(self.group) for t in range(self.Time)],
                            lb=[0 for n in temp for m in self.compartment_info[n]['tocpm'].keys()
                                          for j in range(self.group) for t in range(self.Time)],
                            ub=[cplex.infinity for n in temp for m in self.compartment_info[n]['tocpm'].keys()
                                          for j in range(self.group) for t in range(self.Time)],
                            names=["U." + str(n) + "." + str(m) + "." + str(j) + "." + str(t)
                                   for n in temp for m in self.compartment_info[n]['tocpm'].keys()
                                   for j in range(self.group) for t in range(self.Time)])

            self.model.variables.add(obj=[0 for n in temp for j in range(self.group) for t in range(self.Time)],
                                     lb=[0 for n in temp for j in range(self.group) for t in range(self.Time)],
                                     ub=[cplex.infinity for n in temp for j in range(self.group) for t in range(self.Time)],
                                     names=["W." + str(n) + "." + str(j) + "." + str(t)
                                            for n in temp for j in range(self.group) for t in range(self.Time)])

            if self.__approximation == 'ME':
                self.model.variables.set_lower_bounds([("W." + str(n) + "." + str(j) + "." + str(t),
                                                        self.Wlower[n][j][t]) for n in temp
                                                       for j in range(self.group) for t in range(self.Time)])
                self.model.variables.set_upper_bounds([("W." + str(n) + "." + str(j) + "." + str(t),
                                                        self.Wupper[n][j][t]) for n in temp
                                                       for j in range(self.group) for t in range(self.Time)])

            self.model.variables.add(obj=[0 for n in temp for m in self.compartment_info[n]['tocpm'].keys()
                                            for (j,t) in self.compartment_info[n]['tocpm'][m]['x']],
                                     lb=[self.compartment_info[n]['tocpm'][m]['xlower'][j][t]
                                         for n in temp for m in self.compartment_info[n]['tocpm'].keys()
                                            for (j,t) in self.compartment_info[n]['tocpm'][m]['x']],
                                     ub=[self.compartment_info[n]['tocpm'][m]['xupper'][j][t]
                                         for n in temp for m in self.compartment_info[n]['tocpm'].keys()
                                            for (j,t) in self.compartment_info[n]['tocpm'][m]['x']],
                                     names=["x." + str(n) + "." + str(m) + "." + str(j) + "." + str(t)
                                            for n in temp for m in self.compartment_info[n]['tocpm'].keys()
                                            for (j,t) in self.compartment_info[n]['tocpm'][m]['x']])

            if self.__approximation == 'SF':
                self.model.variables.add(obj=[0 for n in temp for k in range(self.group)
                                                for j in range(self.group) for t in range(self.Time)],
                                         lb=[0 for n in temp for k in range(self.group)
                                                for j in range(self.group) for t in range(self.Time)],
                                         ub=[cplex.infinity for n in temp for k in range(self.group)
                                                for j in range(self.group) for t in range(self.Time)],
                                         names=["V." + str(n) + "." + str(k) + "." + str(j) + "." + str(t)
                                                for n in temp for k in range(self.group)
                                                for j in range(self.group) for t in range(self.Time)])

                self.model.variables.add(obj=[0 for n in temp for k in range(self.group)
                                                for j in range(self.group) for t in range(self.Time)],
                                         lb=[0 for n in temp for k in range(self.group)
                                                for j in range(self.group) for t in range(self.Time)],
                                         ub=[cplex.infinity for n in temp for k in range(self.group)
                                                for j in range(self.group) for t in range(self.Time)],
                                         names=["Vhat." + str(n) + "." + str(k) + "." + str(j) + "." + str(t)
                                                for n in temp for k in range(self.group)
                                                for j in range(self.group) for t in range(self.Time)])

            self.model.variables.add(obj=[0 for i in self.custom_var.keys()],
                                     lb=[self.custom_var[i][0] for i in self.custom_var.keys()],
                                     ub=[self.custom_var[i][1] for i in self.custom_var.keys()],
                                     types = [self.custom_var[i][2] for i in self.custom_var.keys()],
                                     names=[i for i in self.custom_var.keys()])

            if self.types == 'Robustness':
                temp = self.compartment_name.keys()
                self.model.variables.add(obj=[0 for n in temp for j in range(self.group) for t in range(self.Time)],
                                         lb=[0 for n in temp for j in range(self.group) for t in range(self.Time)],
                                         ub=[cplex.infinity if t > 0 else 0
                                             for n in temp for j in range(self.group) for t in range(self.Time)],
                                         names=["Xdagger." + str(n) + "." + str(j) + "." + str(t) for n in temp
                                                for j in range(self.group) for t in range(self.Time)])

                self.model.variables.add(obj=[0 for n in temp for m in self.compartment_info[n]['tocpm'].keys()
                                              for j in range(self.group) for t in range(self.Time)],
                                         lb=[-cplex.infinity for n in temp for m in self.compartment_info[n]['tocpm'].keys()
                                             for j in range(self.group) for t in range(self.Time)],
                                         ub=[cplex.infinity for n in temp for m in
                                             self.compartment_info[n]['tocpm'].keys()
                                             for j in range(self.group) for t in range(self.Time)],
                                         names=["Uhat." + str(n) + "." + str(m) + "." + str(j) + "." + str(t)
                                                for n in temp for m in self.compartment_info[n]['tocpm'].keys()
                                                for j in range(self.group) for t in range(self.Time)])

                self.model.variables.add(obj=[0 for n in temp for j in range(self.group) for t in range(self.Time)],
                                         lb=[-cplex.infinity for n in temp for j in range(self.group)
                                             for t in range(self.Time)],
                                         ub=[cplex.infinity for n in temp for j in range(self.group)
                                             for t in range(self.Time)],
                                         names=["What." + str(n) + "." + str(j) + "." + str(t)
                                                for n in temp for j in range(self.group) for t in range(self.Time)])

                if self.__approximation == 'ME':
                    self.model.variables.set_lower_bounds([("What." + str(n) + "." + str(j) + "." + str(t),
                                                            self.whatlower[n][j][t]) for n in temp
                                                           for j in range(self.group) for t in range(self.Time)])
                    self.model.variables.set_upper_bounds([("What." + str(n) + "." + str(j) + "." + str(t),
                                                            self.whatupper[n][j][t]) for n in temp
                                                           for j in range(self.group) for t in range(self.Time)])

                    self.model.variables.add(obj=[0 for n in temp for m in self.compartment_info[n]['tocpm'].keys()
                                                  for j in range(self.group) for t in range(self.Time)],
                                             lb=[self.eta1lower[n][m][j][t]**2 for n in temp
                                                 for m in self.compartment_info[n]['tocpm'].keys()
                                                 for j in range(self.group) for t in range(self.Time)],
                                             ub=[self.eta1upper[n][m][j][t]**2 for n in temp
                                                 for m in self.compartment_info[n]['tocpm'].keys()
                                                 for j in range(self.group) for t in range(self.Time)],
                                             names=["phi." + str(n) + "." + str(m) + "." + str(j) + "." + str(t)
                                                    for n in temp for m in self.compartment_info[n]['tocpm'].keys()
                                                    for j in range(self.group) for t in range(self.Time)])
            elif self.types != 'Expectation' and self.types != 'Robustness':
                raise ValueError('Incorrect optimization method\n '
                                 'Please either input type = Expectation, or type = Robustness')

        elif self.solver == 'gurobi':
            temp = self.compartment_name.keys()
            for n in temp:
                for j in range(self.group):
                    for t in range(self.Time):
                        if t > 0:
                            self.model.addVar(
                                lb=0, ub=GRB.INFINITY, vtype=GRB.CONTINUOUS,
                                name="X." + str(n) + "." + str(j) + "." + str(t))
                        else:
                            self.model.addVar(
                                lb=self.compartment_info[n]['local']['population'][j],
                                ub=self.compartment_info[n]['local']['population'][j],
                                vtype=GRB.CONTINUOUS,
                                name="X." + str(n) + "." + str(j) + "." + str(t))
            for n in temp:
                for m in self.compartment_info[n]['tocpm'].keys():
                    for j in range(self.group):
                        for t in range(self.Time):
                            self.model.addVar(lb=0, ub=GRB.INFINITY, vtype=GRB.CONTINUOUS,
                                name="U." + str(n) + "." + str(m) + "." + str(j) + "." + str(t))

            for n in temp:
                for j in range(self.group):
                    for t in range(self.Time):
                        self.model.addVar(
                            lb=0, ub=GRB.INFINITY, vtype=GRB.CONTINUOUS,
                            name="W." + str(n) + "." + str(j) + "." + str(t))
                        if self.__approximation == 'ME':
                            self.model.update()
                            self.model.setAttr("lb", self.model.getVarByName("W." + str(n) + "." + str(j) + "." + str(t)),
                                               self.Wlower[n][j][t])
                            self.model.setAttr("ub", self.model.getVarByName("W." + str(n) + "." + str(j) + "." + str(t)),
                                               self.Wupper[n][j][t])
            for n in temp:
                for m in self.compartment_info[n]['tocpm'].keys():
                    for (j, t) in self.compartment_info[n]['tocpm'][m]['x']:
                        self.model.addVar(
                            lb=self.compartment_info[n]['tocpm'][m]['xlower'][j][t],
                            ub=self.compartment_info[n]['tocpm'][m]['xupper'][j][t],
                            vtype=GRB.CONTINUOUS,
                            name="x." + str(n) + "." + str(m) + "." + str(j) + "." + str(t))
            for n in temp:
                if self.__approximation == 'SF':
                    for k in range(self.group):
                        for j in range(self.group):
                            for t in range(self.Time):
                                self.model.addVar(
                                    lb=0,ub=GRB.INFINITY, vtype=GRB.CONTINUOUS,
                                    name="V." + str(n) + "." + str(k) + "." + str(j) + "." + str(t))
                                self.model.addVar(
                                    lb=0, ub=GRB.INFINITY, vtype=GRB.CONTINUOUS,
                                    name="Vhat." + str(n) + "." + str(k) + "." + str(j) + "." + str(t))

            for i in self.custom_var.keys():
                self.model.addVar( lb=self.custom_var[i][0], ub=self.custom_var[i][1], vtype=self.custom_var[i][2], name=i)

            if self.types == 'Robustness':
                temp = self.compartment_name.keys()
                for n in temp:
                    for j in range(self.group):
                        for t in range(self.Time):
                            if t > 0:
                                self.model.addVar(lb=0, ub=GRB.INFINITY, vtype=GRB.CONTINUOUS,
                                    name="Xdagger." + str(n) + "." + str(j) + "." + str(t))
                            else:
                                self.model.addVar(lb=0, ub=0, vtype=GRB.CONTINUOUS,
                                    name="Xdagger." + str(n) + "." + str(j) + "." + str(t))

                for n in temp:
                    for m in self.compartment_info[n]['tocpm'].keys():
                        for j in range(self.group):
                            for t in range(self.Time):
                                self.model.addVar(lb=-GRB.INFINITY, ub=GRB.INFINITY, vtype=GRB.CONTINUOUS,
                                    name="Uhat." + str(n) + "." + str(m) + "." + str(j) + "." + str(t))

                for n in temp:
                    for j in range(self.group):
                        for t in range(self.Time):
                            self.model.addVar(
                                lb=-GRB.INFINITY, ub=GRB.INFINITY, vtype=GRB.CONTINUOUS,
                                name="What." + str(n) + "." + str(j) + "." + str(t))
                            if self.__approximation == 'ME':
                                self.model.update()
                                self.model.setAttr("lb", self.model.getVarByName("What." + str(n) + "." + str(j) + "." + str(t)),
                                                   self.whatlower[n][j][t])
                                self.model.setAttr("ub", [self.model.getVarByName("What." + str(n) + "." + str(j) + "." + str(t))],
                                                   self.whatupper[n][j][t])

                if self.__approximation == 'ME':
                    for n in temp:
                        for m in self.compartment_info[n]['tocpm'].keys():
                            for j in range(self.group):
                                for t in range(self.Time):
                                    self.model.addVar(lb=self.eta1lower[n][m][j][t] ** 2,
                                        ub=self.eta1upper[n][m][j][t] ** 2, vtype=GRB.CONTINUOUS,
                                        name="phi." + str(n) + "." + str(m) + "." + str(j) + "." + str(t))

            elif self.types != 'Expectation' and self.types != 'Robustness':
                raise ValueError('Incorrect optimization method\n '
                                 'Please either input type = Expectation, or type = Robustness')
        else:
            raise ValueError('Current version only supports Cplex and Gurobi as solvers')

    '''=============== Step 7. Define dynamics  ====================='''
    def __Dynamics(self):
        if self.solver == 'cplex':
            self.model.linear_constraints.add(lin_expr=[cplex.SparsePair(
                                               ["X." + str(n) + "." + str(j) + "." + str(t+1)]
                                             + ["x." + str(m) + "." + str(n) + "." + str(j) + "." + str(t)
                                                for m in self.compartment_info[n]['fromcpm'].keys()
                                                if (j,t) in self.compartment_info[n]['fromcpm'][m]['x']]
                                             + ["U." + str(m) + "." + str(n) + "." + str(j) + "." + str(t)
                                                for m in self.compartment_info[n]['fromcpm'].keys()],
                                             [-1]
                                             + [1 for m in self.compartment_info[n]['fromcpm'].keys()
                                                if (j,t) in self.compartment_info[n]['fromcpm'][m]['x'] ]
                                             + [1 for m in self.compartment_info[n]['fromcpm'].keys()])
                                                   for n in self.compartment_name.keys() for j in range(self.group)
                                                        for t in range(self.Time - 1)],
                                         senses=["E" for n in self.compartment_name.keys() for j in range(self.group)
                                                        for t in range(self.Time - 1)],
                                         rhs=[0 for n in self.compartment_name.keys() for j in range(self.group)
                                                        for t in range(self.Time - 1)],
                                         names=["dynamics." + str(n) + "." + str(j) + "." + str(t)
                                                for n in self.compartment_name.keys()
                                                for j in range(self.group) for t in range(self.Time - 1)])

            for n in self.compartment_name.keys():
                for m in self.compartment_info[n]['fromcpm'].keys():
                    for j in range(self.group):
                        for t in range(self.Time-1):
                            if (j,t) in self.compartment_info[n]['fromcpm'][m]['q_dp'].keys():
                                if self.__approximation == 'SF':
                                    self.model.linear_constraints.add(
                                        lin_expr=[cplex.SparsePair(["U." + str(m) + "." + str(n) + "." + str(j) + "." + str(t)]
                                                                   + ["W." + str(m) + "." + str(j) + "." + str(t)]
                                                                   + ["y." + str(m) + "." + str(n) + "." + str(j)
                                                                      + "." + str(t) + "." + str(l)],
                                                                   [-1]
                                                                   + [self.vunder[m][n][j][t][l]]
                                                                   + [self.__M])
                                                  for l in range(self.Lq[m][n][j][t])],
                                        senses=["L" for l in range(self.Lq[m][n][j][t])],
                                        rhs=[self.__M for l in range(self.Lq[m][n][j][t])],
                                        names=["dynamicsUlower." + str(m) + "." + str(n) + "." + str(j) + "." + str(t)
                                               + "." + str(l) for l in range(self.Lq[m][n][j][t])])

                                    self.model.linear_constraints.add(
                                        lin_expr=[cplex.SparsePair(["U." + str(m) + "." + str(n) + "." + str(j) + "." + str(t)]
                                                                   + ["W." + str(m) + "." + str(j) + "." + str(t)]
                                                                   + ["y." + str(m) + "." + str(n) + "." + str(j) + "." + str(t)
                                                                      + "." + str(l)],
                                                                   [-1]
                                                                   + [self.vunder[m][n][j][t][l]]
                                                                   + [-self.__M]) for l in range(self.Lq[m][n][j][t])],
                                        senses=["G" for l in range(self.Lq[m][n][j][t])],
                                        rhs=[-self.__M for l in range(self.Lq[m][n][j][t])],
                                        names=["dynamicsUupper." + str(m) + "." + str(n) + "." + str(j) + "." + str(t)
                                               + "." + str(l) for l in range(self.Lq[m][n][j][t])])
                                elif self.__approximation == 'ME':

                                    self.model.linear_constraints.add(
                                        lin_expr=[cplex.SparsePair(
                                                        ["U." + str(m) + "." + str(n) + "." + str(j) + "." + str(t),
                                                         "W." + str(m) + "." + str(j) + "." + str(t),
                                                         "eta." + str(m) + "." + str(n) + "." + str(j) + "." + str(t)],
                                                        [-1,self.eta1lower[m][n][j][t], self.Wlower[m][j][t]])],
                                        senses=["L"],
                                        rhs=[self.eta1lower[m][n][j][t]* self.Wlower[m][j][t]],
                                        names=["dynamicsUlower." + str(m) + "." + str(n) + "." + str(j) + "." + str(t)+ "." + str(0)])
                                    self.model.linear_constraints.add(
                                        lin_expr=[cplex.SparsePair(
                                                        ["U." + str(m) + "." + str(n) + "." + str(j) + "." + str(t),
                                                         "W." + str(m) + "." + str(j) + "." + str(t),
                                                         "eta." + str(m) + "." + str(n) + "." + str(j) + "." + str(t)],
                                                        [-1, self.eta1upper[m][n][j][t], self.Wupper[m][j][t]])],
                                        senses=["L"],
                                        rhs=[self.eta1upper[m][n][j][t] * self.Wupper[m][j][t]],
                                        names=["dynamicsUlower." + str(m) + "." + str(n) + "." + str(j) + "." + str(
                                            t) + "." + str(1)])
                                    self.model.linear_constraints.add(
                                        lin_expr=[cplex.SparsePair(
                                                        ["U." + str(m) + "." + str(n) + "." + str(j) + "." + str(t),
                                                         "W." + str(m) + "." + str(j) + "." + str(t),
                                                         "eta." + str(m) + "." + str(n) + "." + str(j) + "." + str(t)],
                                            [-1, self.eta1upper[m][n][j][t], self.Wlower[m][j][t]])],
                                        senses=["G"],
                                        rhs=[self.eta1upper[m][n][j][t] * self.Wlower[m][j][t]],
                                        names=["dynamicsUupper." + str(m) + "." + str(n) + "." + str(j) + "." + str(
                                            t) + "." + str(0)])
                                    self.model.linear_constraints.add(
                                        lin_expr=[cplex.SparsePair(
                                                        ["U." + str(m) + "." + str(n) + "." + str(j) + "." + str(t),
                                                         "W." + str(m) + "." + str(j) + "." + str(t),
                                                         "eta." + str(m) + "." + str(n) + "." + str(j) + "." + str(t)],
                                            [-1, self.eta1lower[m][n][j][t], self.Wupper[m][j][t]])],
                                        senses=["G"],
                                        rhs=[self.eta1lower[m][n][j][t] * self.Wupper[m][j][t]],
                                        names=["dynamicsUupper." + str(m) + "." + str(n) + "." + str(j) + "." + str(
                                            t) + "." + str(1)])
                            else:
                                self.model.linear_constraints.add(
                                    lin_expr=[cplex.SparsePair(["U." + str(m) + "." + str(n) + "." + str(j) + "." + str(t)]
                                                               + ["W." + str(m) + "." + str(j) + "." + str(t)],
                                                               [-1]
                                                               + [self.compartment_info[n]['fromcpm'][m]['q_idp'][j][t]]) ],
                                    senses=["E"],
                                    rhs=[0],
                                    names=["dynamicsUequal." + str(m) + "." + str(n) + "." + str(j) + "." + str(t)])
            for m in self.compartment_name.keys():
                for k in range(self.group):
                    for t in range(self.Time-1):
                        self.model.linear_constraints.add(
                        lin_expr=[cplex.SparsePair(["W." + str(m) + "." + str(k) + "." + str(t)]
                                                   + ["X." + str(m) + "." + str(k) + "." + str(t)]
                                                   + ["x." + str(m) + "." + str(mm) + "." + str(k) + "." + str(t)
                                                      for mm in self.compartment_info[m]['tocpm'].keys()
                                                      if (k,t) in self.compartment_info[m]['tocpm'][mm]['x'] ],
                                                   [-1]
                                                   + [self.compartment_info[m]['local']['p_idp'][k][k][t]]
                                                   + [-self.compartment_info[m]['local']['p_idp'][k][k][t]
                                                      for mm in self.compartment_info[m]['tocpm'].keys()
                                                      if (k,t) in self.compartment_info[m]['tocpm'][mm]['x'] ]
                                                   )],
                        senses=["E"],
                        rhs=[0],
                        names=["dynamicsWequal." + str(m) + "." + str(k) + "." + str(t)])

            if self.types == 'Robustness':
                self.model.linear_constraints.add(
                    lin_expr=[cplex.SparsePair(["Xdagger." + str(n) + "." + str(j) + "." + str(t + 1)]
                                               + ["U." + str(m) + "." + str(n) + "." + str(j) + "." + str(t)
                                                  for m in self.compartment_info[n]['fromcpm'].keys()]
                                               + ["Uhat." + str(m) + "." + str(n) + "." + str(j) + "." + str(t)
                                                  for m in self.compartment_info[n]['fromcpm'].keys()],
                                               [-1]
                                               + [1 for m in self.compartment_info[n]['fromcpm'].keys()]
                                               + [1 for m in self.compartment_info[n]['fromcpm'].keys()])
                              for n in self.compartment_name.keys() for j in range(self.group)
                              for t in range(self.Time - 1)],
                    senses=["E" for n in self.compartment_name.keys() for j in range(self.group)
                            for t in range(self.Time - 1)],
                    rhs=[0 for n in self.compartment_name.keys() for j in range(self.group)
                         for t in range(self.Time - 1)],
                    names=["Rdynamics." + str(n) + "." + str(j) + "." + str(t)
                           for n in self.compartment_name.keys()
                           for j in range(self.group) for t in range(self.Time - 1)])

                for n in self.compartment_name.keys():
                    for m in self.compartment_info[n]['fromcpm'].keys():
                        for j in range(self.group):
                            for t in range(self.Time-1):
                                if (j,t) in self.compartment_info[n]['fromcpm'][m]['q_dp'].keys():
                                    if self.__approximation == 'SF':
                                        self.model.linear_constraints.add(
                                            lin_expr=[cplex.SparsePair(["Uhat." + str(m) + "." + str(n) + "." + str(j)
                                                                    + "." + str(t)]
                                                                   + ["What." + str(m) + "." + str(j) + "." + str(t)]
                                                                   + ["y." + str(m) + "." + str(n) + "." + str(j)
                                                                      + "." + str(t) + "." + str(l)],
                                                                   [-1]
                                                                   + [self.vunder[m][n][j][t][l]**2]
                                                                   + [self.__M]) for l in range(self.Lq[m][n][j][t])],
                                            senses=["L" for l in range(self.Lq[m][n][j][t])],
                                            rhs=[self.__M for l in range(self.Lq[m][n][j][t])],
                                            names=["RdynamicsUlower." + str(m) + "." + str(n) + "." + str(j)
                                                   + "." + str(t) + "." + str(l) for l in range(self.Lq[m][n][j][t])])

                                        self.model.linear_constraints.add(
                                            lin_expr=[cplex.SparsePair(["Uhat." + str(m) + "." + str(n) + "." + str(j)
                                                                    + "." + str(t)]
                                                                   + ["What." + str(m) + "." + str(j) + "." + str(t)]
                                                                   + ["y." + str(m) + "." + str(n) + "." + str(j)
                                                                      + "." + str(t) + "." + str(l)],
                                                                   [-1]
                                                                   + [self.vunder[m][n][j][t][l]**2]
                                                                   + [-self.__M]) for l in range(self.Lq[m][n][j][t])],
                                            senses=["G" for l in range(self.Lq[m][n][j][t])],
                                            rhs=[-self.__M for l in range(self.Lq[m][n][j][t])],
                                            names=["RdynamicsUupper." + str(m) + "." + str(n) + "." + str(j)
                                                   + "." + str(t) + "." + str(l)for l in range(self.Lq[m][n][j][t])])
                                    elif self.__approximation == 'ME':
                                        self.model.linear_constraints.add(
                                            lin_expr=[cplex.SparsePair(
                                                ["Uhat." + str(m) + "." + str(n) + "." + str(j) + "." + str(t),
                                                 "What." + str(m) + "." + str(j) + "." + str(t),
                                                 "phi." + str(m) + "." + str(n) + "." + str(j) + "." + str(t)],
                                                [-1, self.eta1lower[m][n][j][t]**2, self.whatlower[m][j][t]])],
                                            senses=["L"],
                                            rhs=[self.eta1lower[m][n][j][t]**2 * self.whatlower[m][j][t]],
                                            names=["RdynamicsUlower." + str(m) + "." + str(n) + "." + str(j)
                                                   + "." + str(t) + "." + str(0)])

                                        self.model.linear_constraints.add(
                                            lin_expr=[cplex.SparsePair(
                                                ["Uhat." + str(m) + "." + str(n) + "." + str(j) + "." + str(t),
                                                 "What." + str(m) + "." + str(j) + "." + str(t),
                                                 "phi." + str(m) + "." + str(n) + "." + str(j) + "." + str(t)],
                                                [-1, self.eta1upper[m][n][j][t] ** 2, self.whatupper[m][j][t]])],
                                            senses=["L"],
                                            rhs=[self.eta1upper[m][n][j][t] ** 2 * self.whatupper[m][j][t]],
                                            names=[
                                                "RdynamicsUlower." + str(m) + "." + str(n) + "." + str(j)
                                                + "." + str(t) + "." + str(1)])

                                        self.model.linear_constraints.add(
                                            lin_expr=[cplex.SparsePair(
                                                ["Uhat." + str(m) + "." + str(n) + "." + str(j) + "." + str(t),
                                                 "What." + str(m) + "." + str(j) + "." + str(t),
                                                 "phi." + str(m) + "." + str(n) + "." + str(j) + "." + str(t)],
                                                [-1, self.eta1upper[m][n][j][t]**2, self.whatlower[m][j][t]])],
                                            senses=["G"],
                                            rhs=[self.eta1upper[m][n][j][t]**2 * self.whatlower[m][j][t]],
                                            names=["RdynamicsUupper." + str(m) + "." + str(n) + "." + str(j)
                                                   + "." + str(t) + "." + str(0)])

                                        self.model.linear_constraints.add(
                                            lin_expr=[cplex.SparsePair(
                                                ["Uhat." + str(m) + "." + str(n) + "." + str(j) + "." + str(t),
                                                 "What." + str(m) + "." + str(j) + "." + str(t),
                                                 "phi." + str(m) + "." + str(n) + "." + str(j) + "." + str(t)],
                                                [-1, self.eta1lower[m][n][j][t] ** 2, self.whatupper[m][j][t]])],
                                            senses=["G"],
                                            rhs=[self.eta1lower[m][n][j][t] ** 2 * self.whatupper[m][j][t]],
                                            names=["RdynamicsUupper." + str(m) + "." + str(n) + "." + str(j)
                                                   + "." + str(t) + "." + str(1)])

                                        self.model.linear_constraints.add(
                                            lin_expr=[cplex.SparsePair(
                                                ["phi." + str(m) + "." + str(n) + "." + str(j) + "." + str(t),
                                                 "eta." + str(m) + "." + str(n) + "." + str(j) + "." + str(t)],
                                                [-1, self.eta1upper[m][n][j][t]+self.eta1lower[m][n][j][t] ])],
                                            senses=["G"],
                                            rhs=[self.eta1upper[m][n][j][t] * self.eta1lower[m][n][j][t]],
                                            names=["Rdynamics_phi_upper." + str(m) + "." + str(n) + "." + str(j)
                                                   + "." + str(t)])

                                        self.model.linear_constraints.add(
                                            lin_expr=[cplex.SparsePair(
                                                ["phi." + str(m) + "." + str(n) + "." + str(j) + "." + str(t),
                                                 "eta." + str(m) + "." + str(n) + "." + str(j) + "." + str(t)],
                                                [-1, 2*self.eta1lower[m][n][j][t]])],
                                            senses=["L"],
                                            rhs=[self.eta1lower[m][n][j][t]**2],
                                            names=["Rdynamics_phi_lower." + str(m) + "." + str(n) + "." + str(j)
                                                   + "." + str(t) + "." + str(0)])

                                        self.model.linear_constraints.add(
                                            lin_expr=[cplex.SparsePair(
                                                ["phi." + str(m) + "." + str(n) + "." + str(j) + "." + str(t),
                                                 "eta." + str(m) + "." + str(n) + "." + str(j) + "." + str(t)],
                                                [-1, 2*self.eta1upper[m][n][j][t]])],
                                            senses=["L"],
                                            rhs=[self.eta1upper[m][n][j][t]**2],
                                            names=["Rdynamics_phi_lower." + str(m) + "." + str(n) + "." + str(j)
                                                   + "." + str(t) + "." + str(1)])
                                else:
                                    self.model.linear_constraints.add(
                                        lin_expr=[cplex.SparsePair(
                                            ["Uhat." + str(m) + "." + str(n) + "." + str(j) + "." + str(t)]
                                           + ["What." + str(m) + "." + str(j)+ "." + str(t)],
                                           [-1, self.compartment_info[n]['fromcpm'][m]['q_idp'][j][t]**2
                                              ])],
                                        senses=["E"],
                                        rhs=[0],
                                        names=["RdynamicsUequal." + str(m) + "." + str(n) + "." + str(j)
                                               + "." + str(t)])

                for m in self.compartment_name.keys():
                    for k in range(self.group):
                        for t in range(self.Time-1):
                            self.model.linear_constraints.add(
                                lin_expr=[
                                    cplex.SparsePair(["What." + str(m) + "." + str(k)  + "." + str(t)]
                                                     + ["Xdagger." + str(m) + "." + str(k) + "." + str(t),
                                                        "X." + str(m) + "." + str(k) + "." + str(t)]
                                                     + ["x." + str(m) + "." + str(mm) + "." + str(k) + "." + str(t)
                                                        for mm in self.compartment_info[m]['tocpm'].keys()
                                                      if (k,t) in self.compartment_info[m]['tocpm'][mm]['x']],
                                                     [-1]
                                                     + [self.compartment_info[m]['local']['p_idp'][k][k][t]**2,
                                                        -self.compartment_info[m]['local']['p_idp'][k][k][t]**2]
                                                     + [self.compartment_info[m]['local']['p_idp'][k][k][t]**2
                                                        for mm in self.compartment_info[m]['tocpm'].keys()
                                                      if (k,t) in self.compartment_info[m]['tocpm'][mm]['x']])],
                                senses=["E"],
                                rhs=[0],
                                names=["RdynamicsWequal." + str(m) + "." + str(k) + "." + str(t)])

            elif self.types != 'Robustness' and self.types != 'Expectation':
                raise ValueError('Incorrect optimization method\n '
                                 'Please either input type = Expectation, or type = Robustness')

        elif self.solver == 'gurobi':
            self.model.update()
            for n in self.compartment_name.keys():
                for j in range(self.group):
                    for t in range(self.Time - 1):
                        self.model.addConstr((-self.model.getVarByName("X." + str(n) + "." + str(j) + "." + str(t+1))
                                              + gp.quicksum(self.model.getVarByName("x." + str(m) + "." + str(n) + "." + str(j) + "." + str(t))
                                                                  for m in self.compartment_info[n]['fromcpm'].keys()
                                                                  if (j, t) in self.compartment_info[n]['fromcpm'][m]['x'])
                                              + gp.quicksum(self.model.getVarByName("U." + str(m) + "." + str(n) + "." + str(j) + "." + str(t))
                                                                  for m in self.compartment_info[n]['fromcpm'].keys()) == 0),
                                             "dynamics." + str(n) + "." + str(j) + "." + str(t))

            for n in self.compartment_name.keys():
                for m in self.compartment_info[n]['fromcpm'].keys():
                    for j in range(self.group):
                        for t in range(self.Time-1):
                            if (j,t) in self.compartment_info[n]['fromcpm'][m]['q_dp'].keys():
                                if self.__approximation == 'SF':
                                    for l in range(self.Lq[m][n][j][t]):
                                        self.model.addConstr(
                                            -self.model.getVarByName("U." + str(m) + "." + str(n) + "." + str(j) + "." + str(t))
                                             + self.model.getVarByName("W." + str(m) + "." + str(j) + "." + str(t))
                                             * self.vunder[m][n][j][t][l]
                                             + self.model.getVarByName("y." + str(m) + "." + str(n) + "." + str(j)
                                               + "." + str(t) + "." + str(l)) * self.__M <= self.__M,
                                             "dynamicsUlower." + str(m) + "." + str(n) + "." + str(j)
                                             + "." + str(t) + "." + str(l))


                                        self.model.addConstr(
                                            -self.model.getVarByName("U." + str(m) + "." + str(n) + "." + str(j) + "." + str(t))
                                             + self.model.getVarByName("W." + str(m) + "." + str(j) + "." + str(t))
                                             * self.vunder[m][n][j][t][l]
                                             - self.model.getVarByName("y." + str(m) + "." + str(n) + "." + str(j)
                                            + "." + str(t) + "." + str(l)) * self.__M >= -self.__M,
                                                             "dynamicsUupper." + str(m) + "." + str(n) + "." + str(j)
                                                             + "." + str(t) + "." + str(l))

                                elif self.__approximation == 'ME':
                                    self.model.addConstr(
                                        -self.model.getVarByName("U." + str(m) + "." + str(n) + "." + str(j) + "." + str(t))
                                         + self.model.getVarByName("W." + str(m) + "." + str(j) + "." + str(t))
                                         * self.eta1lower[m][n][j][t]
                                         + self.model.getVarByName("eta." + str(m) + "." + str(n) + "." + str(j)
                                                                   + "." + str(t)) * self.Wlower[m][j][t]
                                         <= self.eta1lower[m][n][j][t]* self.Wlower[m][j][t],
                                         "dynamicsUlower." + str(m) + "." + str(n) + "." + str(j)
                                         + "." + str(t)+ "." + str(0))
                                    self.model.addConstr(
                                        -self.model.getVarByName("U." + str(m) + "." + str(n) + "." + str(j) + "." + str(t))
                                         + self.model.getVarByName("W." + str(m) + "." + str(j) + "." + str(t)) * self.eta1upper[m][n][j][t]
                                         + self.model.getVarByName("eta." + str(m) + "." + str(n) + "." + str(j)
                                                                   + "." + str(t)) * self.Wupper[m][j][t]
                                         <= self.eta1upper[m][n][j][t] * self.Wupper[m][j][t],
                                         "dynamicsUlower." + str(m) + "." + str(n) + "." + str(j)
                                         + "." + str(t) + "." + str(1))
                                    self.model.addConstr(
                                        -self.model.getVarByName("U." + str(m) + "." + str(n) + "." + str(j) + "." + str(t))
                                         + self.model.getVarByName("W." + str(m) + "." + str(j) + "." + str(t)) * self.eta1upper[m][n][j][t]
                                         + self.model.getVarByName("eta." + str(m) + "." + str(n) + "." + str(j)
                                                                   + "." + str(t)) * self.Wlower[m][j][t]
                                         >= self.eta1upper[m][n][j][t] * self.Wlower[m][j][t],
                                         "dynamicsUupper." + str(m) + "." + str(n) + "." + str(j)
                                         + "." + str(t) + "." + str(0))
                                    self.model.addConstr(
                                        -self.model.getVarByName("U." + str(m) + "." + str(n) + "." + str(j) + "." + str(t))
                                         + self.model.getVarByName("W." + str(m) + "." + str(j) + "." + str(t))
                                        * self.eta1lower[m][n][j][t]
                                         + self.model.getVarByName("eta." + str(m) + "." + str(n) + "." + str(j) + "." + str(t))
                                        * self.Wupper[m][j][t]
                                         >= self.eta1lower[m][n][j][t] * self.Wupper[m][j][t],
                                         "dynamicsUupper." + str(m) + "." + str(n) + "." + str(j)
                                         + "." + str(t) + "." + str(1))
                            else:
                                self.model.addConstr(
                                    -self.model.getVarByName("U." + str(m) + "." + str(n) + "." + str(j) + "." + str(t))
                                     + self.model.getVarByName("W." + str(m) + "." + str(j) + "." + str(t))
                                                     * self.compartment_info[n]['fromcpm'][m]['q_idp'][j][t]
                                     == 0,
                                     "dynamicsUequal." + str(m) + "." + str(n) + "." + str(j) + "." + str(t))
            for m in self.compartment_name.keys():
                for k in range(self.group):
                    for t in range(self.Time-1):
                        self.model.addConstr(-self.model.getVarByName("W." + str(m) + "." + str(k) + "." + str(t))
                                             + self.model.getVarByName("X." + str(m) + "." + str(k) + "." + str(t)) *
                                             self.compartment_info[m]['local']['p_idp'][k][k][t]
                                             - gp.quicksum(self.model.getVarByName("x." + str(m) + "." + str(mm) + "." + str(k) + "." + str(t))
                                                           * self.compartment_info[m]['local']['p_idp'][k][k][t]
                                                           for mm in self.compartment_info[m]['tocpm'].keys()
                                                           if (k, t) in self.compartment_info[m]['tocpm'][mm]['x'])
                                             == 0,
                                             "dynamicsWequal." + str(m) + "." + str(k) + "." + str(t))
            if self.types == 'Robustness':
                for n in self.compartment_name.keys():
                    for j in range(self.group):
                        for t in range(self.Time - 1):
                            self.model.addConstr(
                                - self.model.getVarByName("Xdagger." + str(n) + "." + str(j) + "." + str(t+1))
                                 + gp.quicksum(self.model.getVarByName("U." + str(m) + "." + str(n) + "." + str(j) + "." + str(t))
                                               for m in self.compartment_info[n]['fromcpm'].keys())
                                 + gp.quicksum(self.model.getVarByName("Uhat." + str(m) + "." + str(n) + "." + str(j) + "." + str(t))
                                               for m in self.compartment_info[n]['fromcpm'].keys())
                                 == 0, "Rdynamics." + str(n) + "." + str(j) + "." + str(t))

                    for m in self.compartment_info[n]['fromcpm'].keys():
                        for j in range(self.group):
                            for t in range(self.Time-1):
                                if (j,t) in self.compartment_info[n]['fromcpm'][m]['q_dp'].keys():
                                    if self.__approximation == 'SF':
                                        for l in range(self.Lq[m][n][j][t]):
                                            self.model.addConstr(
                                                -self.model.getVarByName("Uhat." + str(m) + "." + str(n) + "." + str(j) + "." + str(t))
                                                 + self.model.getVarByName("What." + str(m) + "." + str(j) + "." + str(t))
                                                * self.vunder[m][n][j][t][l]**2
                                                 + self.model.getVarByName("y." + str(m) + "." + str(n) + "." + str(j)
                                            + "." + str(t) + "." + str(l)) * self.__M <= self.__M,
                                                 "RdynamicsUlower." + str(m) + "." + str(n) + "." + str(j)
                                                 + "." + str(t) + "." + str(l))
                                            self.model.addConstr(
                                                -self.model.getVarByName("Uhat." + str(m) + "." + str(n) + "." + str(j) + "." + str(t))
                                                 + self.model.getVarByName("What." + str(m) + "." + str(j) + "." + str(t))
                                                * self.vunder[m][n][j][t][l] ** 2
                                                 - self.model.getVarByName("y." + str(m) + "." + str(n) + "." + str(j)
                                            + "." + str(t) + "." + str(l)) * self.__M >= -self.__M,
                                                 "RdynamicsUupper." + str(m) + "." + str(n) + "." + str(j)
                                                 + "." + str(t) + "." + str(l))

                                    elif self.__approximation == 'ME':
                                        self.model.addConstr(-self.model.getVarByName("Uhat." + str(m) + "." + str(n) + "." + str(j) + "." + str(t))
                                                             + self.model.getVarByName("What." + str(m) + "." + str(j) + "." + str(t))
                                                             * self.eta1lower[m][n][j][t]**2
                                                             + self.model.getVarByName("phi." + str(m) + "." + str(n)
                                                                + "." + str(j) + "." + str(t)) * self.whatlower[m][j][t]
                                                             <= self.eta1lower[m][n][j][t]**2 * self.whatlower[m][j][t],
                                                             "RdynamicsUlower." + str(m) + "." + str(n) + "." + str(j)
                                                             + "." + str(t) + "." + str(0))
                                        self.model.addConstr(-self.model.getVarByName("Uhat." + str(m) + "." + str(n) + "." + str(j) + "." + str(t))
                                                             + self.model.getVarByName("What." + str(m) + "." + str(j) + "." + str(t))
                                                             * self.eta1upper[m][n][j][t] ** 2
                                                             + self.model.getVarByName("phi." + str(m) + "." + str(n)
                                                                + "." + str(j) + "." + str(t)) * self.whatupper[m][j][t]
                                                             <= self.eta1upper[m][n][j][t] ** 2 * self.whatupper[m][j][t],
                                                             "RdynamicsUlower." + str(m) + "." + str(n) + "." + str(j)
                                                             + "." + str(t) + "." + str(1))
                                        self.model.addConstr(-self.model.getVarByName("Uhat." + str(m) + "." + str(n) + "." + str(j) + "." + str(t))
                                                             + self.model.getVarByName("What." + str(m) + "." + str(j) + "." + str(t))
                                                             * self.eta1upper[m][n][j][t]**2
                                                             + self.model.getVarByName("phi." + str(m) + "." + str(n)
                                                                + "." + str(j) + "." + str(t)) * self.whatlower[m][j][t]
                                                             >= self.eta1upper[m][n][j][t]**2 * self.whatlower[m][j][t],
                                                             "RdynamicsUlower." + str(m) + "." + str(n) + "." + str(j)
                                                             + "." + str(t) + "." + str(0))
                                        self.model.addConstr(-self.model.getVarByName("Uhat." + str(m) + "." + str(n) + "." + str(j) + "." + str(t))
                                                             + self.model.getVarByName("What." + str(m) + "." + str(j) + "." + str(t))
                                                             * self.eta1lower[m][n][j][t] ** 2
                                                             + self.model.getVarByName("phi." + str(m) + "." + str(n)
                                                                + "." + str(j) + "." + str(t)) * self.whatupper[m][j][t]
                                                             >= self.eta1lower[m][n][j][t] ** 2 * self.whatupper[m][j][t],
                                                             "RdynamicsUupper." + str(m) + "." + str(n) + "." + str(j)
                                                             + "." + str(t) + "." + str(1))
                                        self.model.addConstr(-self.model.getVarByName("phi." + str(m) + "." + str(n)
                                                                + "." + str(j) + "." + str(t))
                                                             + self.model.getVarByName("eta." + str(m) + "." + str(n)
                                                                + "." + str(j) + "." + str(t))
                                                             * (self.eta1upper[m][n][j][t]+self.eta1lower[m][n][j][t])
                                                             >= self.eta1upper[m][n][j][t] * self.eta1lower[m][n][j][t],
                                                             "Rdynamics_phi_upper." + str(m) + "." + str(n) + "." + str(j)
                                                             + "." + str(t))
                                        self.model.addConstr(-self.model.getVarByName("phi." + str(m) + "." + str(n)
                                                                + "." + str(j) + "." + str(t))
                                                             + self.model.getVarByName("eta." + str(m) + "." + str(n)
                                                                + "." + str(j) + "." + str(t)) * (2*self.eta1lower[m][n][j][t])
                                                             <= self.eta1lower[m][n][j][t]**2,
                                                             "Rdynamics_phi_lower." + str(m) + "." + str(n) + "." + str(j)
                                                             + "." + str(t) + "." + str(0))
                                        self.model.addConstr(-self.model.getVarByName("phi." + str(m) + "." + str(n)
                                                                + "." + str(j) + "." + str(t))
                                                             + self.model.getVarByName("eta." + str(m) + "." + str(n)
                                                                + "." + str(j) + "." + str(t)) * (2*self.eta1lower[m][n][j][t])
                                                             <= self.eta1upper[m][n][j][t]**2,
                                                             "Rdynamics_phi_lower." + str(m) + "." + str(n) + "." + str(j)
                                                             + "." + str(t) + "." + str(1))
                                else:
                                    self.model.addConstr(-self.model.getVarByName("Uhat." + str(m) + "." + str(n) + "." + str(j) + "." + str(t))
                                                         + self.model.getVarByName("What." + str(m) + "." + str(j) + "." + str(t))
                                                         * self.compartment_info[n]['fromcpm'][m]['q_idp'][j][t]**2
                                                         == 0,
                                                         "RdynamicsUequal." + str(m) + "." + str(n) + "." + str(j)
                                                         + "." + str(t))

                for m in self.compartment_name.keys():
                    for k in range(self.group):
                        for t in range(self.Time-1):
                            self.model.addConstr(- self.model.getVarByName("What." + str(m) + "." + str(k) + "." + str(t))
                                                 + self.model.getVarByName("Xdagger." + str(m) + "." + str(k) + "." + str(t)) *
                                                 self.compartment_info[m]['local']['p_idp'][k][k][t] ** 2
                                                 - self.model.getVarByName("X." + str(m) + "." + str(k) + "." + str(t))*
                                                 self.compartment_info[m]['local']['p_idp'][k][k][t] ** 2
                                                 + gp.quicksum(
                                self.model.getVarByName("x." + str(m) + "." + str(mm) + "." + str(k) + "." + str(t))
                                * self.compartment_info[m]['local']['p_idp'][k][k][t] ** 2
                                for mm in self.compartment_info[m]['tocpm'].keys()
                                if (k, t) in self.compartment_info[m]['tocpm'][mm]['x'])
                                                 == 0, "RdynamicsWequal." + str(m) + "." + str(k) + "." + str(t))

            elif self.types != 'Robustness' and self.types != 'Expectation':
                raise ValueError('Incorrect optimization method\n '
                                 'Please either input type = Expectation, or type = Robustness')
        else:
            raise ValueError('Current version only supports Cplex and Gurobi as solvers')

    def __DynamicsLinear(self, eta, label=0):
        if self.solver == 'cplex':
            if label == 0:
                self.model.linear_constraints.add(lin_expr=[cplex.SparsePair(
                                               ["X." + str(n) + "." + str(j) + "." + str(t+1)]
                                             + ["x." + str(m) + "." + str(n) + "." + str(j) + "." + str(t)
                                                for m in self.compartment_info[n]['fromcpm'].keys()
                                                if (j,t) in self.compartment_info[n]['fromcpm'][m]['x']]
                                             + ["U." + str(m) + "." + str(n) + "." + str(j) + "." + str(t)
                                                for m in self.compartment_info[n]['fromcpm'].keys()],
                                             [-1]
                                             + [1 for m in self.compartment_info[n]['fromcpm'].keys()
                                                if (j,t) in self.compartment_info[n]['fromcpm'][m]['x']]
                                             + [1 for m in self.compartment_info[n]['fromcpm'].keys()])
                                                   for n in self.compartment_name.keys() for j in range(self.group)
                                                        for t in range(self.Time - 1)],
                                         senses=["E" for n in self.compartment_name.keys() for j in range(self.group)
                                                        for t in range(self.Time - 1)],
                                         rhs=[0 for n in self.compartment_name.keys() for j in range(self.group)
                                                        for t in range(self.Time - 1)],
                                         names=["dynamics." + str(n) + "." + str(j) + "." + str(t)
                                                for n in self.compartment_name.keys()
                                                for j in range(self.group) for t in range(self.Time - 1)])

            for n in self.compartment_name.keys():
                for m in self.compartment_info[n]['fromcpm'].keys():
                    # print(n, m, self.compartment_info[n]['fromcpm'][m]['q_dp'].keys())
                    for j in range(self.group):
                        for t in range(self.Time-1):
                            if (j,t) in self.compartment_info[n]['fromcpm'][m]['q_dp'].keys():
                                if label == 0:
                                    self.model.linear_constraints.add(
                                    lin_expr=[
                                        cplex.SparsePair(["U." + str(m) + "." + str(n) + "." + str(j) + "." + str(t),
                                                          "W." + str(m) + "." + str(j) + "." + str(t)],
                                                         [-1, eta[m][n][j][t]])],
                                    senses=["E"],
                                    rhs=[0],
                                    names=["dynamicsUequal." + str(m) + "." + str(n) + "." + str(j) + "." + str(t)])
                                else:
                                    self.model.linear_constraints.set_coefficients(
                                        "dynamicsUequal." + str(m) + "." + str(n) + "." + str(j) + "." + str(t),
                                         "W." + str(m) + "." + str(j) + "." + str(t),
                                         eta[m][n][j][t]
                                    )
                            else:
                                if label == 0:
                                    self.model.linear_constraints.add(
                                    lin_expr=[
                                        cplex.SparsePair(["U." + str(m) + "." + str(n) + "." + str(j) + "." + str(t),
                                                          "W." + str(m) + "." + str(j) + "." + str(t)],
                                       [-1, self.compartment_info[n]['fromcpm'][m]['q_idp'][j][t]])],
                                    senses=["E"],
                                    rhs=[0],
                                    names=["dynamicsUequal." + str(m) + "." + str(n) + "." + str(j) + "." + str(t)])
            if label == 0:
                for m in self.compartment_name.keys():
                    for k in range(self.group):
                        for t in range(self.Time-1):
                            self.model.linear_constraints.add(
                            lin_expr=[cplex.SparsePair(["W." + str(m) + "." + str(k) + "." + str(t)]
                                                       + ["X." + str(m) + "." + str(k) + "." + str(t)]
                                                       + ["x." + str(m) + "." + str(mm) + "." + str(k) + "." + str(t)
                                                          for mm in self.compartment_info[m]['tocpm'].keys()
                                                          if (k,t) in self.compartment_info[m]['tocpm'][mm]['x']],
                                                       [-1]
                                                       + [self.compartment_info[m]['local']['p_idp'][k][k][t]]
                                                       + [-self.compartment_info[m]['local']['p_idp'][k][k][t]
                                                          for mm in self.compartment_info[m]['tocpm'].keys()
                                                          if (k,t) in self.compartment_info[m]['tocpm'][mm]['x']]
                                                       )],
                            senses=["E"],
                            rhs=[0],
                            names=["dynamicsWequal." + str(m) + "." + str(k) + "." + str(t)])

            if self.types == 'Robustness':
                if label == 0:
                    self.model.linear_constraints.add(
                        lin_expr=[cplex.SparsePair(["Xdagger." + str(n) + "." + str(j) + "." + str(t + 1)]
                                                   + ["U." + str(m) + "." + str(n) + "." + str(j) + "." + str(t)
                                                      for m in self.compartment_info[n]['fromcpm'].keys()]
                                                   + ["Uhat." + str(m) + "." + str(n) + "." + str(j) + "." + str(t)
                                                      for m in self.compartment_info[n]['fromcpm'].keys()],
                                                   [-1]
                                                   + [1 for m in self.compartment_info[n]['fromcpm'].keys()]
                                                   + [1 for m in self.compartment_info[n]['fromcpm'].keys()])
                                  for n in self.compartment_name.keys() for j in range(self.group)
                                  for t in range(self.Time - 1)],
                        senses=["E" for n in self.compartment_name.keys() for j in range(self.group)
                                for t in range(self.Time - 1)],
                        rhs=[0 for n in self.compartment_name.keys() for j in range(self.group)
                             for t in range(self.Time - 1)],
                        names=["Rdynamics." + str(n) + "." + str(j) + "." + str(t)
                               for n in self.compartment_name.keys()
                               for j in range(self.group) for t in range(self.Time - 1)])

                for n in self.compartment_name.keys():
                    for m in self.compartment_info[n]['fromcpm'].keys():
                        for j in range(self.group):
                            for t in range(self.Time-1):
                                if (j,t) in self.compartment_info[n]['fromcpm'][m]['q_dp'].keys():
                                    if label == 0:
                                        self.model.linear_constraints.add(
                                        lin_expr=[cplex.SparsePair(
                                            ["Uhat." + str(m) + "." + str(n) + "." + str(j) + "." + str(t),
                                             "What." + str(m) + "." + str(j) + "." + str(t)],
                                             [-1, eta[m][n][j][t]**2])],
                                        senses=["E"],
                                        rhs=[0],
                                        names=["RdynamicsUequal." + str(m) + "." + str(n) + "." + str(j) + "." + str(t)])
                                    else:
                                        self.model.linear_constraints.set_coefficients(
                                            "RdynamicsUequal." + str(m) + "." + str(n) + "." + str(j) + "." + str(t),
                                             "What." + str(m) + "." + str(j) + "." + str(t),
                                             eta[m][n][j][t]**2)
                                else:
                                    if label == 0:
                                        self.model.linear_constraints.add(
                                            lin_expr=[cplex.SparsePair(
                                                ["Uhat." + str(m) + "." + str(n) + "." + str(j) + "." + str(t),
                                                 "What." + str(m) + "." + str(j) + "." + str(t)],
                                                 [-1, self.compartment_info[n]['fromcpm'][m]['q_idp'][j][t]**2])],
                                            senses=["E"],
                                            rhs=[0],
                                            names=["RdynamicsUequal." + str(m) + "." + str(n) + "." + str(j)
                                                   + "." + str(t)])

                if label == 0:
                    for m in self.compartment_name.keys():
                        for k in range(self.group):
                            for t in range(self.Time-1):
                                self.model.linear_constraints.add(
                                    lin_expr=[
                                        cplex.SparsePair(["What." + str(m) + "." + str(k)  + "." + str(t)]
                                                         + ["Xdagger." + str(m) + "." + str(k) + "." + str(t),
                                                            "X." + str(m) + "." + str(k) + "." + str(t)]
                                                         + ["x." + str(m) + "." + str(mm) + "." + str(k) + "." + str(t)
                                                            for mm in self.compartment_info[m]['tocpm'].keys()
                                                          if (k,t) in self.compartment_info[m]['tocpm'][mm]['x'] ],
                                                         [-1]
                                                         + [self.compartment_info[m]['local']['p_idp'][k][k][t]**2,
                                                            -self.compartment_info[m]['local']['p_idp'][k][k][t]**2]
                                                         + [self.compartment_info[m]['local']['p_idp'][k][k][t]**2
                                                            for mm in self.compartment_info[m]['tocpm'].keys()
                                                          if (k,t) in self.compartment_info[m]['tocpm'][mm]['x'] ])],
                                    senses=["E"],
                                    rhs=[0],
                                    names=["RdynamicsWequal." + str(m) + "." + str(k) + "." + str(t)])

            elif self.types != 'Robustness' and self.types != 'Expectation':
                raise ValueError('Incorrect optimization method\n '
                                 'Please either input type = Expectation, or type = Robustness')

        elif self.solver == 'gurobi':
            self.model.update()
            if label == 0:
                for n in self.compartment_name.keys():
                    for j in range(self.group):
                        for t in range(self.Time - 1):
                            self.model.addConstr(
                                (-self.model.getVarByName("X." + str(n) + "." + str(j) + "." + str(t+1))
                              + gp.quicksum(self.model.getVarByName("x." + str(m) + "." + str(n) + "."
                                                                    + str(j) + "." + str(t))
                                            for m in self.compartment_info[n]['fromcpm'].keys()
                                            if (j, t) in self.compartment_info[n]['fromcpm'][m]['x'])
                              + gp.quicksum(self.model.getVarByName("U." + str(m) + "." + str(n) + "."
                                                                    + str(j) + "." + str(t))
                                            for m in self.compartment_info[n]['fromcpm'].keys()) == 0),
                                'dynamics.' + str(n) + "." + str(j) + "." + str(t))

            for n in self.compartment_name.keys():
                for m in self.compartment_info[n]['fromcpm'].keys():
                    # print(n, m, self.compartment_info[n]['fromcpm'][m]['q_dp'].keys())
                    for j in range(self.group):
                        for t in range(self.Time-1):
                            if (j,t) in self.compartment_info[n]['fromcpm'][m]['q_dp'].keys():
                                if label == 0:
                                    self.model.addConstr(
                                        -self.model.getVarByName("U." + str(m) + "." + str(n) + "." + str(j) + "."
                                                                 + str(t))
                                        + self.model.getVarByName("W." + str(m) + "." + str(j) + "." + str(t))
                                        * eta[m][n][j][t] == 0,
                                        "dynamicsUequal." + str(m) + "." + str(n) + "." + str(j) + "." + str(t))
                                else:
                                    self.model.update()
                                    self.model.chgCoeff(self.model.getConstrByName("dynamicsUequal." + str(m) + "."
                                                        + str(n) + "." + str(j) + "." + str(t)),
                                                        self.model.getVarByName('W.'+str(m)+'.'+str(j) + '.' + str(t)),
                                                        eta[m][n][j][t])
                            else:
                                if label == 0:
                                    self.model.addConstr(
                                        -self.model.getVarByName("U." + str(m) + "." + str(n) + "." + str(j) + "."
                                                                 + str(t))
                                         + self.model.getVarByName("W." + str(m) + "." + str(j) + "." + str(t))
                                         * self.compartment_info[n]['fromcpm'][m]['q_idp'][j][t] == 0,
                                         "dynamicsUequal." + str(m) + "." + str(n) + "." + str(j) + "." + str(t))
            if label == 0:
                for m in self.compartment_name.keys():
                    for k in range(self.group):
                        for t in range(self.Time-1):
                            self.model.update()
                            self.model.addConstr(-self.model.getVarByName("W." + str(m) + "." + str(k) + "." + str(t))
                                                 + self.model.getVarByName("X." + str(m) + "." + str(k) + "." + str(t)) *
                                                 self.compartment_info[m]['local']['p_idp'][k][k][t]
                                                 - gp.quicksum(self.model.getVarByName("x." + str(m) + "." + str(mm)
                                                                                       + "." + str(k) + "." + str(t))
                                                               * self.compartment_info[m]['local']['p_idp'][k][k][t]
                                                               for mm in self.compartment_info[m]['tocpm'].keys()
                                                               if (k, t) in self.compartment_info[m]['tocpm'][mm]['x'])
                                                 == 0,
                                                 "dynamicsWequal." + str(m) + "." + str(k) + "." + str(t))

            if self.types == 'Robustness':
                if label == 0:
                    for n in self.compartment_name.keys():
                        for j in range(self.group):
                            for t in range(self.Time - 1):
                                self.model.addConstr(
                                    - self.model.getVarByName("Xdagger." + str(n) + "." + str(j) + "." + str(t+1))
                                 + gp.quicksum(self.model.getVarByName("U." + str(m) + "." + str(n) + "." + str(j)
                                                                       + "." + str(t))
                                               for m in self.compartment_info[n]['fromcpm'].keys())
                                 + gp.quicksum(self.model.getVarByName("Uhat." + str(m) + "." + str(n) + "." + str(j)
                                                                       + "." + str(t))
                                               for m in self.compartment_info[n]['fromcpm'].keys())
                                 == 0, "Rdynamics." + str(n) + "." + str(j) + "." + str(t))

                for n in self.compartment_name.keys():
                    for m in self.compartment_info[n]['fromcpm'].keys():
                        for j in range(self.group):
                            for t in range(self.Time-1):
                                if (j,t) in self.compartment_info[n]['fromcpm'][m]['q_dp'].keys():
                                    if label == 0:
                                        self.model.addConstr(
                                            -self.model.getVarByName("Uhat." + str(m) + "." + str(n)
                                                                  + "." + str(j) + "." + str(t))
                                         + self.model.getVarByName("W." + str(m) + "." + str(j)
                                                                   + "." + str(t)) * eta[m][n][j][t]**2
                                         == 0,
                                         "RdynamicsUequal." + str(m) + "." + str(n) + "." + str(j)
                                         + "." + str(t))
                                    else:
                                        self.model.update()
                                        self.model.chgCoeff(self.model.getConstrByName("RdynamicsUequal." + str(m)
                                                            + "." + str(n) + "." + str(j) + "." + str(t)),
                                                            self.model.getVarByName('What.' + str(m) + '.' + str(j) + '.' + str(t)),
                                                            eta[m][n][j][t]**2)
                                else:
                                    if label == 0:
                                        self.model.addConstr(-self.model.getVarByName("Uhat." + str(m) + "." + str(n) + "." + str(j) + "." + str(t))
                                                             + self.model.getVarByName("What." + str(m) + "." + str(j) + "." + str(t)) *
                                                             self.compartment_info[n]['fromcpm'][m]['q_idp'][j][t] ** 2
                                                             == 0,
                                                             "RdynamicsUequal." + str(m) + "." + str(n) + "." + str(j)
                                                             + "." + str(t))

                if label == 0:
                    for m in self.compartment_name.keys():
                        for k in range(self.group):
                            for t in range(self.Time-1):
                                self.model.addConstr(
                                    - self.model.getVarByName("What." + str(m) + "." + str(k) + "." + str(t))
                                    + self.model.getVarByName("Xdagger." + str(m) + "." + str(k) + "." + str(t)) *
                                    self.compartment_info[m]['local']['p_idp'][k][k][t] ** 2
                                    - self.model.getVarByName("X." + str(m) + "." + str(k) + "." + str(t)) *
                                    self.compartment_info[m]['local']['p_idp'][k][k][t] ** 2
                                    + gp.quicksum(self.model.getVarByName(
                                        "x." + str(m) + "." + str(mm) + "." + str(k) + "." + str(t)) *
                                                  self.compartment_info[m]['local']['p_idp'][k][k][t] ** 2
                                                  for mm in self.compartment_info[m]['tocpm'].keys()
                                                  if (k, t) in self.compartment_info[m]['tocpm'][mm]['x'])
                                    == 0,
                                    "RdynamicsWequal." + str(m) + "." + str(k) + "." + str(t))
            elif self.types != 'Robustness' and self.types != 'Expectation':
                raise ValueError('Incorrect optimization method\n '
                                 'Please either input type = Expectation, or type = Robustness')
        else:
            raise ValueError('Current version only supports Cplex and Gurobi as solvers')

    '''=============== Step 6. Define step function  ====================='''
    def stepfunction(self, label=None):
        if label == 'p':
            if self.solver == 'cplex':
                for m in self.compartment_name.keys():
                    state = self.compartment_info[m]['local']['p_dp'].keys()
                    self.model.variables.add(obj=[0 for (j, k, t) in state for l in range(self.Lp[m][j][k][t])if t != self.Time-1],
                                             lb=[0 for (j, k, t) in state for l in range(self.Lp[m][j][k][t])if t != self.Time-1],
                                             ub=[1 for (j, k, t) in state for l in range(self.Lp[m][j][k][t])if t != self.Time-1],
                                             types = ['B' for (j, k, t) in state for l in range(self.Lp[m][j][k][t])if t != self.Time-1],
                                             names=["z." + str(m) + "." + str(j) + "." + str(k) + "." + str(t) + "." + str(l)
                                                    for (j, k, t) in state for l in range(self.Lp[m][j][k][t]) if t != self.Time-1])
                    for (j, k, t) in state:
                        if t != self.Time - 1:
                            var = ["z." + str(m) + "." + str(j) + "." + str(k) + "." + str(t) + "." + str(l)
                                   for l in range(self.Lp[m][j][k][t])]
                            coef = [1 for l in range(self.Lp[m][j][k][t])]
                            sense = 'E'
                            rhs = 1
                            names = "System_stepfunctionp." + str(m) + "." + str(j) + "." + str(k) + "." + str(t)
                            self.model.linear_constraints.add(
                                lin_expr=[cplex.SparsePair(var, coef)],
                                senses=[sense],
                                rhs=[rhs],
                                names=[names])

                            a = self.compartment_info[m]['local']['p_dp'][(j,k,t)][0]
                            b = self.compartment_info[m]['local']['p_dp'][(j,k,t)][1]
                            for l in range(self.Lp[m][j][k][t]):
                                var = ["z." + str(m) + "." + str(j) + "." + str(k) + "." + str(t) + "." + str(l)] \
                                      + ["X." + str(n) + "." + str(jj) + "." + str(tau)
                                        for n in self.compartment_name.keys() for jj in range(self.group)
                                        for tau in range(t+1)] \
                                      + ["x." + str(n) + "." + str(nn) + "." + str(jj) + "." + str(tau)
                                        for n in self.compartment_name.keys() for nn in self.compartment_info[n]['tocpm'].keys()
                                        for jj in range(self.group) for tau in range(t + 1)
                                        if (jj,tau) in self.compartment_info[n]['tocpm'][nn]['x'] ]
                                coef = [self.__M] \
                                       + [a[n][jj][tau] for n in self.compartment_name.keys()
                                          for jj in range(self.group) for tau in range(t+1)]  \
                                       + [b[n][nn][jj][tau] for n in self.compartment_name.keys()
                                          for nn in self.compartment_info[n]['tocpm'].keys()
                                          for jj in range(self.group) for tau in range(t+1)
                                        if (jj,tau) in self.compartment_info[n]['tocpm'][nn]['x'] ]
                                sense = "L"
                                rhs = self.__M+ (self.uover[m][j][k][t][l+1]-self.gap/100)-self.compartment_info[m]['local']['p_idp'][j][k][t]
                                names = "System_stepfunction_upper_p." + str(m) + "." + str(j) + "." + str(k) \
                                        + "." + str(t)+ "." + str(l)
                                self.model.linear_constraints.add(
                                    lin_expr=[cplex.SparsePair(var, coef)],
                                    senses=[sense],
                                    rhs=[rhs],
                                    names=[names])

                                coef = [-self.__M] \
                                       + [a[n][jj][tau] for n in self.compartment_name.keys()
                                          for jj in range(self.group) for tau in range(t+1)] \
                                       + [b[n][nn][jj][tau] for n in self.compartment_name.keys()
                                          for nn in self.compartment_info[n]['tocpm'].keys()
                                          for jj in range(self.group) for tau in range(t+1)
                                        if (jj,tau) in self.compartment_info[n]['tocpm'][nn]['x'] ]
                                sense = "G"
                                rhs = -self.__M+ self.uover[m][j][k][t][l]-self.compartment_info[m]['local']['p_idp'][j][k][t]
                                names = "System_stepfunction_lower_p." + str(m) + "." + str(j) + "." + str(k) \
                                        + "." + str(t)+ "." + str(l)
                                self.model.linear_constraints.add(
                                    lin_expr=[cplex.SparsePair(var, coef)],
                                    senses=[sense],
                                    rhs=[rhs],
                                    names=[names])
            elif self.solver == 'gurobi':
                for m in self.compartment_name.keys():
                    state = self.compartment_info[m]['local']['p_dp'].keys()
                    for (j, k, t) in state:
                        if t != self.Time - 1:
                            for l in range(self.Lp[m][j][k][t]):
                                self.model.addVar(vtype=GRB.BINARY,name="z." + str(m) + "." + str(j)
                                                                        + "." + str(k) + "." + str(t) + "." + str(l))

                    for (j, k, t) in state:
                        if t != self.Time - 1:
                            self.model.addConstr(
                                (gp.quicksum(self.model.getVarByName("z." + str(m) + "." + str(j)
                                             + "." + str(k) + "." + str(t) + "." + str(l)) for l in range(self.Lp[m][j][k][t])) == 1),
                                "System_stepfunctionp." + str(m) + "." + str(j) + "." + str(k) + "." + str(t))

                            a = self.compartment_info[m]['local']['p_dp'][(j, k, t)][0]
                            b = self.compartment_info[m]['local']['p_dp'][(j, k, t)][1]
                            for l in range(self.Lp[m][j][k][t]):
                                rhs = self.__M + (self.uover[m][j][k][t][l + 1] - self.gap / 100) - \
                                      self.compartment_info[m]['local']['p_idp'][j][k][t]
                                self.model.addConstr(
                                    (self.model.getVarByName("z." + str(m) + "." + str(j)
                                                             + "." + str(k) + "." + str(t) + "." + str(l))*self.__M
                                     + gp.quicksum(self.model.getVarByName("X." + str(n) + "." + str(jj) + "." + str(tau)) * a[n][jj][t]
                                                   for n in self.compartment_name.keys()
                                                   for jj in range(self.group) for tau in range(t + 1))
                                     + gp.quicksum(self.model.getVarByName("x." + str(n) + "." + str(nn) + "." + str(jj)
                                                                           + "." + str(tau)) * b[n][nn][jj][tau]
                                                   for n in self.compartment_name.keys()
                                                  for nn in self.compartment_info[n]['tocpm'].keys()
                                                  for jj in range(self.group) for tau in range(t + 1)
                                                  if (jj, tau) in self.compartment_info[n]['tocpm'][nn]['x']) <= rhs),
                                    "System_stepfunction_upper_p." + str(m) + "." + str(j) + "." + str(k) \
                                        + "." + str(t) + "." + str(l))
                                rhs = -self.__M + self.uover[m][j][k][t][l] - \
                                      self.compartment_info[m]['local']['p_idp'][j][k][t]
                                self.model.addConstr(
                                    (- self.model.getVarByName("z." + str(m) + "." + str(j)
                                                             + "." + str(k) + "." + str(t) + "." + str(l)) * self.__M
                                     + gp.quicksum(self.model.getVarByName("X." + str(n) + "." + str(jj) + "." + str(tau)) * a[n][jj][tau]
                                                   for n in self.compartment_name.keys()
                                                   for jj in range(self.group) for tau in range(t + 1))
                                     + gp.quicksum(self.model.getVarByName("x." + str(n) + "." + str(nn) + "." + str(jj)
                                                                           + "." + str(tau)) * b[n][nn][jj][tau]
                                                   for n in self.compartment_name.keys()
                                                   for nn in self.compartment_info[n]['tocpm'].keys()
                                                   for jj in range(self.group) for tau in range(t + 1)
                                                   if (jj, tau) in self.compartment_info[n]['tocpm'][nn]['x'])
                                     >= rhs),
                                    "System_stepfunction_upper_p." + str(m) + "." + str(j) + "." + str(k) \
                                    + "." + str(t) + "." + str(l))
        elif label == 'q':
            for m in self.compartment_name.keys():
                for n in self.compartment_info[m]['tocpm'].keys():
                    state = list(self.compartment_info[m]['tocpm'][n]['q_dp'].keys())
                    state.sort()
            if self.solver == 'cplex':
                for m in self.compartment_name.keys():
                    for n in self.compartment_info[m]['tocpm'].keys():
                        state = list(self.compartment_info[m]['tocpm'][n]['q_dp'].keys())
                        state.sort()
                        self.model.variables.add(obj=[0 for (j,t) in state if t<self.Time-1 for l in range(self.Lq[m][n][j][t])],
                                         lb=[0 for (j,t) in state if t<self.Time-1 for l in range(self.Lq[m][n][j][t])],
                                         ub=[1 for (j,t) in state if t<self.Time-1 for l in range(self.Lq[m][n][j][t])],
                                         types = ['B' for (j,t) in state if t<self.Time-1 for l in range(self.Lq[m][n][j][t])],
                                         names=["y." + str(m) + "." + str(n) + "." + str(j) + "." + str(t) + "." + str(l)
                                                for (j,t) in state if t<self.Time-1 for l in range(self.Lq[m][n][j][t])])

                        for (j, t) in state:
                            if t < self.Time-1:
                                var = ["y." + str(m) + "." + str(n) + "." + str(j) + "." + str(t) + "." + str(l)
                                       for l in range(self.Lq[m][n][j][t])]
                                coef = [1 for l in range(self.Lq[m][n][j][t])]
                                sense = 'E'
                                rhs = 1
                                names = "System_stepfunctionq." + str(m) + "." + str(n) + "." + str(j) + "." + str(t)
                                self.model.linear_constraints.add(
                                    lin_expr=[cplex.SparsePair(var, coef)],
                                    senses=[sense],
                                    rhs=[rhs],
                                    names=[names])

                        for (j, t) in state:
                            if t < self.Time-1:
                                betabar = self.compartment_info[m]['tocpm'][n]['q_norm'][(j,t)][0]
                                gammabar = self.compartment_info[m]['tocpm'][n]['q_norm'][(j, t)][1]
                                phibar = self.compartment_info[m]['tocpm'][n]['q_norm'][(j, t)][2]
                                psibar = self.compartment_info[m]['tocpm'][n]['q_norm'][(j, t)][3]
                                alphabar = self.compartment_info[m]['tocpm'][n]['q_norm'][(j,t)][4]
                                beta = self.compartment_info[m]['tocpm'][n]['q_dp'][(j,t)][0]
                                gamma = self.compartment_info[m]['tocpm'][n]['q_dp'][(j, t)][1]
                                phi = self.compartment_info[m]['tocpm'][n]['q_dp'][(j, t)][2]
                                psi = self.compartment_info[m]['tocpm'][n]['q_dp'][(j, t)][3]
                                alpha = self.compartment_info[m]['tocpm'][n]['q_dp'][(j, t)][4]
                                for l in range(self.Lq[m][n][j][t]):
                                    var = ["y." + str(m) + "." + str(n) + "." + str(j) + "." + str(t) + "." + str(l)] \
                                          + ['X.' + str(mm) + "." + str(k) + "." + str(t)
                                             for mm in self.compartment_name.keys() for k in range(self.group)] \
                                          + ['x.' + str(mm) + "." + str(nn) + "." + str(k) + "." + str(t)
                                             for mm in self.compartment_name.keys()
                                             for nn in self.compartment_info[mm]['tocpm'].keys() for k in range(self.group)
                                             if (k,t) in self.compartment_info[mm]['tocpm'][nn]['x']] \
                                          + ['V.' + str(mm) + "." + str(k) + "." + str(j) + "." + str(t)
                                             for mm in self.compartment_name.keys() for k in range(self.group)] \
                                          + ['Vhat.' + str(mm) + "." + str(k) + "." + str(j) + "." + str(t)
                                             for mm in self.compartment_name.keys() for k in range(self.group)]
                                    coef = [-self.__M] \
                                           + [(self.vover[m][n][j][t][l + 1]-self.gap/1000) * gammabar[mm][k] - gamma[mm][k]
                                              for mm in self.compartment_name.keys() for k in range(self.group)] \
                                           + [(self.vover[m][n][j][t][l + 1]-self.gap/1000) * psibar[mm][k] - psi[mm][k]
                                              for mm in self.compartment_name.keys() for nn in
                                              self.compartment_info[mm]['tocpm'].keys() for k in range(self.group)
                                              if (k, t) in self.compartment_info[mm]['tocpm'][nn]['x']] \
                                           + [(self.vover[m][n][j][t][l + 1]-self.gap/1000) * betabar[mm][k] - beta[mm][k]
                                             for mm in self.compartment_name.keys() for k in range(self.group)] \
                                          + [(self.vover[m][n][j][t][l + 1]-self.gap/1000) * phibar[mm][k] - phi[mm][k]
                                             for mm in self.compartment_name.keys() for k in range(self.group)]
                                    rhs = -self.__M - (self.vover[m][n][j][t][l + 1]-self.gap/1000) * alphabar + alpha
                                    sense = 'G'
                                    names = "System_stepfunction_upper_q." + str(m) + "." + str(n) + "." + str(j)\
                                            + "." + str(t) + "." + str(l)
                                    self.model.linear_constraints.add(
                                        lin_expr=[cplex.SparsePair(var, coef)],
                                        senses=[sense],
                                        rhs=[rhs],
                                        names=[names])

                                    coef = [self.__M] \
                                           + [self.vover[m][n][j][t][l] * gammabar[mm][k] - gamma[mm][k]
                                              for mm in self.compartment_name.keys() for k in range(self.group)] \
                                           + [self.vover[m][n][j][t][l] * psibar[mm][k] - psi[mm][k]
                                              for mm in self.compartment_name.keys() for nn in
                                              self.compartment_info[mm]['tocpm'].keys() for k in range(self.group)
                                              if (k, t) in self.compartment_info[mm]['tocpm'][nn]['x']] \
                                           + [self.vover[m][n][j][t][l] * betabar[mm][k] - beta[mm][k]
                                              for mm in self.compartment_name.keys() for k in range(self.group)] \
                                           + [self.vover[m][n][j][t][l] * phibar[mm][k] - phi[mm][k]
                                              for mm in self.compartment_name.keys() for k in range(self.group)]
                                    rhs = self.__M - self.vover[m][n][j][t][l] * alphabar + alpha
                                    sense = 'L'
                                    names = "System_stepfunction_lower_q." + str(m) + "." + str(n) + "." + str(j)\
                                            + "." + str(t) + "." + str(l)
                                    self.model.linear_constraints.add(
                                        lin_expr=[cplex.SparsePair(var, coef)],
                                        senses=[sense],
                                        rhs=[rhs],
                                        names=[names])

                for mm in self.compartment_name.keys():
                    for k in range(self.group):
                        for j in range(self.group):
                            for t in range(self.Time-1):
                                if (k,j,t) in self.compartment_info[mm]['local']['p_dp'].keys():
                                    for l in range(self.Lp[mm][k][j][t]):
                                        var = ['V.' + str(mm) + "." + str(k) + "." + str(j) + "." + str(t),
                                               'X' + "." + str(mm) + "." + str(k) + "." + str(t),
                                               "z." + str(mm) + "." + str(k) + "." + str(j) + "." + str(t) + "." + str(l)]
                                        coef = [-1, self.uunder[mm][k][j][t][l], self.__M]
                                        rhs = self.__M
                                        sense = 'L'
                                        names = "System_stepfunction_lower_V." + str(mm) + "." + str(k) + "." + str(
                                            j) + "." + str(t) + "." + str(l)
                                        self.model.linear_constraints.add(
                                            lin_expr=[cplex.SparsePair(var, coef)],
                                            senses=[sense],
                                            rhs=[rhs],
                                            names=[names])

                                        coef = [-1, self.uunder[mm][k][j][t][l], -self.__M]
                                        rhs = -self.__M
                                        sense = 'G'
                                        names = "System_stepfunction_upper_V." + str(mm) + "." + str(k) + "." + str(
                                            j) + "." + str(t) + "." + str(l)
                                        self.model.linear_constraints.add(
                                            lin_expr=[cplex.SparsePair(var, coef)],
                                            senses=[sense],
                                            rhs=[rhs],
                                            names=[names])


                                        var = ['Vhat.' + str(mm) + "." + str(k) + "." + str(j) + "." + str(t)]\
                                               + ['x' + "." + str(mm) + "." + str(nn) + "." + str(k) + "." + str(t)
                                                  for nn in self.compartment_info[mm]['tocpm'].keys()
                                                    if (k,t) in self.compartment_info[mm]['tocpm'][nn]['x']]\
                                               + ["z." + str(mm) + "." + str(k) + "." + str(j) + "." + str(t) + "." + str(l)]
                                        coef = [-1] \
                                               + [self.uunder[mm][k][j][t][l]
                                                  for nn in self.compartment_info[mm]['tocpm'].keys()
                                                    if (k,t) in self.compartment_info[mm]['tocpm'][nn]['x']]\
                                               + [self.__M]
                                        rhs = self.__M
                                        sense = 'L'
                                        names = "System_stepfunction_lower_Vhat." + str(mm) + "." + str(k) + "." + str(
                                            j) + "." + str(t) + "." + str(l)
                                        self.model.linear_constraints.add(
                                            lin_expr=[cplex.SparsePair(var, coef)],
                                            senses=[sense],
                                            rhs=[rhs],
                                            names=[names])

                                        coef = [-1] \
                                               + [self.uunder[mm][k][j][t][l]
                                                  for nn in self.compartment_info[mm]['tocpm'].keys()
                                                    if (k,t) in self.compartment_info[mm]['tocpm'][nn]['x']]\
                                               + [-self.__M]
                                        rhs = -self.__M
                                        sense = 'G'
                                        names = "System_stepfunction_upper_Vhat." + str(mm) + "." + str(k) + "." + str(
                                            j) + "." + str(t) + "." + str(l)
                                        self.model.linear_constraints.add(
                                            lin_expr=[cplex.SparsePair(var, coef)],
                                            senses=[sense],
                                            rhs=[rhs],
                                            names=[names])
                                else:
                                    var = ['V.'  + str(mm) + "." + str(k) + "." + str(j) + "." + str(t),
                                           'X' + "." + str(mm) + "." + str(k) + "." + str(t)]
                                    coef = [-1, self.compartment_info[mm]['local']['p_idp'][k][j][t] ]
                                    rhs = 0
                                    sense = 'E'
                                    names = "System_stepfunction_equal_V." + str(mm) + "." + str(k) + "." + str(
                                        j) + "." + str(t)
                                    self.model.linear_constraints.add(
                                        lin_expr=[cplex.SparsePair(var, coef)],
                                        senses=[sense],
                                        rhs=[rhs],
                                        names=[names])

                                    var = ['Vhat.' + str(mm) + "." + str(k) + "." + str(j) + "." + str(t)] \
                                           + ['x' + "." + str(mm) + "." + str(nn) + "." + str(k) + "." + str(t)
                                                  for nn in self.compartment_info[mm]['tocpm'].keys()
                                                    if (k,t) in self.compartment_info[mm]['tocpm'][nn]['x']]
                                    coef = [-1]\
                                           + [self.compartment_info[mm]['local']['p_idp'][k][j][t]
                                              for nn in self.compartment_info[mm]['tocpm'].keys()
                                                    if (k,t) in self.compartment_info[mm]['tocpm'][nn]['x'] ]
                                    rhs = 0
                                    sense = 'E'
                                    names = "System_stepfunction_equal_Vhat." + str(mm) + "." + str(k) + "." + str(
                                        j) + "." + str(t)
                                    self.model.linear_constraints.add(
                                        lin_expr=[cplex.SparsePair(var, coef)],
                                        senses=[sense],
                                        rhs=[rhs],
                                        names=[names])
            elif self.solver == 'gurobi':
                self.Decision_y = {}
                for m in self.compartment_name.keys():
                    for n in self.compartment_info[m]['tocpm'].keys():
                        state = list(self.compartment_info[m]['tocpm'][n]['q_dp'].keys())
                        state.sort()
                        for (j, t) in state:
                            if t < self.Time - 1:
                                for l in range(self.Lq[m][n][j][t]):
                                    self.Decision_y[m, n, j, t, l] = self.model.addVar(
                                        vtype=GRB.BINARY,
                                        name="y." + str(m) + "." + str(n) + "." + str(j) + "." + str(t) + "." + str(l))
                        for (j, t) in state:
                            if t < self.Time - 1:
                                self.model.addConstr(
                                    (gp.quicksum(
                                        self.Decision_y[m, n, j, t, l] for l in range(self.Lq[m][n][j][t])) == 1),
                                    "System_stepfunctionq." + str(m) + "." + str(n) + "." + str(j) + "." + str(t))

                        for (j, t) in state:
                            if t < self.Time - 1:
                                betabar = self.compartment_info[m]['tocpm'][n]['q_norm'][(j, t)][0]
                                gammabar = self.compartment_info[m]['tocpm'][n]['q_norm'][(j, t)][1]
                                phibar = self.compartment_info[m]['tocpm'][n]['q_norm'][(j, t)][2]
                                psibar = self.compartment_info[m]['tocpm'][n]['q_norm'][(j, t)][3]
                                alphabar = self.compartment_info[m]['tocpm'][n]['q_norm'][(j, t)][4]
                                beta = self.compartment_info[m]['tocpm'][n]['q_dp'][(j, t)][0]
                                gamma = self.compartment_info[m]['tocpm'][n]['q_dp'][(j, t)][1]
                                phi = self.compartment_info[m]['tocpm'][n]['q_dp'][(j, t)][2]
                                psi = self.compartment_info[m]['tocpm'][n]['q_dp'][(j, t)][3]
                                alpha = self.compartment_info[m]['tocpm'][n]['q_dp'][(j, t)][4]
                                for l in range(self.Lq[m][n][j][t]):
                                    rhs = -self.__M - (self.vover[m][n][j][t][
                                                           l + 1] - self.gap / 1000) * alphabar + alpha
                                    self.model.addConstr(
                                        (-self.Decision_y[m, n, j, t, l] * self.__M
                                         + gp.quicksum(self.Decision_X[mm][k][t]
                                                       * (
                                                       (self.vover[m][n][j][t][l + 1] - self.gap / 1000) * gammabar[mm][
                                                           k] - gamma[mm][k])
                                                       for mm in self.compartment_name.keys() for k in
                                                       range(self.group))
                                         + gp.quicksum(self.Decision_x[mm, nn, k, t]
                                                       * (
                                                       (self.vover[m][n][j][t][l + 1] - self.gap / 1000) * psibar[mm][
                                                           k] - psi[mm][k])
                                                       for mm in self.compartment_name.keys() for nn in
                                                       self.compartment_info[mm]['tocpm'].keys() for k in
                                                       range(self.group))
                                         + gp.quicksum(self.Decision_V[mm, k, j, t]
                                                       * (
                                                       (self.vover[m][n][j][t][l + 1] - self.gap / 1000) * betabar[mm][
                                                           k] - beta[mm][k])
                                                       for mm in self.compartment_name.keys() for k in
                                                       range(self.group))
                                         + gp.quicksum(self.Decision_Vhat[mm, k, j, t]
                                                       * (
                                                       (self.vover[m][n][j][t][l + 1] - self.gap / 1000) * phibar[mm][
                                                           k] - phi[mm][k])
                                                       for mm in self.compartment_name.keys() for k in
                                                       range(self.group)) >= rhs),
                                        "System_stepfunction_upper_q." + str(m) + "." + str(n) + "." + str(j) \
                                        + "." + str(t) + "." + str(l))
                                    rhs = self.__M - self.vover[m][n][j][t][l] * alphabar + alpha
                                    self.model.addConstr(
                                        (self.Decision_y[m, n, j, t, l] * self.__M
                                         + gp.quicksum(self.Decision_X[mm][k][t]
                                                       * (self.vover[m][n][j][t][l] * gammabar[mm][k] - gamma[mm][k])
                                                       for mm in self.compartment_name.keys() for k in
                                                       range(self.group))
                                         + gp.quicksum(self.Decision_x[mm, nn, k, t]
                                                       * (self.vover[m][n][j][t][l] * psibar[mm][k] - psi[mm][k])
                                                       for mm in self.compartment_name.keys() for nn in
                                                       self.compartment_info[mm]['tocpm'].keys() for k in
                                                       range(self.group))
                                         + gp.quicksum(self.Decision_V[mm, k, j, t]
                                                       * (self.vover[m][n][j][t][l] * betabar[mm][k] - beta[mm][k])
                                                       for mm in self.compartment_name.keys() for k in
                                                       range(self.group))
                                         + gp.quicksum(self.Decision_Vhat[mm, k, j, t]
                                                       * (self.vover[m][n][j][t][l] * phibar[mm][k] - phi[mm][k])
                                                       for mm in self.compartment_name.keys() for k in
                                                       range(self.group)) <= rhs),
                                        "System_stepfunction_lower_q." + str(m) + "." + str(n) + "." + str(j) \
                                        + "." + str(t) + "." + str(l))


                for mm in self.compartment_name.keys():
                    for k in range(self.group):
                        for j in range(self.group):
                            for t in range(self.Time-1):
                                if (k,j,t) in self.compartment_info[mm]['local']['p_dp'].keys():
                                    for l in range(self.Lp[mm][k][j][t]):
                                        self.model.addConstr(
                                            (-self.model.getVarByName("V." + str(mm) + "." + str(k) + "." + str(j) + "." + str(t))
                                             + self.model.getVarByName("X." + str(mm) + "." + str(k) + "." + str(t)) * self.uunder[mm][k][j][t][l]
                                             + self.model.getVarByName("z." + str(mm) + "." + str(k)
                                        + "." + str(j) + "." + str(t) + "." + str(l)) * self.__M <= self.__M),
                                            "System_stepfunction_lower_V." + str(mm) + "." + str(k) + "." + str(
                                                j) + "." + str(t) + "." + str(l)
                                        )
                                        self.model.addConstr(
                                            (-self.model.getVarByName("V." + str(mm) + "." + str(k) + "." + str(j) + "." + str(t))
                                             + self.model.getVarByName("X." + str(mm) + "." + str(k) + "." + str(t)) * self.uunder[mm][k][j][t][l]
                                             - self.model.getVarByName("z." + str(mm) + "." + str(k)
                                        + "." + str(j) + "." + str(t) + "." + str(l)) * self.__M >= -self.__M),
                                            "System_stepfunction_upper_V." + str(mm) + "." + str(k) + "." + str(
                                            j) + "." + str(t) + "." + str(l)
                                        )
                                        self.model.addConstr(
                                            (-self.model.getVarByName("Vhat." + str(mm) + "." + str(k) + "." + str(j) + "." + str(t))
                                             + gp.quicksum(self.model.getVarByName("x." + str(mm) + "." + str(nn) + "." + str(k) + "." + str(t))
                                                           * self.uunder[mm][k][j][t][l]
                                                  for nn in self.compartment_info[mm]['tocpm'].keys()
                                                    if (k,t) in self.compartment_info[mm]['tocpm'][nn]['x'])
                                             + self.model.getVarByName("z." + str(mm) + "." + str(k)
                                        + "." + str(j) + "." + str(t) + "." + str(l)) * self.__M <= self.__M),
                                            "System_stepfunction_lower_Vhat." + str(mm) + "." + str(k) + "." + str(
                                                j) + "." + str(t) + "." + str(l)
                                        )
                                        self.model.addConstr(
                                            (-self.model.getVarByName("Vhat." + str(mm) + "." + str(k) + "." + str(j) + "." + str(t))
                                             + gp.quicksum(self.model.getVarByName("x." + str(mm) + "." + str(nn) + "." + str(k) + "." + str(t))
                                                           * self.uunder[mm][k][j][t][l]
                                                  for nn in self.compartment_info[mm]['tocpm'].keys()
                                                    if (k,t) in self.compartment_info[mm]['tocpm'][nn]['x'])
                                             - self.model.getVarByName("z." + str(mm) + "." + str(k)
                                        + "." + str(j) + "." + str(t) + "." + str(l)) * self.__M >= -self.__M),
                                            "System_stepfunction_upper_Vhat." + str(mm) + "." + str(k) + "." + str(
                                            j) + "." + str(t) + "." + str(l)
                                        )
                                else:
                                    self.model.addConstr(
                                        (-self.model.getVarByName("V." + str(mm) + "." + str(k) + "." + str(j) + "." + str(t))
                                         + self.model.getVarByName("X." + str(mm) + "." + str(k) + "." + str(t))
                                         * self.compartment_info[mm]['local']['p_idp'][k][j][t]
                                         == 0),
                                        "System_stepfunction_equal_V." + str(mm) + "." + str(k) + "." + str(
                                        j) + "." + str(t)
                                    )
                                    self.model.addConstr(
                                        (-self.model.getVarByName("Vhat." + str(mm) + "." + str(k) + "." + str(j) + "." + str(t))
                                         + gp.quicksum(self.model.getVarByName("x." + str(mm) + "." + str(nn) + "." + str(k) + "." + str(t))
                                                       * self.compartment_info[mm]['local']['p_idp'][k][j][t]
                                                       for nn in self.compartment_info[mm]['tocpm'].keys()
                                                       if (k, t) in self.compartment_info[mm]['tocpm'][nn]['x'])
                                         == 0),
                                        "System_stepfunction_equal_Vhat." + str(mm) + "." + str(k) + "." + str(
                                        j) + "." + str(t)
                                    )
        else:
            raise ValueError('label should be p or q')

    def set_approximation(self,opt='ME'):
        '''opt in {ME (0), SF (1), SO (2)}'''
        if opt == 0:
            opt ='ME'
        elif opt == 1:
            opt = 'SF'
        elif opt == 2:
            opt = 'SO'

        if opt in ['ME','SF','SO']:
            self.__approximation = opt
        else:
            raise ValueError('opt should belong to the set {ME, SF, SO}')

    def __Wbar(self,Sbar, Sunder, Ibar, Iunder, S, I):
        return min(Sbar * I + Iunder * S - Sbar * Iunder, Sunder * I + Ibar * S - Sunder * Ibar)

    def __Wunder(self, Sbar, Sunder, Ibar, Iunder, S, I):
        return max(Sbar * I + Ibar * S - Sbar * Ibar, Sunder * I + Iunder * S - Sunder * Iunder)

    def __set_bound_McCormickEnvelopes(self):
        Xupper = [[[self.compartment_info[n]['local']['population'][j] if t == 0 else 0 for t in range(self.Time)]
                   for j in range(self.group)] for n in range(max(self.compartment_name.keys())+1)]
        Xlower = [[[self.compartment_info[n]['local']['population'][j] if t == 0 else 0 for t in range(self.Time)]
                   for j in range(self.group)] for n in range(max(self.compartment_name.keys())+1)]
        xupper, xlower = self.__set_dvar_bound()

        scale = 1
        gamma1 = [[[[self.compartment_info[m]['tocpm'][n]['q_dp'][(j, t)][1]
                   if n in self.compartment_info[m]['tocpm'].keys()
                      and self.compartment_info[m]['tocpm'][n]['q_dp'] != {} else 0
                     for t in range(self.Time)]for j in range(self.group)]
                   for n in self.compartment_name.keys()] for m in self.compartment_name.keys()]
        psi1 = [[[[self.compartment_info[m]['tocpm'][n]['q_dp'][(j, t)][3]
                 if n in self.compartment_info[m]['tocpm'].keys()
                    and self.compartment_info[m]['tocpm'][n]['q_dp'] != {} else 0
                   for t in range(self.Time)]for j in range(self.group)]
                 for n in self.compartment_name.keys()] for m in self.compartment_name.keys()]
        alpha1 = [[[[self.compartment_info[m]['tocpm'][n]['q_dp'][(j, t)][4]
                   if n in self.compartment_info[m]['tocpm'].keys()
                      and self.compartment_info[m]['tocpm'][n]['q_dp'] != {} else 0
                     for t in range(self.Time)]for j in range(self.group)]
                   for n in self.compartment_name.keys()] for m in self.compartment_name.keys()]

        if self.types == 'Robustness':
            Xdupper = [[[0 for t in range(self.Time)]
                        for j in range(self.group)] for n in range(max(self.compartment_name.keys()) + 1)]
            Xdlower = [[[0 for t in range(self.Time)]
                        for j in range(self.group)] for n in range(max(self.compartment_name.keys()) + 1)]

        for t in range(self.Time - 1):
            for j in range(self.group):
                eta1upper = [[sum([Xupper[mm][k][t] / scale * gamma1[m][n][j][t][mm][k]
                                   for mm in self.compartment_name.keys() for k in range(self.group)]) \
                              + sum([xlower[mm][nn][k][t] / scale * psi1[m][n][j][t][mm][k]
                                     for mm in self.compartment_name.keys()
                                     for nn in self.compartment_info[mm]['tocpm'].keys() for k in range(self.group)
                                     if (k, t) in self.compartment_info[mm]['tocpm'][nn]['x']]) + alpha1[m][n][j][t]
                              if n in self.compartment_info[m]['tocpm'].keys()
                                 and self.compartment_info[m]['tocpm'][n]['q_dp'] != {} else 0
                              for n in self.compartment_name.keys()] for m in self.compartment_name.keys()]
                eta1lower = [[sum([Xlower[mm][k][t] / scale * gamma1[m][n][j][t][mm][k]
                                   for mm in self.compartment_name.keys() for k in range(self.group)]) \
                              + sum([xupper[mm][nn][k][t] / scale * psi1[m][n][j][t][mm][k]
                                     for mm in self.compartment_name.keys()
                                     for nn in self.compartment_info[mm]['tocpm'].keys() for k in range(self.group)
                                     if (k, t) in self.compartment_info[mm]['tocpm'][nn]['x']]) + alpha1[m][n][j][t]
                              if n in self.compartment_info[m]['tocpm'].keys()
                                 and self.compartment_info[m]['tocpm'][n]['q_dp'] != {} else 0
                              for n in self.compartment_name.keys()] for m in self.compartment_name.keys()]
                u1upper = [max(Xupper[m][j][t] - sum([xlower[m][mm][j][t] for mm in self.compartment_name.keys()]), 0)
                           for m in self.compartment_name.keys()]
                u1lower = [max(Xlower[m][j][t] - sum([xupper[m][mm][j][t] for mm in self.compartment_name.keys()]), 0)
                           for m in self.compartment_name.keys()]

                if self.types == 'Robustness':
                    whatupper = [Xdupper[m][j][t] - max(Xupper[m][j][t] - sum([xlower[m][mm][j][t]
                                for mm in self.compartment_name.keys()]), 0) for m in self.compartment_name.keys()]
                    whatlower = [Xdlower[m][j][t] - max(Xlower[m][j][t] - sum([xupper[m][mm][j][t]
                                for mm in self.compartment_name.keys()]), 0) for m in self.compartment_name.keys()]

                for n in self.compartment_name.keys():
                    Xupper[n][j][t + 1] = max(0, sum([xupper[m][n][j][t] for m in self.compartment_name.keys()]) \
                                              + u1upper[n] \
                                              + sum([self.compartment_info[m]['tocpm'][n]['q_idp'][j][t] * u1upper[m]
                                                     for m in self.compartment_name.keys()
                                                     if n != m and n in self.compartment_info[m]['tocpm'].keys()]) \
                                              + sum([u1upper[m] * eta1upper[m][n]
                                                     if m != n and n in self.compartment_info[m]['tocpm'].keys()
                                                        and self.compartment_info[m]['tocpm'][n]['q_dp'] != {} else 0
                                                     for m in self.compartment_name.keys()]) \
                                              - sum([self.compartment_info[n]['tocpm'][m]['q_idp'][j][t] * u1upper[n]
                                                     for m in self.compartment_name.keys()
                                                     if n != m and m in self.compartment_info[n]['tocpm'].keys()]) \
                                              - sum([u1upper[n] * eta1lower[n][m]
                                                     if m != n and m in self.compartment_info[n]['tocpm'].keys() \
                                                        and self.compartment_info[n]['tocpm'][m]['q_dp'] != {} else 0
                                                     for m in self.compartment_name.keys()]))

                    Xlower[n][j][t + 1] = max(0, sum([xlower[m][n][j][t] for m in self.compartment_name.keys()]) \
                                              + u1lower[n] \
                                              + sum([self.compartment_info[m]['tocpm'][n]['q_idp'][j][t] * u1lower[m]
                                                     for m in self.compartment_name.keys()
                                                     if n != m and n in self.compartment_info[m]['tocpm'].keys()]) \
                                              + sum([u1lower[m] * eta1lower[m][n]
                                                     if m != n and n in self.compartment_info[m]['tocpm'].keys() \
                                                        and self.compartment_info[m]['tocpm'][n]['q_dp'] != {} else 0
                                                     for m in self.compartment_name.keys()]) \
                                              - sum([self.compartment_info[n]['tocpm'][m]['q_idp'][j][t] * u1lower[n]
                                                     for m in self.compartment_name.keys()
                                                     if n != m and m in self.compartment_info[n]['tocpm'].keys()]) \
                                              - sum([u1lower[n] * eta1upper[n][m]
                                                     if m != n and m in self.compartment_info[n]['tocpm'].keys() \
                                                        and self.compartment_info[n]['tocpm'][m]['q_dp'] != {} else 0
                                                     for m in self.compartment_name.keys()]))
                    if self.types == 'Robustness':
                        if whatupper[n] >= 0:
                            qq = (1 - sum([self.compartment_info[n]['tocpm'][m]['q_idp'][j][t]
                                           for m in self.compartment_name.keys()
                                           if n != m and m in self.compartment_info[n]['tocpm'].keys()])
                                  - sum([eta1lower[n][m]
                                         if m != n and m in self.compartment_info[n]['tocpm'].keys()
                                            and self.compartment_info[n]['tocpm'][m]['q_dp'] != {} else 0
                                         for m in self.compartment_name.keys()])) ** 2
                        else:
                            qq = (1 - sum([self.compartment_info[n]['tocpm'][m]['q_idp'][j][t]
                                           for m in self.compartment_name.keys()
                                           if n != m and m in self.compartment_info[n]['tocpm'].keys()])
                                  - sum([eta1upper[n][m]
                                         if m != n and m in self.compartment_info[n]['tocpm'].keys()
                                            and self.compartment_info[n]['tocpm'][m]['q_dp'] != {} else 0
                                         for m in self.compartment_name.keys()])) ** 2

                        Xdupper[n][j][t + 1] = max(0, u1upper[n] \
                               + sum([self.compartment_info[m]['tocpm'][n]['q_idp'][j][t] * u1upper[m]
                                 for m in self.compartment_name.keys()
                                 if n != m and n in self.compartment_info[m]['tocpm'].keys()]) \
                               + sum([u1upper[m] * eta1upper[m][n]
                                      if m != n and n in self.compartment_info[m]['tocpm'].keys()
                                         and self.compartment_info[m]['tocpm'][n]['q_dp'] != {} else 0
                                      for m in self.compartment_name.keys()]) \
                               - sum([self.compartment_info[n]['tocpm'][m]['q_idp'][j][t] * u1upper[n]
                                 for m in self.compartment_name.keys()
                                 if n != m and m in self.compartment_info[n]['tocpm'].keys()]) \
                               - sum([u1upper[n] * eta1lower[n][m]
                                      if m != n and m in self.compartment_info[n]['tocpm'].keys()
                                         and self.compartment_info[n]['tocpm'][m]['q_dp'] != {} else 0
                                      for m in self.compartment_name.keys()]) \
                               + sum([self.compartment_info[m]['tocpm'][n]['q_idp'][j][t] ** 2 * whatupper[m]
                                 for m in self.compartment_name.keys()
                                 if n != m and n in self.compartment_info[m]['tocpm'].keys()]) \
                               + sum([max(whatupper[m] * eta1upper[m][n] ** 2, whatupper[m] * eta1lower[m][n] ** 2)
                                 if m != n and n in self.compartment_info[m]['tocpm'].keys()
                                    and self.compartment_info[m]['tocpm'][n]['q_dp'] != {} else 0
                                 for m in self.compartment_name.keys()]) \
                               + qq * whatupper[n])

                        if whatlower[n] < 0:
                            qq = (1 - sum([self.compartment_info[n]['tocpm'][m]['q_idp'][j][t]
                                           for m in self.compartment_name.keys()
                                           if n != m and m in self.compartment_info[n]['tocpm'].keys()])
                                  - sum([eta1lower[n][m] if m != n and m in self.compartment_info[n]['tocpm'].keys()
                                            and self.compartment_info[n]['tocpm'][m]['q_dp'] != {} else 0
                                         for m in self.compartment_name.keys()])) ** 2
                        else:
                            qq = (1 - sum([self.compartment_info[n]['tocpm'][m]['q_idp'][j][t]
                                           for m in self.compartment_name.keys()
                                           if n != m and m in self.compartment_info[n]['tocpm'].keys()])
                                  - sum([eta1upper[n][m] if m != n and m in self.compartment_info[n]['tocpm'].keys()
                                            and self.compartment_info[n]['tocpm'][m]['q_dp'] != {} else 0
                                         for m in self.compartment_name.keys()])) ** 2
                        Xdlower[n][j][t + 1] = max(0, u1lower[n] \
                               + sum( [self.compartment_info[m]['tocpm'][n]['q_idp'][j][t] * u1lower[m]
                                     for m in self.compartment_name.keys()
                                     if n != m and n in self.compartment_info[m]['tocpm'].keys()]) \
                               + sum([u1lower[m] * eta1lower[m][n]
                                      if m != n and n in self.compartment_info[m]['tocpm'].keys() \
                                         and self.compartment_info[m]['tocpm'][n]['q_dp'] != {} else 0
                                      for m in self.compartment_name.keys()]) \
                               - sum([self.compartment_info[n]['tocpm'][m]['q_idp'][j][t] * u1lower[n]
                                     for m in self.compartment_name.keys()
                                     if n != m and m in self.compartment_info[n]['tocpm'].keys()]) \
                               - sum([u1lower[n] * eta1upper[n][m]
                                      if m != n and m in self.compartment_info[n]['tocpm'].keys() \
                                         and self.compartment_info[n]['tocpm'][m]['q_dp'] != {} else 0
                                      for m in self.compartment_name.keys()]) \
                               + sum([self.compartment_info[m]['tocpm'][n]['q_idp'][j][t] ** 2 * whatlower[m]
                                     for m in self.compartment_name.keys()
                                     if n != m and n in self.compartment_info[m]['tocpm'].keys()]) \
                               + sum([min(whatlower[m] * eta1lower[m][n] ** 2, whatlower[m] * eta1upper[m][n] ** 2)
                                     if m != n and n in self.compartment_info[m]['tocpm'].keys() \
                                        and self.compartment_info[m]['tocpm'][n]['q_dp'] != {} else 0
                                     for m in self.compartment_name.keys()]) \
                               + qq * whatlower[n])

        self.Wupper = [[[max(0,Xupper[n][j][t] - sum([xlower[n][m][j][t] for m in self.compartment_name.keys()]))
                    for t in range(self.Time)] for j in range(self.group)] for n in self.compartment_name.keys()]
        self.Wlower = [[[max(0,Xlower[n][j][t] - sum([xupper[n][m][j][t] for m in self.compartment_name.keys()]))
                    for t in range(self.Time)] for j in range(self.group)] for n in self.compartment_name.keys()]

        self.eta1upper = [[[[sum([Xupper[mm][k][t] / scale * gamma1[m][n][j][t][mm][k]
                           for mm in self.compartment_name.keys() for k in range(self.group)]) \
                      + sum([xlower[mm][nn][k][t] / scale * psi1[m][n][j][t][mm][k]
                             for mm in self.compartment_name.keys()
                             for nn in self.compartment_info[mm]['tocpm'].keys() for k in range(self.group)
                             if (k, t) in self.compartment_info[mm]['tocpm'][nn]['x']]) + alpha1[m][n][j][t]
                      if n in self.compartment_info[m]['tocpm'].keys()
                         and self.compartment_info[m]['tocpm'][n]['q_dp'] != {} and n!=m else 0
                      for t in range(self.Time)] for j in range(self.group)]
                      for n in self.compartment_name.keys()] for m in self.compartment_name.keys()]
        self.eta1lower = [[[[sum([Xlower[mm][k][t] / scale * gamma1[m][n][j][t][mm][k]
                           for mm in self.compartment_name.keys() for k in range(self.group)]) \
                      + sum([xupper[mm][nn][k][t] / scale * psi1[m][n][j][t][mm][k]
                             for mm in self.compartment_name.keys()
                             for nn in self.compartment_info[mm]['tocpm'].keys() for k in range(self.group)
                             if (k, t) in self.compartment_info[mm]['tocpm'][nn]['x']]) + alpha1[m][n][j][t]
                      if n in self.compartment_info[m]['tocpm'].keys()
                         and self.compartment_info[m]['tocpm'][n]['q_dp'] != {} and n!=m else 0
                      for t in range(self.Time)] for j in range(self.group)]
                      for n in self.compartment_name.keys()] for m in self.compartment_name.keys()]

        for n in self.compartment_name.keys():
            self.eta1upper[n][n] = [[1 - sum([self.eta1lower[n][m][j][t] for m in self.compartment_name.keys() if n != m])
                                     for t in range(self.Time)] for j in range(self.group)]
            self.eta1lower[n][n] = [[1 - sum([self.eta1upper[n][m][j][t] for m in self.compartment_name.keys() if n != m])
                                     for t in range(self.Time)] for j in range(self.group)]

        if self.types == 'Robustness':
            self.whatupper = [[[Xdupper[m][j][t] - max(Xlower[m][j][t] - sum([xupper[m][mm][j][t]
                                for mm in self.compartment_name.keys()]), 0) for t in range(self.Time)]
                               for j in range(self.group)] for m in self.compartment_name.keys()]

            self.whatlower = [[[Xdlower[m][j][t] - max(Xupper[m][j][t] - sum([xlower[m][mm][j][t]
                                for mm in self.compartment_name.keys()]), 0) for t in range(self.Time)]
                               for j in range(self.group)] for m in self.compartment_name.keys()]

    def __McCormickEnvelopesfunction(self):
        if self.solver == 'cplex':
            for m in self.compartment_name.keys():
                for n in self.compartment_info[m]['tocpm'].keys():
                    state = list(self.compartment_info[m]['tocpm'][n]['q_dp'].keys())
                    state.sort()
                    self.model.variables.add(
                        obj=[0 for (j, t) in state if t < self.Time - 1],
                        lb=[self.eta1lower[m][n][j][t] for (j, t) in state if t < self.Time - 1],
                        ub=[self.eta1upper[m][n][j][t] for (j, t) in state if t < self.Time - 1],
                        names=["eta." + str(m) + "." + str(n) + "." + str(j) + "." + str(t)
                               for (j, t) in state if t < self.Time - 1])
            for m in self.compartment_name.keys():
                for n in self.compartment_info[m]['tocpm'].keys():
                    state = list(self.compartment_info[m]['tocpm'][n]['q_dp'].keys())
                    state.sort()
                    for (j, t) in state:
                        if t < self.Time-1:
                            gamma = self.compartment_info[m]['tocpm'][n]['q_dp'][(j, t)][1]
                            psi = self.compartment_info[m]['tocpm'][n]['q_dp'][(j, t)][3]
                            alpha = self.compartment_info[m]['tocpm'][n]['q_dp'][(j, t)][4]
                            var = ["eta." + str(m) + "." + str(n) + "." + str(j) + "." + str(t)] \
                                  + ['X.' + str(mm) + "." + str(k) + "." + str(t)
                                     for mm in self.compartment_name.keys() for k in range(self.group)] \
                                  + ['x.' + str(mm) + "." + str(nn) + "." + str(k) + "." + str(t)
                                     for mm in self.compartment_name.keys()
                                     for nn in self.compartment_info[mm]['tocpm'].keys() for k in range(self.group)
                                     if (k, t) in self.compartment_info[mm]['tocpm'][nn]['x']]
                            coef = [1] \
                                   + [-gamma[mm][k]
                                      for mm in self.compartment_name.keys() for k in range(self.group)] \
                                   + [-psi[mm][k]
                                      for mm in self.compartment_name.keys() for nn in
                                      self.compartment_info[mm]['tocpm'].keys() for k in range(self.group)
                                      if (k, t) in self.compartment_info[mm]['tocpm'][nn]['x']]
                            rhs = alpha
                            sense = 'E'
                            names = "System_McCormickEnvelopes." + str(m) + "." + str(n) + "." + str(j) \
                                    + "." + str(t)
                            self.model.linear_constraints.add(
                                lin_expr=[cplex.SparsePair(var, coef)],
                                senses=[sense],
                                rhs=[rhs],
                                names=[names])
        elif self.solver == 'gurobi':
            for m in self.compartment_name.keys():
                for n in self.compartment_info[m]['tocpm'].keys():
                    state = list(self.compartment_info[m]['tocpm'][n]['q_dp'].keys())
                    state.sort()
                    for (j, t) in state:
                        if t < self.Time - 1:
                            self.model.addVar(lb=self.eta1lower[m][n][j][t], ub=self.eta1upper[m][n][j][t],
                                              name="eta." + str(m) + "." + str(n) + "." + str(j) + "." + str(t))
            self.model.update()
            for m in self.compartment_name.keys():
                for n in self.compartment_info[m]['tocpm'].keys():
                    state = list(self.compartment_info[m]['tocpm'][n]['q_dp'].keys())
                    state.sort()
                    for (j, t) in state:
                        if t < self.Time-1:
                            gamma = self.compartment_info[m]['tocpm'][n]['q_dp'][(j, t)][1]
                            psi = self.compartment_info[m]['tocpm'][n]['q_dp'][(j, t)][3]
                            alpha = self.compartment_info[m]['tocpm'][n]['q_dp'][(j, t)][4]
                            self.model.addConstr(self.model.getVarByName("eta." + str(m) + "." + str(n) + "." + str(j) + "." + str(t))
                                   - gp.quicksum(self.model.getVarByName("X." + str(mm) + "." + str(k) + "." + str(t)) * gamma[mm][k]
                                                 for mm in self.compartment_name.keys() for k in range(self.group))
                                   - gp.quicksum(self.model.getVarByName("x." + str(mm) + "." + str(nn) + "." + str(k)
                                                                         + "." + str(t)) * psi[mm][k]
                                      for mm in self.compartment_name.keys() for nn in
                                      self.compartment_info[mm]['tocpm'].keys() for k in range(self.group)
                                      if (k, t) in self.compartment_info[mm]['tocpm'][nn]['x']) == alpha,
                                   "System_McCormickEnvelopes." + str(m) + "." + str(n) + "." + str(j) + "." + str(t))

    '''=============== Step 9. Define other constraints  ====================='''
    def custom_lp(self, fvar= [], fcoef = [], dvar= [], dcoef = [], sense=None,
                  rhs=None, name=None, label=None, option=None):
        if option == 'clear':
            self.con_var_para = []
        elif option == None:
            if sense == 'G':
                if fcoef != []:
                    fcoef = [-i for i in fcoef]
                if dcoef != []:
                    dcoef = [-i for i in dcoef]
                rhs = -rhs
                sense = 'L'
            var1 = []
            var2 = []
            coef1 = []
            coef2 = []
            var3 = []
            coef3 = []
            for i in range(len(fvar)):
                if len(fvar[i]) == 3:
                    n = get_keys(self.compartment_name, fvar[i][0])[0]
                    var1.append('X.' + str(n) + '.' + str(fvar[i][1]) + '.' + str(fvar[i][2]))
                    coef1.append(fcoef[i])
                elif len(fvar[i]) == 4:
                    n = get_keys(self.compartment_name, fvar[i][0])[0]
                    m = get_keys(self.compartment_name, fvar[i][1])[0]
                    var2.append('x.' + str(n) + '.' + str(m) + '.' + str(fvar[i][2]) + '.' + str(fvar[i][3]))
                    coef2.append(fcoef[i])
            if len(dvar) == len(dcoef) != [] and \
                    all([isinstance(dvar[i], str) and dvar[i] in self.custom_var.keys() for i in range(len(dvar))]):
                # for i in range(len(dvar)):
                var3.extend(dvar)
                coef3.extend(dcoef)

            if var1 == []:
                label = 'Expectation'
            else:
                if label != 'Expectation':
                    label = 'Robustness'
                else:
                    label = 'Expectation'
            self.con_var_para.append([var1+var2+var3, coef1+coef2+coef3, sense, rhs, name, label])

    def custom_qp(self, fvarlp= [], fcoeflp = [], dvarlp= [], dcoeflp = [],
                  fvarqp= [], fcoefqp = [], dvarqp= [], dcoefqp = [], sense=None, rhs=None, name=None, option=None):
        if sense not in ['G', 'L']:
            raise ValueError('sense should be G (greater than) or L (less than)')
        if option == 'clear':
            self.con_var_para_qp = []
        elif option == None:
            if sense == 'G':
                if fcoeflp != []:
                    fcoeflp = [-i for i in fcoeflp]
                if dcoeflp != []:
                    dcoeflp = [-i for i in dcoeflp]
                if fcoefqp != []:
                    fcoefqp = [-i for i in fcoefqp]
                if dcoefqp != []:
                    dcoefqp = [-i for i in dcoefqp]
                rhs = -rhs
                sense = 'L'
            var1 = []
            coef1 = []
            var2 = []
            coef2 = []
            var3 = []
            coef3 = []
            for i in range(len(fvarlp)):
                if len(fvarlp[i]) == 3:
                    n = get_keys(self.compartment_name, fvarlp[i][0])[0]
                    var1.append('X.' + str(n) + '.' + str(fvarlp[i][1]) + '.' + str(fvarlp[i][2]))
                    coef1.append(fcoeflp[i])
                elif len(fvarlp[i]) == 4:
                    n = get_keys(self.compartment_name, fvarlp[i][0])[0]
                    m = get_keys(self.compartment_name, fvarlp[i][1])[0]
                    var2.append('x.' + str(n) + '.' + str(m) + '.' + str(fvarlp[i][2]) + '.' + str(fvarlp[i][3]))
                    coef2.append(fcoeflp[i])
            if len(dvarlp) == len(dcoeflp) != [] and \
                    all([isinstance(dvarlp[i], str) and dvarlp[i] in self.custom_var.keys() for i in range(len(dvarlp))]):
                # for i in range(len(dvar)):
                var3.extend(dvarlp)
                coef3.extend(dcoeflp)

            var4 = []
            coef4 = []
            var5 = []
            coef5 = []
            var6 = []
            coef6 = []
            for i in range(len(fvarqp)):
                if len(fvarqp[i]) == 3:
                    n = get_keys(self.compartment_name, fvarqp[i][0])[0]
                    var4.append('X.' + str(n) + '.' + str(fvarqp[i][1]) + '.' + str(fvarqp[i][2]))
                    coef4.append(fcoefqp[i])
                elif len(fvarqp[i]) == 4:
                    n = get_keys(self.compartment_name, fvarqp[i][0])[0]
                    m = get_keys(self.compartment_name, fvarqp[i][1])[0]
                    var5.append('x.' + str(n) + '.' + str(m) + '.' + str(fvarqp[i][2]) + '.' + str(fvarqp[i][3]))
                    coef5.append(fcoefqp[i])

            if len(dvarqp) == len(dcoefqp) != [] and \
                    all([isinstance(dvarqp[i], str) and dvarqp[i] in self.custom_var.keys() for i in range(len(dvarqp))]):
                # for i in range(len(dvar)):
                var6.extend(dvarqp)
                coef6.extend(dcoefqp)

            label = 'Expectation'
            self.con_var_para_qp.append([var1+var2+var3, coef1+coef2+coef3,
                                         var4+var5+var6, coef4+coef5+coef6, sense, rhs, name, label])

    def set_initial_solution(self, compartment1=None, compartment2=None, val = None):
        if val == None or len(val) != self.group or all([len(val[i]) == self.Time for i in range(len(val))])==0:
            raise ValueError('val is not defined or its dimension is wrong')
        if compartment1 == None or compartment2 == None:
            raise ValueError('compartment1 or compartment2 is not given')
        if compartment1 not in self.compartment_name.values() or compartment2 not in self.compartment_name.values():
            raise ValueError('the name of compartment1 or compartment2 is not defined')
        # if compartment2 == None:
        #     n = get_keys(self.compartment_name,compartment1)[0]
        #     self.__initialsol.update({n: val})
        # else:
        n = get_keys(self.compartment_name, compartment1)[0]
        m = get_keys(self.compartment_name, compartment2)[0]
        self.__initialsol.update({(n,m): val})

    def solve(self, label='Expectation', ct=None, target=None):
        self.set_transition_compartment_self()
        self.Lq = None
        self.Lp = None
        self.vunder = None
        self.vover = None
        self.uunder = None
        self.uover = None

        self.expobj = []
        self.robustobj = []
        self.constraint = {}
        self.quadraticconstraint = {}

        for i in self.compartment_name:
            self.set_transition_group(compartment=i, prob=[[[1 if k==j else 0 for t in range(self.Time)]
                                                            for k in range(self.group)] for j in range(self.group)])
        if label == 'Expectation':
            self.types = 'Expectation'
            if self.__approximation in ['SF','ME']:
                if self.__approximation == 'ME':
                    self.__set_bound_McCormickEnvelopes()
                    # self.__McCormickEnvelopesfunction()
                self.dvar_conpartment()
                self.set_objectivetype(sense=self.objective_type)
                for i in self.objectiveterm:
                    self.set_objective(state=i[0], value=i[2])
                for i in self.con_var_para:
                    self.defineLinearConstraints(var=i[0], coef=i[1], sense=i[2], rhs=i[3], name=i[4],
                                                 types='Expectation')
                for i in self.con_var_para_qp:
                    self.defineQuadraticConstraints(varlp=i[0], coeflp=i[1], varqp=i[2], coefqp=i[3], sense=i[4],
                                                    rhs=i[5], name=i[6])
                if self.__approximation == 'SF':
                    self.set_step()
                    self.stepfunction(label='p')
                    self.stepfunction(label='q')
                elif self.__approximation == 'ME':
                    self.__McCormickEnvelopesfunction()
                self.__Dynamics()
                self.constraintgeneration()
                self.model.write('model_expectation_ME.mps')
                self.model.write('model_expectation_ME.lp')
                solution = self.solvemodel(ct=ct)
                if isinstance(solution, (int, float)):
                    self.status = solution
                    return self.status
                else:
                    return solution
            elif self.__approximation == 'SO':
                if self.__log_stream == 1:
                    startSO = time.clock()
                    outcome = open(self.__log_file + '.txt', 'w')

                bestobj = [1e15]
                # generate the initial solution
                if self.__initialsol == {}:
                    self.set_approximation('ME')
                    sol = self.solve(label='Expectation', ct=ct, target=target)
                    self.set_approximation('SO')
                    if isinstance(sol, (int, float)):
                        self.status = sol
                        return self.status
                    else:
                        (self.status, obj, Xopt, xopt, Xc) = sol
                    label2 = 0
                else:
                    xopt = [[[[0 for t in range(self.Time)] for j in range(self.group)]
                             for m in self.compartment_info.keys()] for n in self.compartment_info.keys()]
                    for i in self.__initialsol.keys():
                        xopt[i[0]][i[1]] = self.__initialsol[i]
                    label2 = 1
                label1 = 0
                if self.solver == 'cplex':
                    self.model.linear_constraints.delete()
                    self.model.quadratic_constraints.delete()
                    self.model.variables.delete()
                elif self.solver =='gurobi':
                    self.model.remove(self.model.getConstrs())
                    self.model.remove(self.model.getQConstrs())
                    self.model.remove(self.model.getVars())

                self.model.write('123.mps')
                self.model.write('123.lp')
                iter = 0
                while 1:
                    iter += 1
                    print('|-- Iteration ' + str(iter))
                    if self.__log_stream == 1:
                        outcome.write('|-- Iteration ' + str(iter)+'\n')
                    # step 1. simulation and generate value of compartment and probability eta
                    self.x = xopt
                    Xopt, xopt, eta = self.DeterPrediction()

                    # step 2 generate or update the model
                    if label1 == 0:
                        self.dvar_conpartment()
                        self.set_objectivetype(sense=self.objective_type)
                        for i in self.objectiveterm:
                            self.set_objective(state=i[0], value=i[2])
                        if label2 == 1:
                            for i in self.con_var_para:
                                self.defineLinearConstraints( var=i[0], coef=i[1], sense=i[2], rhs=i[3],
                                                        name=i[4], types='Expectation')
                            for i in self.con_var_para_qp:
                                self.defineQuadraticConstraints(varlp=i[0], coeflp=i[1], varqp=i[2], coefqp=i[3],
                                                                sense=i[4], rhs=i[5], name=i[6])
                    self.__DynamicsLinear(eta=eta,label=label1)
                    if label1 == 0:
                        self.constraintgeneration()
                    self.model.write('model_expectation_SO_' + str(iter) + '.mps')
                    self.model.write('model_expectation_SO_' + str(iter) + '.lp')
                    label1 = 1

                    # step 5. solve model
                    solution = self.solvemodel(ct=ct)
                    if isinstance(solution, (int, float)) and len(bestobj) == 1:
                        print('Initial solution is infeasible for the optimization model...')
                        print('|-----------------------------------')
                        if self.__log_stream == 1:
                            outcome.write('Initial solution is infeasible for the optimization model...\n')
                            outcome.write('|-----------------------------------\n')
                        bestobj.append(bestobj[-1]*1.01)
                    elif isinstance(solution, (int, float)) and len(bestobj) > 1 and min(bestobj) >= bestobj[0]:
                        print('Cannot generate a feasible solution, please check the initial solution '
                              'or model setting...')
                        print('|-----------------------------------')
                        if self.__log_stream == 1:
                            outcome.write('Cannot generate a feasible solution, please check the initial solution '
                                  'or model setting...\n')
                            outcome.write('|-----------------------------------\n')
                            endSO = time.clock()
                            outcome.write('Computational time ' + str(endSO - startSO) + '\n')
                            outcome.close()
                        return 0
                    elif isinstance(solution, (int, float)) and len(bestobj) > 1 and min(bestobj) < bestobj[0]:
                        print('|-- Optimal model at current iteration is infeasible but a feasible solution is '
                              'generated at the previous iteration')
                        print('|-- Objective values at all iterations:', bestobj)
                        print('|-- Optimal objective:', bestsol[1])
                        print('|-- Best x =')
                        if self.__log_stream == 1:
                            outcome.write('|-- Optimal model at current iteration is infeasible but a feasible '
                                          'solution is generated at the previous iteration\n')
                            outcome.write('|-- Objective values at all iterations: ' +str(bestobj)+'\n')
                            outcome.write('|-- Optimal objective: ' + str(bestsol[1])+'\n')
                            outcome.write('|-- Best x =\n')
                        for n in self.compartment_name.keys():
                            for m in self.compartment_info[n]['tocpm'].keys():
                                if sum([sum(bestsol[3][n][m][j]) for j in range(self.group)]) != 0:
                                    print('|---- From compartment ' + self.compartment_name[n] +
                                          ' to compartment ' + self.compartment_name[m] + ':')
                                    if self.__log_stream == 1:
                                        outcome.write('|---- From compartment ' + self.compartment_name[n] +
                                              ' to compartment ' + self.compartment_name[m] + ':\n')
                                    for j in range(self.group):
                                        print('|-------- ', bestsol[3][n][m][j])
                                        if self.__log_stream == 1:
                                            outcome.write('|-------- ')
                                            outcome.write(str(bestsol[3][n][m][j])+'\n')
                        print('|-----------------------------------')
                        if self.__log_stream == 1:
                            outcome.write('|-----------------------------------\n')
                            endSO = time.clock()
                            outcome.write('Computational time ' + str(endSO - startSO) + '\n')
                            outcome.close()
                        self.bestsol = bestsol
                        return bestsol
                    else:
                        (self.status, obj, Xopt, xopt, Xc) = solution
                        if obj < bestobj[-1]:
                            bestobj.append(obj)
                            bestsol = solution
                            print('|-- Objective values at all iterations:', bestobj)
                            print('|-- Current objective:', obj)
                            print('|-- Current x (Iteration ' + str(iter) + ' ) =')
                            if self.__log_stream == 1:
                                outcome.write('|-- Objective values at all iterations: '+ str(bestobj)+'\n')
                                outcome.write('|-- Current objective: '+ str(obj) + '\n')
                                outcome.write('|-- Current x (Iteration ' + str(iter) + ' ) =\n')
                            for n in self.compartment_name.keys():
                                for m in self.compartment_info[n]['tocpm'].keys():
                                    if sum([sum(xopt[n][m][j]) for j in range(self.group)]) != 0:
                                        print('|---- From compartment ' + self.compartment_name[n] +
                                              ' to compartment ' + self.compartment_name[m] + ':')
                                        if self.__log_stream == 1:
                                            outcome.write('|---- From compartment ' + self.compartment_name[n] +
                                                  ' to compartment ' + self.compartment_name[m] + ':\n')
                                        for j in range(self.group):
                                            print('|-------- ', xopt[n][m][j])
                                            if self.__log_stream == 1:
                                                outcome.write('|-------- ')
                                                outcome.write(str(xopt[n][m][j])+'\n')
                            print('|-----------------------------------')
                        else:
                            bestobj.append(obj)
                            print('|-- Convergence')
                            print('|-- Objective values at all iterations:', bestobj)
                            print('|-- Optimal objective:', bestsol[1])
                            print('|-- Best x =')
                            if self.__log_stream == 1:
                                outcome.write('|-- Convergence\n')
                                outcome.write('|-- Objective values at all iterations: '+ str(bestobj)+'\n')
                                outcome.write('|-- Optimal objective: ' + str(bestsol[1])+'\n')
                                outcome.write('|-- Best x =\n')
                            for n in self.compartment_name.keys():
                                for m in self.compartment_info[n]['tocpm'].keys():
                                    if sum([sum(bestsol[3][n][m][j]) for j in range(self.group)]) != 0:
                                        print('|---- From compartment ' + self.compartment_name[n] +
                                              ' to compartment ' + self.compartment_name[m] + ':')
                                        if self.__log_stream == 1:
                                            outcome.write('|---- From compartment ' + self.compartment_name[n] +
                                                  ' to compartment ' + self.compartment_name[m] + ':\n')
                                        for j in range(self.group):
                                            print('|-------- ', bestsol[3][n][m][j])
                                            if self.__log_stream == 1:
                                                outcome.write('|-------- ')
                                                outcome.write(str(bestsol[3][n][m][j]) + '\n')
                            print('|-----------------------------------')
                            if self.__log_stream == 1:
                                outcome.write('|-----------------------------------\n')
                                endSO = time.clock()
                                outcome.write('Computational time '+str(endSO -startSO) + '\n')
                                outcome.close()
                            self.bestsol = bestsol
                            return bestsol

        elif label == 'Robustness':
            if self.solver == 'gurobi':
                self.model.update()
            self.types = 'Robustness'
            if self.__approximation in ['ME','SF']:
                if self.__approximation == 'ME':
                    self.__set_bound_McCormickEnvelopes()
                self.dvar_conpartment()
                self.set_objectivetype(sense=self.objective_type)
                for i in self.objectiveterm:
                    self.set_objective(state=i[0], value=i[2])
                for i in self.con_var_para:
                    self.defineLinearConstraints(var=i[0], coef=i[1], sense=i[2], rhs=i[3], name=i[4], types=i[5])
                for i in self.con_var_para_qp:
                    self.defineQuadraticConstraints(varlp=i[0], coeflp=i[1], varqp=i[2], coefqp=i[3], sense=i[4],
                                                    rhs=i[5], name=i[6])
                if self.__approximation == 'SF':
                    self.set_step()
                    self.stepfunction(label='p')
                    self.stepfunction(label='q')
                elif self.__approximation == 'ME':
                    self.__McCormickEnvelopesfunction()
                self.__Dynamics()
                self.constraintgeneration()

                self.model.write('modelrobust_ME_('+str(target)+').lp')
                if self.solver == 'cplex':
                    self.model.write('modelrobust_ME_(' + str(target) + 'cplex).mps')
                elif self.solver == 'gurobi':
                    self.model.write('modelrobust_ME_(' + str(target) + 'gurobi).mps')
                solution = self.solvemodel(ct=ct,target=target)
                if isinstance(solution, (int, float)):
                    self.status = solution
                    return self.status
                else:
                    return solution
            elif self.__approximation == 'SO':
                if self.__log_stream == 1:
                    startSO = time.clock()
                    self.outcome = open(self.__log_file + '.txt', 'w')
                if self.__initialsol == {}:
                    self.set_approximation('ME')
                    sol = self.solve(label='Robustness', ct=ct, target=target)
                    self.set_approximation('SO')
                    if isinstance(sol, (int, float)):
                        self.status = sol
                        return self.status
                    else:
                        (self.status, bestobj, Xopt, xopt, Xc) = sol
                    label2 = 0
                else:
                    xopt = [[[[0 for t in range(self.Time)] for j in range(self.group)]
                             for m in self.compartment_info.keys()] for n in self.compartment_info.keys()]
                    for i in self.__initialsol.keys():
                        xopt[i[0]][i[1]] = self.__initialsol[i]
                    label2 = 1

                besttheta = [[1e15]]
                label1 = 0
                if self.solver == 'cplex':
                    self.model.linear_constraints.delete()
                    self.model.quadratic_constraints.delete()
                    self.model.variables.delete()
                elif self.solver =='gurobi':
                    self.model.remove(self.model.getConstrs())
                    self.model.remove(self.model.getQConstrs())
                    self.model.remove(self.model.getVars())
                pp = 0
                while 1:
                    pp += 1
                    # step 1. simulation and generate value of compartment and probability eta
                    self.x = xopt
                    Xopt, xopt, eta = self.DeterPrediction()

                    # step 2 generate or update the model
                    if label1 == 0:
                        self.dvar_conpartment()
                        self.set_objectivetype(sense=self.objective_type)
                        for i in self.objectiveterm:
                            self.set_objective(state=i[0], value=i[2])
                        if label2 == 1:
                            for i in self.con_var_para:
                                self.defineLinearConstraints(var=i[0], coef=i[1], sense=i[2], rhs=i[3],
                                                       name=i[4], types=i[5])
                            for i in self.con_var_para_qp:
                                self.defineQuadraticConstraints(varlp=i[0], coeflp=i[1], varqp=i[2], coefqp=i[3],
                                                                sense=i[4], rhs=i[5], name=i[6])
                    self.__DynamicsLinear(eta=eta,label=label1)
                    if label1 == 0:
                        self.constraintgeneration()
                    self.model.write('model_robustness_SO_' + str(pp) + '.lp')
                    self.model.write('model_robustness_SO_' + str(pp) + '.mps')
                    label1 = 1
                    solution = self.solvemodel(ct=ct, target=target)
                    theta = [self.robustconstraint[name][2] for name in self.robustconstraint.keys()]
                    theta.sort(reverse=True)

                    # step 4. stop condition
                    if isinstance(solution, (int, float)) and len(besttheta) == 1:
                        print('|-- Initial solution is infeasible for the optimization model...')
                        print('|-----------------------------------')
                        if self.__log_stream == 1:
                            self.outcome.write('|-- Initial solution is infeasible for the optimization model...\n')
                            self.outcome.write('|-----------------------------------\n')
                        besttheta.append(besttheta[-1])
                    elif isinstance(solution, (int, float)) and len(besttheta) > 1 \
                            and min([min(ii) for ii in besttheta]) >= self.theta * 10-0.01:
                        print('|-- Cannot generate a feasible solution, please check the initial solution '
                              'or model setting...')
                        print('|-----------------------------------')
                        if self.__log_stream == 1:
                            self.outcome.write('|-- Cannot generate a feasible solution, please check the initial solution '
                                  'or model setting...\n')
                            self.outcome.write('|-----------------------------------\n')
                            endSO = time.clock()
                            self.outcome.write('Computational time ' + str(endSO - startSO) + '\n')
                            self.outcome.close()
                        return 0
                    elif isinstance(solution, (int, float)) and len(besttheta) > 1 \
                            and min([min(ii) for ii in besttheta]) < self.theta * 10-0.01:
                        print('|-- Optimal model at current iteration is infeasible but a feasible solution is '
                              'generated at the previous iteration')
                        print('|-- Objective theta at all iterations:')
                        if self.__log_stream == 1:
                            self.outcome.write('|-- Optimal model at current iteration is infeasible but a feasible solution is '
                                  'generated at the previous iteration\n')
                            self.outcome.write('|-- Objective theta at all iterations:\n')
                        for jj in besttheta:
                            print('|---- ', jj)
                            if self.__log_stream == 1:
                                self.outcome.write('|---- '+str(jj)+'\n')
                        print('|-- Current theta (Iteration ' + str(pp) + ' ):', theta)
                        print('|-- Best x =')
                        if self.__log_stream == 1:
                            self.outcome.write('|-- Current theta (Iteration ' + str(pp) + ' ): '+str(theta)+'\n')
                            self.outcome.write('|-- Best x =\n')
                        for n in self.compartment_name.keys():
                            for m in self.compartment_info[n]['tocpm'].keys():
                                if sum([sum(bestsol[3][n][m][j]) for j in range(self.group)]) != 0:
                                    print('|---- From compartment ' + self.compartment_name[n] +
                                          ' to compartment ' + self.compartment_name[m] + ':')
                                    if self.__log_stream == 1:
                                        self.outcome.write('|---- From compartment ' + self.compartment_name[n] +
                                              ' to compartment ' + self.compartment_name[m] + ':\n')
                                    for j in range(self.group):
                                        print('|-------- ', bestsol[3][n][m][j])
                                        if self.__log_stream == 1:
                                            self.outcome.write('|-------- '+ str(bestsol[3][n][m][j])+'\n')
                        print('|-----------------------------------')
                        if self.__log_stream == 1:
                            self.outcome.write('|-----------------------------------\n')
                            endSO = time.clock()
                            self.outcome.write('Computational time ' + str(endSO - startSO) + '\n')
                            self.outcome.close()
                        self.bestsol = bestsol
                        return bestsol
                    else:
                        (self.status, obj, Xopt, xopt, Xc) = solution
                        if theta < besttheta[-1]:
                            besttheta.append(theta)
                            bestsol = solution
                            print('|-- Objective theta at all iterations:')
                            if self.__log_stream == 1:
                                self.outcome.write('|-- Objective theta at all iterations:\n')
                            for jj in besttheta:
                                print('|---- ', jj)
                                if self.__log_stream == 1:
                                    self.outcome.write('|---- ' + str(jj)+'\n')
                            print('|-- Current theta (Iteration ' + str(pp) + ' ):', theta)
                            print('|-- Current x  (Iteration ' + str(pp) + ' ) =')
                            if self.__log_stream == 1:
                                self.outcome.write('|-- Current theta (Iteration ' + str(pp) + ' ): ' + str(theta) + '\n')
                                self.outcome.write('|-- Current x  (Iteration ' + str(pp) + ' ) =\n')
                            for n in self.compartment_name.keys():
                                for m in self.compartment_info[n]['tocpm'].keys():
                                    if sum([sum(xopt[n][m][j]) for j in range(self.group)]) != 0:
                                        print('|---- From compartment ' + self.compartment_name[n] +
                                              ' to compartment ' + self.compartment_name[m] + ':')
                                        if self.__log_stream == 1:
                                            self.outcome.write('|---- From compartment ' + self.compartment_name[n] +
                                                        ' to compartment ' + self.compartment_name[m] + ':\n')
                                        for j in range(self.group):
                                            print('|-------- ', xopt[n][m][j])
                                            if self.__log_stream == 1:
                                                self.outcome.write('|-------- ' + str(xopt[n][m][j])+'\n')
                            print('|-----------------------------------')
                            if self.__log_stream == 1:
                                self.outcome.write('|-----------------------------------\n')
                        else:
                            besttheta.append(theta)
                            print('|-- Convergence')
                            print('|-- Objective theta at all iterations:')
                            for jj in besttheta:
                                print('|---- ', jj)
                            print('|-- Current theta (Iteration ' + str(pp) + ' ):', theta)
                            print('|-- Best x =')
                            if self.__log_stream == 1:
                                self.outcome.write('|-- Convergence\n')
                                self.outcome.write('|-- Objective theta at all iterations:\n')
                            for jj in besttheta:
                                print('|---- ', jj)
                                if self.__log_stream == 1:
                                    self.outcome.write('|---- ' + str(jj)+'\n')
                            if self.__log_stream == 1:
                                self.outcome.write('|-- Current theta (Iteration ' + str(pp) + ' ): ' + str(theta) + '\n')
                                self.outcome.write('|-- Best x =\n')
                            for n in self.compartment_name.keys():
                                for m in self.compartment_info[n]['tocpm'].keys():
                                    if sum([sum(bestsol[3][n][m][j]) for j in range(self.group)]) != 0:
                                        print('|---- From compartment ' + self.compartment_name[n] +
                                              ' to compartment ' + self.compartment_name[m] + ':')
                                        if self.__log_stream == 1:
                                            self.outcome.write('|---- From compartment ' + self.compartment_name[n] +
                                                          ' to compartment ' + self.compartment_name[m] + ':\n')
                                        for j in range(self.group):
                                            print('|-------- ', bestsol[3][n][m][j])
                                            if self.__log_stream == 1:
                                                self.outcome.write('|-------- ' + str(bestsol[3][n][m][j]) + '\n')
                            print('|-----------------------------------')
                            if self.__log_stream == 1:
                                self.outcome.write('|-----------------------------------\n')
                                endSO = time.clock()
                                self.outcome.write('Computational time ' + str(endSO - startSO) + '\n')
                                self.outcome.close()
                            self.bestsol = bestsol
                            return bestsol

    def defineLinearConstraints(self, var= [], coef = [], sense=None, rhs=None, name=None, types = 'Expectation'):
        if types == 'Robustness' and self.types == 'Expectation':
            types = self.types

        if name == None:
            raise ValueError('Please define the name of constraints')
        elif name in self.constraint.keys():
            self.constraint[name][4] = types
            if sense in ['L','G','E']:
                self.constraint[name][2] = sense
            if rhs != None and isinstance(rhs, (float,int)):
                self.constraint[name][3] = rhs

            if (var == [] or coef == []) and rhs == None:
                raise ValueError('Please define variable and coefficient in constraint' + name)
            elif (var == [] or coef == []) and rhs != None:
                pass
            else:
                if isinstance(var, str) and isinstance(coef, (float,int)):
                    if var in self.constraint[name][0]:
                        index = self.constraint[name][0].index(var)
                        self.constraint[name][1][index] = coef
                    else:
                        self.constraint[name][0].append(var)
                        self.constraint[name][1].append(coef)
                elif isinstance(var, list) and isinstance(coef, list) and len(var) == len(coef):
                    for i in range(len(var)):
                        if var[i] in self.constraint[name][0]:
                            index = self.constraint[name][0].index(var[i])
                            self.constraint[name][1][index] = coef[i]
                        else:
                            self.constraint[name][0].append(var[i])
                            self.constraint[name][1].append(coef[i])
        else:
            if isinstance(var, str) and isinstance(coef, (float, int)):
                self.constraint.update({name: [[var], [coef], sense, rhs, types]})
            elif isinstance(var, list) and isinstance(coef, list) and len(var) == len(coef):
                self.constraint.update({name: [var, coef, sense, rhs, types]})
            else:
                raise ValueError('Domains of variable and coefficient are inconsistent')

    def defineQuadraticConstraints(self, varlp=[], coeflp=[], varqp=[], coefqp=[], sense=None, rhs=None, name=None):
        if name is None:
            raise ValueError('Please define the name of constraints')
        elif name in self.quadraticconstraint.keys():
            if sense in ['L','G']:
                self.quadraticconstraint[name][4] = sense
            if rhs is not None and isinstance(rhs, (float,int)):
                self.quadraticconstraint[name][5] = rhs

            if isinstance(varlp, str):
                varlp = [varlp]
            if isinstance(coeflp, (float,int)):
                coeflp = [coeflp]
            if isinstance(varqp, str):
                varqp = [varqp]
            if isinstance(coefqp, (float,int)):
                coefqp = [coefqp]

            if len(varqp) != len(coefqp) or len(varlp) != len(coeflp):
                raise ValueError('Domains of variable and coefficient are inconsistent')
            elif varqp == [] or coefqp ==[]:
                raise ValueError('Please define variable and coefficient in '
                                 'the quadratic term of constraint' + name)
            else:
                for i in range(len(varlp)):
                    if varlp[i] in self.quadraticconstraint[name][0]:
                        index = self.quadraticconstraint[name][0].index(varlp[i])
                        self.quadraticconstraint[name][1][index] += coeflp[i]
                    else:
                        self.quadraticconstraint[name][0].append(varlp[i])
                        self.quadraticconstraint[name][1].append(coeflp[i])

                for i in range(len(varqp)):
                    if varqp[i] in self.quadraticconstraint[name][0]:
                        index = self.quadraticconstraint[name][2].index(varqp[i])
                        self.quadraticconstraint[name][3][index] += coefqp[i]
                    else:
                        self.quadraticconstraint[name][2].append(varqp[i])
                        self.quadraticconstraint[name][3].append(coefqp[i])
        else:
            self.quadraticconstraint.update({name: [varlp, coeflp, varqp, coefqp, sense, rhs]})

    def constraintgeneration(self):
        if self.solver == 'cplex':
            self.dvar_robust = {}
            self.dcoef_robust = {}
            for name in self.constraint.keys():
                if self.constraint[name][-1] == 'Expectation':
                    if self.constraint[name][2] in ['L', 'E']:
                        self.model.linear_constraints.add(
                            lin_expr=[cplex.SparsePair(self.constraint[name][0], self.constraint[name][1])],
                            senses=[self.constraint[name][2]],
                            rhs=[self.constraint[name][3]],
                            names=[name])
                    else:
                        self.model.linear_constraints.add(
                            lin_expr=[cplex.SparsePair(self.constraint[name][0], [-self.constraint[name][1][ii]
                                                        for ii in range(len(self.constraint[name][1]))])],
                            senses=['L'],
                            rhs=[-self.constraint[name][3]],
                            names=[name])
                elif self.constraint[name][-1] == 'Robustness':
                    self.dvar_robust.update({name: []})
                    self.dcoef_robust.update({name: []})
                    for i in range(len(self.constraint[name][0])):
                        if self.constraint[name][0][i][0] == 'X':
                            self.dvar_robust[name].append('Xdagger' + self.constraint[name][0][i][1:])
                            self.dcoef_robust[name].append(self.constraint[name][1][i])

                    if self.constraint[name][2] in ['L','E']:
                        self.model.linear_constraints.add(
                            lin_expr=[cplex.SparsePair(self.constraint[name][0] + self.dvar_robust[name],
                                                       self.constraint[name][1]
                                                       + [self.dcoef_robust[name][ii] ** 2 * 0.5/self.theta
                                                        for ii in range(len(self.dcoef_robust[name]))])],
                            senses=[self.constraint[name][2]],
                            rhs=[self.constraint[name][3]],
                            names=[name])
                    else:
                        self.model.linear_constraints.add(
                            lin_expr=[cplex.SparsePair(self.constraint[name][0] + self.dvar_robust[name],
                                                        [-self.constraint[name][1][ii]
                                                         for ii in range(len(self.constraint[name][1]))]
                                                       + [self.dcoef_robust[name][ii] ** 2 * 0.5/self.theta
                                                          for ii in range(len(self.dcoef_robust[name]))])],
                            senses=['L'],
                            rhs=[-self.constraint[name][3]],
                            names=[name])
                    self.robustconstraint.update({name: [0, self.theta, self.theta*10, 0]})

            for name in self.quadraticconstraint.keys():
                if self.quadraticconstraint[name][4] == 'L':
                    self.model.quadratic_constraints.add(
                        lin_expr=[self.quadraticconstraint[name][0],self.quadraticconstraint[name][1]],
                        quad_expr=[self.quadraticconstraint[name][2],self.quadraticconstraint[name][2],
                                   self.quadraticconstraint[name][3]],
                        sense=self.quadraticconstraint[name][4],
                        rhs=self.quadraticconstraint[name][5],name=name)
                else:
                    self.model.quadratic_constraints.add(
                        lin_expr=[self.quadraticconstraint[name][0],
                                  [-self.quadraticconstraint[name][1][ii]
                                   for ii in range(len(self.quadraticconstraint[name][1]))]],
                        quad_expr=[self.quadraticconstraint[name][2],self.quadraticconstraint[name][2],
                                   [-self.quadraticconstraint[name][3][ii] for ii in
                                    range(len(self.quadraticconstraint[name][3]))]],
                        sense='L',
                        rhs=-self.quadraticconstraint[name][5],name=name)

        elif self.solver == 'gurobi':
            self.dvar_robust = {}
            self.dcoef_robust = {}
            for name in self.constraint.keys():
                if self.constraint[name][-1] == 'Expectation':
                    if self.constraint[name][2] in ['L']:
                        self.model.addConstr(
                            (gp.quicksum(self.model.getVarByName(self.constraint[name][0][ii]) * self.constraint[name][1][ii]
                                         for ii in range(len(self.constraint[name][0]))) <= self.constraint[name][3]),
                            name)
                    elif self.constraint[name][2] in ['E']:
                        self.model.addConstr(
                            (gp.quicksum(self.model.getVarByName(self.constraint[name][0][ii]) * self.constraint[name][1][ii]
                                         for ii in range(len(self.constraint[name][0]))) == self.constraint[name][3]),
                            name)
                    else:
                        self.model.addConstr(
                            (gp.quicksum(-self.model.getVarByName(self.constraint[name][0][ii]) * self.constraint[name][1][ii]
                                         for ii in range(len(self.constraint[name][0]))) <= -self.constraint[name][3]),
                            name)
                elif self.constraint[name][-1] == 'Robustness':
                    print(name)
                    self.dvar_robust.update({name: []})
                    self.dcoef_robust.update({name: []})
                    for i in range(len(self.constraint[name][0])):
                        if self.constraint[name][0][i][0] == 'X':
                            self.dvar_robust[name].append('Xdagger' + self.constraint[name][0][i][1:])
                            self.dcoef_robust[name].append(self.constraint[name][1][i])

                    if self.constraint[name][2] in ['L']:
                        self.model.addConstr(
                            (gp.quicksum(self.model.getVarByName(self.constraint[name][0][ii]) * self.constraint[name][1][ii]
                                         for ii in range(len(self.constraint[name][0])))
                             + gp.quicksum(self.model.getVarByName(self.dvar_robust[name][ii])
                                           * self.dcoef_robust[name][ii] ** 2 * 0.5 / self.theta
                                           for ii in range(len(self.dcoef_robust[name]))) <= self.constraint[name][3]),
                            name)
                    elif self.constraint[name][2] in ['E']:
                        self.model.addConstr(
                            (gp.quicksum(self.model.getVarByName(self.constraint[name][0][ii]) * self.constraint[name][1][ii]
                                         for ii in range(len(self.constraint[name][0])))
                             + gp.quicksum(self.model.getVarByName(self.dvar_robust[name][ii])
                                           * self.dcoef_robust[name][ii] ** 2 * 0.5 / self.theta
                                           for ii in range(len(self.dcoef_robust[name]))) == self.constraint[name][3]),
                            name)
                    else:
                        self.model.addConstr(
                            (- gp.quicksum(self.model.getVarByName(self.constraint[name][0][ii]) * self.constraint[name][1][ii]
                                         for ii in range(len(self.constraint[name][0])))
                             + gp.quicksum(self.model.getVarByName(self.dvar_robust[name][ii])
                                           * self.dcoef_robust[name][ii] ** 2 * 0.5 / self.theta
                                           for ii in range(len(self.dcoef_robust[name]))) <= -self.constraint[name][3]),
                            name)
                    self.robustconstraint.update({name: [0, self.theta, self.theta*10, 0]})

            self.model.update()
            for name in self.quadraticconstraint.keys():
                if self.quadraticconstraint[name][4] == 'L':
                    self.model.addConstr(
                        gp.quicksum(self.model.getVarByName(self.quadraticconstraint[name][2][ii])
                                     * self.model.getVarByName(self.quadraticconstraint[name][2][ii])
                                     * self.quadraticconstraint[name][3][ii]
                                       for ii in range(len(self.quadraticconstraint[name][3])))
                         + gp.quicksum(self.model.getVarByName(self.quadraticconstraint[name][0][ii])
                                       * self.quadraticconstraint[name][1][ii]
                                       for ii in range(len(self.quadraticconstraint[name][1]))) <= self.quadraticconstraint[name][5],
                        name)
                else:
                    self.model.addConstr(
                        -gp.quicksum(self.model.getVarByName(self.quadraticconstraint[name][2][ii])**2 * self.quadraticconstraint[name][3][ii]
                                       for ii in range(len(self.quadraticconstraint[name][3])))
                         - gp.quicksum(self.model.getVarByName(self.quadraticconstraint[name][0][ii])
                                       * self.quadraticconstraint[name][1][ii]
                                       for ii in range(len(self.quadraticconstraint[name][1]))) <= -self.quadraticconstraint[name][5],
                        str(name))
        else:
            raise ValueError('Current version only supports Cplex and Gurobi as solver')

    '''=============== Step 10. Define objective  ====================='''
    def set_objectivetype(self,sense='min'):
        try:
            if sense == 'min':
                if self.solver == 'cplex':
                    self.model.objective.set_sense(self.model.objective.sense.minimize)
                elif self.solver == 'gurobi':
                    self.model.setAttr('ModelSense', GRB.MINIMIZE)
                self.objective_type = sense
            elif sense == 'max':
                if self.solver == 'cplex':
                    self.model.objective.set_sense(self.model.objective.sense.minimize)
                elif self.solver == 'gurobi':
                    self.model.setAttr('ModelSense', GRB.MINIMIZE)
                self.objective_type = sense
            else:
                raise ValueError('Sense of objective function should be min or max')
        except:
            self.objective_type = sense

    def set_objective(self, state=None, value=None, option=None):
        if option == 'clear':
            self.objectiveterm = []
        elif state != None and value!=None and option==None:
            try:
                if self.solver == 'cplex':
                    if isinstance(state, str):
                        if self.objective_type == 'max':
                            self.model.objective.set_linear([(state, -value)])
                            self.expobj.append([state, -value])
                        else:
                            self.model.objective.set_linear([(state, value)])
                            self.expobj.append([state, value])
                    elif len(state) == 3:
                        if isinstance(state[0], str):
                            n = get_keys(self.compartment_name, state[0])[0]
                            state = (n,state[1],state[2])
                        (n,j,t) = state
                        if self.objective_type == 'max':
                            self.model.objective.set_linear([("X." + str(n) + "." + str(j) + "." + str(t), -value)])
                            self.expobj.append(["X." + str(n) + "." + str(j) + "." + str(t), -value])
                        else:
                            self.model.objective.set_linear([("X." + str(n) + "." + str(j) + "." + str(t), value)])
                            self.expobj.append(["X." + str(n) + "." + str(j) + "." + str(t), value])
                        if self.types == 'Robustness':
                            self.model.objective.set_linear([("Xdagger." + str(n) + "." + str(j) + "." + str(t),
                                                              0.5 * value**2 / self.theta)])
                            self.objtype = 1
                            self.robustobj.append(["Xdagger." + str(n) + "." + str(j) + "." + str(t), value])
                        else:
                            self.objtype = 0
                    elif len(state) == 4:
                        if isinstance(state[0], str):
                            n = get_keys(self.compartment_name, state[0])[0]
                            state = (n,state[1],state[2],state[3])
                        if isinstance(state[1], str):
                            m = get_keys(self.compartment_name, state[1])[0]
                            state = (state[0],m,state[2],state[3])
                        (n, m, j, t) = state
                        if (j, t) in self.compartment_info[n]['tocpm'][m]['x']:
                            if self.objective_type == 'max':
                                self.model.objective.set_linear(
                                    [("x." + str(n) + "." + str(m) + "." + str(j) + "." + str(t), -value)])
                            else:
                                self.model.objective.set_linear(
                                    [("x." + str(n) + "." + str(m) + "." + str(j) + "." + str(t), value)])
                        else:
                            raise ValueError('Decision flow is not defined')
                    else:
                        raise ValueError('length of state should be three or four, or state should be a string')

                elif self.solver == 'gurobi':
                    self.model.update()
                    if isinstance(state, str):
                        if self.objective_type == 'max':
                            self.model.getVarByName(state).Obj = -value
                            self.expobj.append([state, -value])
                        else:
                            self.model.getVarByName(state).Obj = value
                            self.expobj.append([state, value])
                    elif len(state) == 3:
                        if isinstance(state[0], str):
                            n = get_keys(self.compartment_name, state[0])[0]
                            state = (n, state[1], state[2])
                        (n, j, t) = state
                        if self.objective_type == 'max':
                            self.model.getVarByName('X.'+str(n)+'.'+str(j)+'.'+str(t)).Obj = -value
                            self.expobj.append(["X." + str(n) + "." + str(j) + "." + str(t), -value])
                        else:
                            self.model.getVarByName('X.' + str(n) + '.' + str(j) + '.' + str(t)).Obj = value
                            self.expobj.append(["X." + str(n) + "." + str(j) + "." + str(t), value])
                        if self.types == 'Robustness':
                            self.model.getVarByName('Xdagger.' + str(n) + '.' + str(j) + '.' + str(t)).Obj \
                                = 0.5 * value**2 / self.theta
                            self.objtype = 1
                            self.robustobj.append(["Xdagger." + str(n) + "." + str(j) + "." + str(t), value])
                        else:
                            self.objtype = 0
                    elif len(state) == 4:
                        if isinstance(state[0], str):
                            n = get_keys(self.compartment_name, state[0])[0]
                            state = (n,state[1],state[2],state[3])
                        if isinstance(state[1], str):
                            m = get_keys(self.compartment_name, state[1])[0]
                            state = (state[0],m,state[2],state[3])
                        (n,m,j,t) = state
                        if (j,t) in self.compartment_info[n]['tocpm'][m]['x']:
                            if self.objective_type == 'max':
                                self.model.getVarByName(
                                    'x.' + str(n) + '.' + str(m) + '.' + str(j) + '.' + str(t)).Obj = -value
                            else:
                                self.model.getVarByName(
                                    'x.' + str(n) + '.' + str(m) + '.' + str(j) + '.' + str(t)).Obj = value
                        else:
                            raise ValueError('Decision flow is not defined')
                    else:
                        raise ValueError('length of state should be three or four, or state should be a string')
                else:
                    raise ValueError('Current version only supports Cplex and Gurobi as solvers')
            except:
                self.objectiveterm.append([state, 0, value])
        else:
            raise ValueError('loss of input')

    '''=============== Step 11. Solve and output  ====================='''

    def solvemodel(self, ct=None, target=None):
        if self.solver == 'cplex':
            sol = self.__solvemodel_cplex(ct=ct, target=target)
        elif self.solver == 'gurobi':
            sol = self.__solvemodel_gurobi(ct=ct, target=target)
        return sol
    def __solvemodel_cplex(self, ct=None, target=None):
        if self.types == 'Expectation':
            X = [[[0 for t in range(self.Time)] for j in range(self.group)]
                 for n in range(max(self.compartment_name.keys()) + 1)]
            x = [[[[0 for t in range(self.Time)] for j in range(self.group)]
                  for m in range(max(self.compartment_name.keys()) + 1)]
                 for n in range(max(self.compartment_name.keys()) + 1)]
            Xc = {}
            # self.model.set_results_stream(None)
            # self.model.set_log_stream(None)
            if ct != None:
                self.model.parameters.timelimit.set(ct)
            self.model.solve()
            try:
                bestobj = self.model.solution.get_objective_value()
                for n in self.compartment_name.keys():
                    X[n] = [[round(self.model.solution.get_values('X.' + str(n) + '.' + str(j) + '.' + str(t)), 6)
                             for t in range(self.Time)] for j in range(self.group)]
                    for m in self.compartment_info[n]['tocpm'].keys():
                        x[n][m] = [
                            [self.model.solution.get_values('x.' + str(n) + '.' + str(m) + '.' + str(j) + '.' + str(t))
                             if (j, t) in self.compartment_info[n]['tocpm'][m]['x'] else 0
                             for t in range(self.Time)] for j in range(self.group)]
                for n in self.custom_var.keys():
                    Xc.update({n: self.model.solution.get_values(n)})
                sol = (self.model.solution.get_status(), bestobj, X, x, Xc)
                return sol
            except:
                return 0
        elif self.types == 'Robustness' and target != None:
            # self.model.set_results_stream(None)
            # self.model.set_log_stream(None)
            T0 = []
            if self.objtype:
                self.robustconstraint.update({'obj': [0, self.theta, self.theta * 10, 0]})
            T1 = list(self.robustconstraint.keys())

            # print('T1=',T1)
            # print(self.robustconstraint)
            T1.sort()
            p = 0
            print('\n============================================')
            while T1 != []:
                if p > 0:
                    for name in T0:
                        self.robustconstraint[name][0] = self.robustconstraint[name][2]
                        self.robustconstraint[name][1] = self.robustconstraint[name][2]
                        self.robustconstraint[name][3] = 0
                    for name in T1:
                        self.robustconstraint[name][0] = 0
                        self.robustconstraint[name][1] = self.robustconstraint[name][2]
                        self.robustconstraint[name][3] = 0
                    p += 1
                else:
                    for name in T1:
                        self.robustconstraint[name] = [0, self.theta, self.theta * 10, 0]
                    p += 1
                num_infeasible = 0
                gaptheta = 1e10
                pp = 0
                print('+- Begin round ' + str(p))
                if self.__log_stream == 1:
                    self.outcome.write('+- Begin round ' + str(p) + '\n')
                while max([self.robustconstraint[name][2] / (self.robustconstraint[name][0] + 0.0001) for name in
                           T1]) > 1.01:
                    pp += 1
                    for name in T1:
                        if name != 'obj':
                            self.model.linear_constraints.set_coefficients([(name, self.dvar_robust[name][i],
                                                                             self.dcoef_robust[name][i] ** 2 * 0.5 /
                                                                             self.robustconstraint[name][1])
                                                                            for i in range(len(self.dcoef_robust[name]))
                                                                            if self.dcoef_robust[name][i] != 0])
                        else:
                            for ii in range(len(self.robustobj)):
                                self.model.objective.set_linear([(self.robustobj[ii][0],
                                                                  0.5 * self.robustobj[ii][1] ** 2 /
                                                                  self.robustconstraint[name][1])])
                    if ct != None:
                        self.model.parameters.timelimit.set(ct)
                    print('|- Iteration ' + str(pp) + ' of round ' + str(p) + ':')
                    if self.__log_stream == 1:
                        self.outcome.write('|- Iteration ' + str(pp) + ' of round ' + str(p) + ':\n')
                    self.model.solve()
                    try:
                        objective_value = self.model.solution.get_objective_value()
                    except:
                        objective_value = 1e+30
                    if objective_value < target:
                        num_infeasible = 0
                        time.sleep(5)
                        self.model.write('modelrobust_SO_' + str(round(target, 2)) + '.mps')
                        status = self.model.solution.get_status()
                        bestobj = self.model.solution.get_objective_value()
                        X = [[[0 for t in range(self.Time)] for j in range(self.group)]
                             for n in range(max(self.compartment_name.keys()) + 1)]
                        Xdagger = [[[0 for t in range(self.Time)] for j in range(self.group)]
                                   for n in range(max(self.compartment_name.keys()) + 1)]
                        x = [[[[0 for t in range(self.Time)] for j in range(self.group)]
                              for m in range(max(self.compartment_name.keys()) + 1)]
                             for n in range(max(self.compartment_name.keys()) + 1)]
                        Xc = {}
                        for n in self.compartment_name.keys():
                            X[n] = [[self.model.solution.get_values('X.' + str(n) + '.' + str(j) + '.' + str(t))
                                     for t in range(self.Time)] for j in range(self.group)]
                            Xdagger[n] = [
                                [self.model.solution.get_values('Xdagger.' + str(n) + '.' + str(j) + '.' + str(t))
                                 for t in range(self.Time)] for j in range(self.group)]
                            for m in self.compartment_info[n]['tocpm'].keys():
                                x[n][m] = [[round(self.model.solution.get_values(
                                    'x.' + str(n) + '.' + str(m) + '.' + str(j) + '.' + str(t)), 6)
                                            if (j, t) in self.compartment_info[n]['tocpm'][m]['x'] else 0
                                            for t in range(self.Time)] for j in range(self.group)]
                        for n in self.custom_var.keys():
                            Xc.update({n: self.model.solution.get_values(n)})
                        for name in T1:
                            if name != 'obj':
                                drhs = self.model.linear_constraints.get_rhs(name)

                                dvar1 = self.model.solution.get_values(self.constraint[name][0])
                                dcoef1 = self.model.linear_constraints.get_coefficients(
                                    [(name, iii) for iii in self.constraint[name][0]])
                                lhs1 = sum([dvar1[i] * dcoef1[i] for i in range(len(dvar1))])

                                dvar2 = self.model.solution.get_values(self.dvar_robust[name])
                                dcoef2 = self.model.linear_constraints.get_coefficients(
                                    [(name, iii) for iii in self.dvar_robust[name]])
                                lhs2 = sum([dvar2[i] * dcoef2[i] for i in range(len(dvar2))])
                                if lhs1 + lhs2 < 0.999 * drhs:
                                    if lhs2 != 0:
                                        gaptheta = min(gaptheta, (drhs - lhs1) / lhs2 - 0.1)
                                    else:
                                        gaptheta = 0
                                else:
                                    self.robustconstraint[name][3] = 1
                                    gaptheta = 0
                            else:
                                if self.model.solution.get_objective_value() < 0.999 * target:
                                    dvar3 = self.model.solution.get_values(
                                        [self.expobj[iii][0] for iii in range(len(self.expobj))])
                                    dcoef3 = self.model.objective.get_linear(
                                        [self.expobj[iii][0] for iii in range(len(self.expobj))])
                                    dvar4 = self.model.solution.get_values(
                                        [self.robustobj[iii][0] for iii in range(len(self.robustobj))])
                                    dcoef4 = self.model.objective.get_linear(
                                        [self.robustobj[iii][0] for iii in range(len(self.robustobj))])
                                    lhs3 = sum([dvar3[iii] * dcoef3[iii] for iii in range(len(dvar3))])
                                    lhs4 = sum([dvar4[iii] * dcoef4[iii] for iii in range(len(dvar4))])
                                    if lhs4 != 0:
                                        gaptheta = min(gaptheta, (target - lhs3) / lhs4 - 0.1)
                                    else:
                                        gaptheta = 0
                                else:
                                    self.robustconstraint[name][3] = 1
                                    gaptheta = 0
                        print('|-- Feasible ',
                              [(name, self.robustconstraint[name]) for name in self.robustconstraint.keys()])
                        if self.__log_stream == 1:
                            self.outcome.write('|-- Feasible ' + str([(name, self.robustconstraint[name])
                                                                      for name in self.robustconstraint.keys()]) + '\n')
                        if gaptheta <= 1:
                            for name in T1:
                                self.robustconstraint[name][2] = self.robustconstraint[name][1]
                                self.robustconstraint[name][1] = (self.robustconstraint[name][0]
                                                                  + self.robustconstraint[name][1]) / 2.0
                        else:
                            for name in T1:
                                self.robustconstraint[name][2] = self.robustconstraint[name][1]
                                self.robustconstraint[name][1] = self.robustconstraint[name][1] / gaptheta
                            gaptheta = 0
                        print('|-- Current objective and target:', self.model.solution.get_objective_value(), target)
                        print('|-- Current x =')
                        if self.__log_stream == 1:
                            self.outcome.write('|-- Current objective and target: '
                                               + str(self.model.solution.get_objective_value()) + ',' + str(
                                target) + '\n')
                            self.outcome.write('|-- Current x =\n')
                        for n in self.compartment_name.keys():
                            for m in self.compartment_info[n]['tocpm'].keys():
                                if sum([sum(x[n][m][j]) for j in range(self.group)]) != 0:
                                    print('|---- From compartment ' + self.compartment_name[n] +
                                          ' to compartment ' + self.compartment_name[m] + ':')
                                    for j in range(self.group):
                                        print('|-------- ', x[n][m][j])
                                    if self.__log_stream == 1:
                                        self.outcome.write('|---- From compartment ' + self.compartment_name[n] +
                                                           ' to compartment ' + self.compartment_name[m] + ':\n')
                                        for j in range(self.group):
                                            self.outcome.write('|-------- ' + str(x[n][m][j]) + '\n')
                        print('|--------------------------------------------')
                        if self.__log_stream == 1:
                            self.outcome.write('|--------------------------------------------\n')
                    else:
                        if num_infeasible < 3:
                            for name in T1:
                                self.robustconstraint[name][0] = self.robustconstraint[name][1]
                                self.robustconstraint[name][1] = (self.robustconstraint[name][2]
                                                                  + self.robustconstraint[name][1]) / 2.0
                            num_infeasible += 1
                        else:
                            for name in T1:
                                self.robustconstraint[name][0] = self.robustconstraint[name][1]
                                self.robustconstraint[name][1] = max(self.robustconstraint[name][2] - 0.1,
                                                                     (self.robustconstraint[name][2] +
                                                                      self.robustconstraint[name][1]) / 2.0)
                            num_infeasible = 0
                        print('|-- Infeasible ',
                              [(name, self.robustconstraint[name]) for name in self.robustconstraint.keys()])
                        print('|--------------------------------------------')
                        if self.__log_stream == 1:
                            self.outcome.write('|-- Infeasible '
                                               + str(
                                [(name, self.robustconstraint[name]) for name in self.robustconstraint.keys()])
                                               + '\n')
                            self.outcome.write('|--------------------------------------------\n')
                if sum([self.robustconstraint[name][3] for name in self.robustconstraint.keys()]) == 0:
                    T1 = []
                    T0 = list(self.robustconstraint.keys())
                else:
                    for name in T1:
                        if self.robustconstraint[name][3] == 1:
                            T0.append(name)
                            T1.remove(name)
                print('+- Now sets T0 and T1 are')
                print('|- T1 = ', T1)
                print('|- T0 = ', T0)
                print('|--------------------------------------------')
                if self.__log_stream == 1:
                    self.outcome.write('+- Now sets T0 and T1 are\n')
                    self.outcome.write('|- T1 = ' + str(T1) + '\n')
                    self.outcome.write('|- T0 = ' + str(T0) + '\n')
                    self.outcome.write('|--------------------------------------------\n')
            try:
                try:
                    time.sleep(5)
                    self.model = cplex.Cplex('modelrobust_SO_' + str(round(target, 2)) + '.mps')
                    self.model.set_results_stream(None)
                    self.model.set_log_stream(None)
                    self.model.solve()
                    try:
                        objective_value = self.model.solution.get_objective_value()
                    except:
                        objective_value = 1e+30
                    if objective_value < bestobj:
                        print('|-- Optimal objective value:', self.model.solution.get_objective_value(), bestobj)
                        if self.__log_stream == 1:
                            self.outcome.write('|-- Optimal objective value: '
                                               + str(self.model.solution.get_objective_value()) + '\n')
                        status = self.model.solution.get_status()
                        bestobj = self.model.solution.get_objective_value()
                        X = [[[0 for t in range(self.Time)] for j in range(self.group)]
                             for n in range(max(self.compartment_name.keys()) + 1)]
                        x = [[[[0 for t in range(self.Time)] for j in range(self.group)]
                              for m in range(max(self.compartment_name.keys()) + 1)]
                             for n in range(max(self.compartment_name.keys()) + 1)]
                        Xc = {}
                        print('|-- Optimal x=')
                        if self.__log_stream == 1:
                            self.outcome.write('|-- Optimal x=\n')
                        for n in self.compartment_name.keys():
                            X[n] = [
                                [round(self.model.solution.get_values('X.' + str(n) + '.' + str(j) + '.' + str(t)), 6)
                                 for t in range(self.Time)] for j in range(self.group)]
                            for m in self.compartment_info[n]['tocpm'].keys():
                                x[n][m] = [[round(
                                    self.model.solution.get_values(
                                        'x.' + str(n) + '.' + str(m) + '.' + str(j) + '.' + str(t)),
                                    6)
                                            if (j, t) in self.compartment_info[n]['tocpm'][m]['x'] else 0
                                            for t in range(self.Time)] for j in range(self.group)]
                                if sum([sum(x[n][m][j]) for j in range(self.group)]) != 0:
                                    print('|---- From compartment ' + self.compartment_name[n] +
                                          ' to compartment ' + self.compartment_name[m] + ':')
                                    for j in range(self.group):
                                        print('|-------- ', x[n][m][j])
                                    if self.__log_stream == 1:
                                        self.outcome.write('|---- From compartment ' + self.compartment_name[n] +
                                                           ' to compartment ' + self.compartment_name[m] + ':\n')
                                        for j in range(self.group):
                                            self.outcome.write('|-------- ' + str(x[n][m][j]) + '\n')
                        for n in self.custom_var.keys():
                            Xc.update({n: self.model.solution.get_values(n)})
                        print('|============================================\n\n')
                        if self.__log_stream == 1:
                            self.outcome.write('|============================================\n\n')
                    else:
                        print('|-- Optimal objective value:', bestobj)
                        print('|-- Optimal x =')
                        for n in self.compartment_name.keys():
                            for m in self.compartment_info[n]['tocpm'].keys():
                                if sum([sum(x[n][m][j]) for j in range(self.group)]) != 0:
                                    print('|---- From compartment ' + self.compartment_name[n] +
                                          ' to compartment ' + self.compartment_name[m] + ':')
                                    for j in range(self.group):
                                        print('|-------- ', x[n][m][j])
                        print('|============================================\n\n')
                        self.outcome.write('|-- Optimal objective value: ' + str(bestobj) + '\n')
                        self.outcome.write('|-- Optimal x =\n')
                        for n in self.compartment_name.keys():
                            for m in self.compartment_info[n]['tocpm'].keys():
                                if sum([sum(x[n][m][j]) for j in range(self.group)]) != 0:
                                    self.outcome.write('|---- From compartment ' + self.compartment_name[n] +
                                                       ' to compartment ' + self.compartment_name[m] + ':\n')
                                    for j in range(self.group):
                                        self.outcome.write('|-------- ' + str(x[n][m][j]) + '\n')
                        self.outcome.write('|============================================\n\n')
                except:
                    pass
                sol = (status, bestobj, X, x, Xc)
                return sol
            except:
                return 0

    def __solvemodel_gurobi(self, ct=None, target=None):
        if self.types == 'Expectation':
            X = [[[0 for t in range(self.Time)] for j in range(self.group)]
                 for n in range(max(self.compartment_name.keys()) + 1)]
            x = [[[[0 for t in range(self.Time)] for j in range(self.group)]
                  for m in range(max(self.compartment_name.keys()) + 1)]
                 for n in range(max(self.compartment_name.keys()) + 1)]
            Xc = {}
            if ct != None:
                self.model.setParam("timelimit", ct)
            self.model.optimize()

            try:
                bestobj = self.model.getAttr('ObjVal')
                for n in self.compartment_name.keys():
                    X[n] = [[round(self.model.getVarByName('X.' + str(n) + '.' + str(j) + '.' + str(t)).x, 6)
                             for t in range(self.Time)] for j in range(self.group)]
                    for m in self.compartment_info[n]['tocpm'].keys():
                        x[n][m] = [
                            [self.model.getVarByName('x.' + str(n) + '.' + str(m) + '.' + str(j) + '.' + str(t)).x
                             if (j, t) in self.compartment_info[n]['tocpm'][m]['x'] else 0
                             for t in range(self.Time)] for j in range(self.group)]
                for n in self.custom_var.keys():
                    Xc.update({n: self.model.getVarByName(n).x})
                sol = (self.model.getAttr('status'), bestobj, X, x, Xc)
                return sol
            except:
                return 0
        elif self.types == 'Robustness' and target != None:
            # self.model.set_results_stream(None)
            # self.model.set_log_stream(None)
            T0 = []
            if self.objtype:
                self.robustconstraint.update({'obj': [0, self.theta, self.theta * 10, 0]})
            T1 = list(self.robustconstraint.keys())
            T1.sort()
            p = 0
            print('\n============================================')
            while T1 != []:
                if p > 0:
                    for name in T0:
                        self.robustconstraint[name][0] = self.robustconstraint[name][2]
                        self.robustconstraint[name][1] = self.robustconstraint[name][2]
                        self.robustconstraint[name][3] = 0
                    for name in T1:
                        self.robustconstraint[name][0] = 0
                        self.robustconstraint[name][1] = self.robustconstraint[name][2]
                        self.robustconstraint[name][3] = 0
                    p += 1
                else:
                    for name in T1:
                        self.robustconstraint[name] = [0, self.theta, self.theta * 10, 0]
                    p += 1
                num_infeasible = 0
                gaptheta = 1e10
                pp = 0
                print('+- Begin round ' + str(p))
                if self.__log_stream == 1:
                    self.outcome.write('+- Begin round ' + str(p) + '\n')
                while max([self.robustconstraint[name][2] / (self.robustconstraint[name][0] + 0.0001) for name in
                           T1]) > 1.01:
                    pp += 1
                    for name in T1:
                        if name != 'obj':
                            for i in range(len(self.dcoef_robust[name])):
                                if self.dcoef_robust[name][i] != 0:
                                    self.model.chgCoeff(self.model.getConstrByName(name),
                                                        self.model.getVarByName(self.dvar_robust[name][i]),
                                                        self.dcoef_robust[name][i] ** 2 * 0.5 / self.robustconstraint[name][1])
                        else:
                            # m.chgCoeff(m.getConstrByName('c0'), m.getVarByName('x'), 2)
                            for ii in range(len(self.robustobj)):
                                self.model.getVarByName(self.robustobj[ii][0]).Obj = \
                                    0.5 * self.robustobj[ii][1] ** 2 / self.robustconstraint[name][1]
                    if ct != None:
                        self.model.setParam("timelimit", ct)
                    print('|- Iteration ' + str(pp) + ' of round ' + str(p) + ':')
                    if self.__log_stream == 1:
                        self.outcome.write('|- Iteration ' + str(pp) + ' of round ' + str(p) + ':\n')
                    self.model.optimize()
                    try:
                        objective_value = self.model.getAttr('ObjVal')
                    except:
                        objective_value = 1e+30
                    if objective_value < target:
                        num_infeasible = 0
                        time.sleep(5)
                        # self.model.write('2222.lp')
                        self.model.write('modelrobust_SO_' + str(round(target, 2)) + '.mps')
                        status = self.model.getAttr('status')
                        bestobj = self.model.getAttr('ObjVal')
                        X = [[[0 for t in range(self.Time)] for j in range(self.group)]
                             for n in range(max(self.compartment_name.keys()) + 1)]
                        Xdagger = [[[0 for t in range(self.Time)] for j in range(self.group)]
                                   for n in range(max(self.compartment_name.keys()) + 1)]
                        x = [[[[0 for t in range(self.Time)] for j in range(self.group)]
                              for m in range(max(self.compartment_name.keys()) + 1)]
                             for n in range(max(self.compartment_name.keys()) + 1)]
                        Xc = {}
                        for n in self.compartment_name.keys():
                            X[n] = [[self.model.getVarByName('X.' + str(n) + '.' + str(j) + '.' + str(t)).x
                                     for t in range(self.Time)] for j in range(self.group)]
                            Xdagger[n] = [
                                [self.model.getVarByName('Xdagger.' + str(n) + '.' + str(j) + '.' + str(t)).x
                                 for t in range(self.Time)] for j in range(self.group)]
                            for m in self.compartment_info[n]['tocpm'].keys():
                                x[n][m] = [[round(self.model.getVarByName(
                                    'x.' + str(n) + '.' + str(m) + '.' + str(j) + '.' + str(t)).x, 6)
                                            if (j, t) in self.compartment_info[n]['tocpm'][m]['x'] else 0
                                            for t in range(self.Time)] for j in range(self.group)]
                        for n in self.custom_var.keys():
                            Xc.update({n: self.model.getVarByName(n).x})
                        for name in T1:
                            if name != 'obj':
                                drhs = self.model.getConstrByName(name).rhs
                                dvar1 = [self.model.getVarByName(iii).x for iii in self.constraint[name][0]]
                                dcoef1 = [self.model.getCoeff(self.model.getConstrByName(name), self.model.getVarByName(iii))
                                          for iii in self.constraint[name][0]]
                                lhs1 = sum([dvar1[i] * dcoef1[i] for i in range(len(dvar1))])

                                dvar2 = [self.model.getVarByName(iii).x for iii in self.dvar_robust[name]]
                                dcoef2 = [self.model.getCoeff(self.model.getConstrByName(name), self.model.getVarByName(iii))
                                          for iii in self.dvar_robust[name]]
                                lhs2 = sum([dvar2[i] * dcoef2[i] for i in range(len(dvar2))])
                                if lhs1 + lhs2 < 0.999 * drhs:
                                    if lhs2 != 0:
                                        gaptheta = min(gaptheta, (drhs - lhs1) / lhs2 - 0.1)
                                    else:
                                        gaptheta = 0
                                else:
                                    self.robustconstraint[name][3] = 1
                                    gaptheta = 0
                            else:
                                if self.model.getAttr('ObjVal') < 0.999 * target:
                                    dvar3 = [self.model.getVarByName(self.expobj[iii][0]).x for iii in range(len(self.expobj))]
                                    dcoef3 = [self.model.getVarByName(self.expobj[iii][0]).obj for iii in range(len(self.expobj))]
                                    dvar4 = [self.model.getVarByName(self.robustobj[iii][0]).x for iii in range(len(self.robustobj))]
                                    dcoef4 = [self.model.getVarByName(self.robustobj[iii][0]).obj for iii in range(len(self.robustobj))]

                                    lhs3 = sum([dvar3[iii] * dcoef3[iii] for iii in range(len(dvar3))])
                                    lhs4 = sum([dvar4[iii] * dcoef4[iii] for iii in range(len(dvar4))])
                                    if lhs4 != 0:
                                        gaptheta = min(gaptheta, (target - lhs3) / lhs4 - 0.1)
                                    else:
                                        gaptheta = 0
                                else:
                                    self.robustconstraint[name][3] = 1
                                    gaptheta = 0
                        print('|-- Feasible ',
                              [(name, self.robustconstraint[name]) for name in self.robustconstraint.keys()])
                        if self.__log_stream == 1:
                            self.outcome.write('|-- Feasible ' + str([(name, self.robustconstraint[name])
                                                                      for name in self.robustconstraint.keys()]) + '\n')
                        if gaptheta <= 1:
                            for name in T1:
                                # print(self.robustconstraint[name])
                                self.robustconstraint[name][2] = self.robustconstraint[name][1]
                                self.robustconstraint[name][1] = (self.robustconstraint[name][0]
                                                                  + self.robustconstraint[name][1]) / 2.0
                        else:
                            for name in T1:
                                self.robustconstraint[name][2] = self.robustconstraint[name][1]
                                self.robustconstraint[name][1] = self.robustconstraint[name][1] / gaptheta
                            gaptheta = 0
                        print('|-- Current objective and target:', self.model.getAttr('ObjVal'), target)
                        print('|-- Current x =')
                        if self.__log_stream == 1:
                            self.outcome.write('|-- Current objective and target: '+ str(self.model.getAttr('ObjVal'))
                                               + ',' + str(target) + '\n')
                            self.outcome.write('|-- Current x =\n')
                        for n in self.compartment_name.keys():
                            for m in self.compartment_info[n]['tocpm'].keys():
                                if sum([sum(x[n][m][j]) for j in range(self.group)]) != 0:
                                    print('|---- From compartment ' + self.compartment_name[n] +
                                          ' to compartment ' + self.compartment_name[m] + ':')
                                    for j in range(self.group):
                                        print('|-------- ', x[n][m][j])
                                    if self.__log_stream == 1:
                                        self.outcome.write('|---- From compartment ' + self.compartment_name[n] +
                                                           ' to compartment ' + self.compartment_name[m] + ':\n')
                                        for j in range(self.group):
                                            self.outcome.write('|-------- ' + str(x[n][m][j]) + '\n')
                        print('|--------------------------------------------')
                        if self.__log_stream == 1:
                            self.outcome.write('|--------------------------------------------\n')
                            # self.model = cplex.Cplex('model_robust_' + str(target) + '.mps')
                    else:
                        if num_infeasible < 3:
                            for name in T1:
                                # print(self.robustconstraint[name])
                                self.robustconstraint[name][0] = self.robustconstraint[name][1]
                                self.robustconstraint[name][1] = (self.robustconstraint[name][2]
                                                                  + self.robustconstraint[name][1]) / 2.0
                            num_infeasible += 1
                        else:
                            for name in T1:
                                # print(self.robustconstraint[name])
                                self.robustconstraint[name][0] = self.robustconstraint[name][1]
                                self.robustconstraint[name][1] = max(self.robustconstraint[name][2] - 0.1,
                                                                     (self.robustconstraint[name][2] +
                                                                      self.robustconstraint[name][1]) / 2.0)
                            num_infeasible = 0
                        print('|-- Infeasible ',
                              [(name, self.robustconstraint[name]) for name in self.robustconstraint.keys()])
                        print('|--------------------------------------------')
                        if self.__log_stream == 1:
                            self.outcome.write('|-- Infeasible '
                                               + str(
                                [(name, self.robustconstraint[name]) for name in self.robustconstraint.keys()])
                                               + '\n')
                            self.outcome.write('|--------------------------------------------\n')
                # print('iter = ',[self.robustconstraint[name][3] for name in self.robustconstraint.keys()])
                if sum([self.robustconstraint[name][3] for name in self.robustconstraint.keys()]) == 0:
                    T1 = []
                    T0 = list(self.robustconstraint.keys())
                else:
                    for name in T1:
                        if self.robustconstraint[name][3] == 1:
                            T0.append(name)
                            T1.remove(name)
                print('+- Now sets T0 and T1 are')
                print('|- T1 = ', T1)
                print('|- T0 = ', T0)
                print('|--------------------------------------------')
                if self.__log_stream == 1:
                    self.outcome.write('+- Now sets T0 and T1 are\n')
                    self.outcome.write('|- T1 = ' + str(T1) + '\n')
                    self.outcome.write('|- T0 = ' + str(T0) + '\n')
                    self.outcome.write('|--------------------------------------------\n')
            try:
                try:
                    time.sleep(5)
                    self.model = gp.read('modelrobust_SO_' + str(round(target, 2)) + '.mps')
                    # self.model.set_results_stream(None)
                    # self.model.set_log_stream(None)
                    self.model.optimize()
                    try:
                        objective_value = self.model.getAttr('ObjVal')
                    except:
                        objective_value = 1e+30
                    if objective_value < bestobj:
                        print('|-- Optimal objective value:', self.model.getAttr('ObjVal'), bestobj)
                        if self.__log_stream == 1:
                            self.outcome.write('|-- Optimal objective value: '
                                               + str(self.model.getAttr('ObjVal')) + '\n')
                        status = self.model.getAttr('status')
                        bestobj = self.model.getAttr('ObjVal')
                        X = [[[0 for t in range(self.Time)] for j in range(self.group)]
                             for n in range(max(self.compartment_name.keys()) + 1)]
                        x = [[[[0 for t in range(self.Time)] for j in range(self.group)]
                              for m in range(max(self.compartment_name.keys()) + 1)]
                             for n in range(max(self.compartment_name.keys()) + 1)]
                        Xc = {}
                        print('|-- Optimal x=')
                        if self.__log_stream == 1:
                            self.outcome.write('|-- Optimal x=\n')
                        for n in self.compartment_name.keys():
                            X[n] = [[round(self.model.getVarByName('X.' + str(n) + '.' + str(j) + '.' + str(t)).x,6)
                                     for t in range(self.Time)] for j in range(self.group)]
                            for m in self.compartment_info[n]['tocpm'].keys():
                                x[n][m] = [[round(self.model.getVarByName(
                                    'x.' + str(n) + '.' + str(m) + '.' + str(j) + '.' + str(t)).x, 6)
                                            if (j, t) in self.compartment_info[n]['tocpm'][m]['x'] else 0
                                            for t in range(self.Time)] for j in range(self.group)]
                                if sum([sum(x[n][m][j]) for j in range(self.group)]) != 0:
                                    print('|---- From compartment ' + self.compartment_name[n] +
                                          ' to compartment ' + self.compartment_name[m] + ':')
                                    for j in range(self.group):
                                        print('|-------- ', x[n][m][j])
                                    if self.__log_stream == 1:
                                        self.outcome.write('|---- From compartment ' + self.compartment_name[n] +
                                                           ' to compartment ' + self.compartment_name[m] + ':\n')
                                        for j in range(self.group):
                                            self.outcome.write('|-------- ' + str(x[n][m][j]) + '\n')
                        for n in self.custom_var.keys():
                            Xc.update({n: self.model.getVarByName(n).x})
                        print('|============================================\n\n')
                        if self.__log_stream == 1:
                            self.outcome.write('|============================================\n\n')
                    else:
                        print('|-- Optimal objective value:', bestobj)
                        print('|-- Optimal x =')
                        for n in self.compartment_name.keys():
                            for m in self.compartment_info[n]['tocpm'].keys():
                                if sum([sum(x[n][m][j]) for j in range(self.group)]) != 0:
                                    print('|---- From compartment ' + self.compartment_name[n] +
                                          ' to compartment ' + self.compartment_name[m] + ':')
                                    for j in range(self.group):
                                        print('|-------- ', x[n][m][j])
                        print('|============================================\n\n')
                        self.outcome.write('|-- Optimal objective value: ' + str(bestobj) + '\n')
                        self.outcome.write('|-- Optimal x =\n')
                        for n in self.compartment_name.keys():
                            for m in self.compartment_info[n]['tocpm'].keys():
                                if sum([sum(x[n][m][j]) for j in range(self.group)]) != 0:
                                    self.outcome.write('|---- From compartment ' + self.compartment_name[n] +
                                                       ' to compartment ' + self.compartment_name[m] + ':\n')
                                    for j in range(self.group):
                                        self.outcome.write('|-------- ' + str(x[n][m][j]) + '\n')
                        self.outcome.write('|============================================\n\n')
                except:
                    pass
                sol = (status, bestobj, X, x, Xc)
                return sol
            except:
                return 0
    '''================================================================='''
    '''==========================Prediction============================='''
    def set_x(self, x=None):
        setattr(self, 'x', x)

    def functionp(self, state= None, X = None, x = None, types = None):
        (m,j,k,t) = state

        if self.compartment_info[m]['local']['p_dp'] != {}:
            temp = self.compartment_info[m]['local']['p_dp'].keys()
            timemax = max([temp[i][2] for i in range(len(temp)) if temp[i][0]==j and temp[i][1]==k])
            a = self.compartment_info[m]['local']['p_dp'][(j,k,min(t,timemax))][0]
            b = self.compartment_info[m]['local']['p_dp'][(j,k,min(t,timemax))][1]

            p = sum([a[mm][j][min(tau,timemax)] * X[mm][jj][tau] for mm in self.compartment_name.keys()
                 for jj in range(self.group) for tau in range(t+1)]) \
                + sum([b[nn][mm][jj][min(tau,timemax)] * x[nn][mm][jj][tau]
                       for nn in self.compartment_name.keys() for mm in self.compartment_info[nn]['tocpm'].keys()
                       for jj in range(self.group) for tau in range(t+1)
                       if (jj, tau) in self.compartment_info[nn]['tocpm'][mm]['x']]) \
                + self.compartment_info[m]['local']['p_idp'][j][k][t]
            if types == 'stair':
                if j == k:
                    p = int(p / self.gap)*self.gap
                else:
                    p = int(p / self.gap + 1) * self.gap
        else:
            # print(self.compartment_info[m]['local']['p_idp'])
            timemax = len(self.compartment_info[m]['local']['p_idp'][j][k]) - 1
            p = self.compartment_info[m]['local']['p_idp'][j][k][min(t,timemax)]
        return p

    def functionq(self, state=None, X=None, x=None, phat = None, types = None, scale = 1):
        (m,n,j,t) = state
        if n in self.compartment_info[m]['tocpm'].keys():
            if self.compartment_info[m]['tocpm'][n]['q_dp'] != {}:
                temp = list(self.compartment_info[m]['tocpm'][n]['q_norm'].keys())
                timemax = max([temp[i][1] for i in range(len(temp)) if temp[i][0] == j])
                t = min(t, timemax)
                betabar = self.compartment_info[m]['tocpm'][n]['q_norm'][(j, t)][0]
                gammabar = self.compartment_info[m]['tocpm'][n]['q_norm'][(j, t)][1]
                phibar = self.compartment_info[m]['tocpm'][n]['q_norm'][(j, t)][2]
                psibar = self.compartment_info[m]['tocpm'][n]['q_norm'][(j, t)][3]
                alphabar = self.compartment_info[m]['tocpm'][n]['q_norm'][(j, t)][4]
                beta = self.compartment_info[m]['tocpm'][n]['q_dp'][(j, t)][0]
                gamma = self.compartment_info[m]['tocpm'][n]['q_dp'][(j, t)][1]
                phi = self.compartment_info[m]['tocpm'][n]['q_dp'][(j, t)][2]
                psi = self.compartment_info[m]['tocpm'][n]['q_dp'][(j, t)][3]
                alpha = self.compartment_info[m]['tocpm'][n]['q_dp'][(j, t)][4]
                qhat1 = sum([X[mm][k][t]/ scale * (beta[mm][k] * phat[mm][k][j][t] + gamma[mm][k])
                                              for mm in self.compartment_name.keys() for k in range(self.group)])\
                         + sum([x[mm][nn][k][t]/ scale * (phi[mm][k] * phat[mm][k][j][t] + psi[mm][k])
                                              for mm in self.compartment_name.keys()
                              for nn in self.compartment_info[mm]['tocpm'].keys() for k in range(self.group)
                              if (k, t) in self.compartment_info[mm]['tocpm'][nn]['x'] ]) + alpha
                qhat2 = sum([X[mm][k][t]/ scale * (betabar[mm][k] * phat[mm][k][j][t] + gammabar[mm][k])
                                              for mm in self.compartment_name.keys() for k in range(self.group)])\
                         + sum([x[mm][nn][k][t]/ scale * (phibar[mm][k] * phat[mm][k][j][t] + psibar[mm][k])
                                              for mm in self.compartment_name.keys()
                              for nn in self.compartment_info[mm]['tocpm'].keys() for k in range(self.group)
                              if (k,t) in self.compartment_info[mm]['tocpm'][nn]['x'] ]) + alphabar
                if types == 'stair':
                    if m == n:
                        qhat = int(qhat1 * 1.0 / qhat2 /self.gap) * self.gap
                    else:
                        qhat = int(qhat1 * 1.0 / qhat2 / self.gap + 1) * self.gap
                else:
                    qhat = qhat1 * 1.0 / qhat2
                if m != n:
                    qhat = min(max(self.compartment_info[m]['tocpm'][n]['q_lb'][(j, t)], qhat),
                               self.compartment_info[m]['tocpm'][n]['q_ub'][(j, t)])
            else:
                timemax = len(self.compartment_info[m]['tocpm'][n]['q_idp'][j]) - 1
                qhat = self.compartment_info[m]['tocpm'][n]['q_idp'][j][min(t,timemax)]

            return qhat
        else:
            return 0

    def DeterPrediction2(self, types=None, randomq=None):
        '''Deterministic Epidemiological Prediction Model'''
        # population.insert(0,[sum([self.population[n][j] for n in range(self.compartment)]) for j in range(self.group)])
        X = [[[0 for t in range(self.Time)] for j in range(self.group)]
             for n in range(max(self.compartment_name.keys())+1)]
        randomp = None
        Xbound = [1e+20,1e+20,1000,150,1e+20,1e+20]
        if self.x != None:
            x = copy.deepcopy(self.x)
        else:
            x = [[[[0 for t in range(self.Time)] for j in range(self.group)]
                  for m in range(max(self.compartment_name.keys())+1)]
                 for n in range(max(self.compartment_name.keys())+1)]
        qhat = [[[[0 for t in range(self.Time)] for j in range(self.group)]
              for m in range(max(self.compartment_name.keys())+1)]
             for n in range(max(self.compartment_name.keys())+1)]
        phat = [[[[0 for t in range(self.Time)] for k in range(self.group)] for j in range(self.group)]
             for n in range(max(self.compartment_name.keys())+1)]


        for n in self.compartment_name.keys():
            for j in range(self.group):
                X[n][j][0] = self.compartment_info[n]['local']['population'][j]

        for t in range(self.Time - 1):
            for m in self.compartment_name.keys():
                for k in range(self.group):
                    if sum([x[m][mm][k][t] for mm in self.compartment_name.keys()]) > X[m][k][t]:
                        tempx = sum([x[m][mm][k][t] for mm in self.compartment_name.keys()])
                        for mm in self.compartment_name.keys():
                            x[m][mm][k][t] = x[m][mm][k][t] * X[m][k][t] / tempx

            if randomp == None:
                for n in self.compartment_name.keys():
                    for j in range(self.group):
                        for k in range(self.group):
                            phat[n][j][k][t] = self.functionp(state=(n, j, k, t), X=X, x=x, types=types)
            else:
                for n in self.compartment_name.keys():
                    for j in range(self.group):
                        for k in range(self.group):
                            phat[n][j][k][t] = self.functionp(state=(n, j, k, t), X=X, x=x, types=types) * randomp[n]
                        ptotal = max(1, sum([phat[n][j][k][t] for k in range(self.group) if j != k]))
                        for k in range(self.group):
                            if j != k:
                                phat[n][j][k][t] = phat[n][j][k][t] / ptotal
                        phat[n][j][j][t] = 1 - sum([phat[n][j][kk][t] for kk in range(self.group) if j != kk])
            if randomq == None:
                for j in range(self.group):
                    for n in self.compartment_name.keys():
                        for m in self.compartment_info[n]['tocpm'].keys():
                            qhat[n][m][j][t] = self.functionq(state=(n, m, j, t), X=X, x=x, phat=phat, types=types)
                        qhat_fromn = sum([qhat[n][m][j][t] for m in self.compartment_info[n]['tocpm'].keys() if m != n])
                        if qhat_fromn > 1:
                            qhat_fromn_dp = sum([qhat[n][m][j][t] for m in self.compartment_info[n]['tocpm'].keys()
                                                 if self.compartment_info[n]['tocpm'][m]['q_dp'] != {}])
                            qhat_fromn_idp = qhat_fromn - qhat_fromn_dp
                            for m in self.compartment_info[n]['tocpm'].keys():
                                if self.compartment_info[n]['tocpm'][m]['q_dp'] != {}:
                                    qhat[n][m][j][t] = qhat[n][m][j][t] * max(0, 1 - qhat_fromn_idp) / qhat_fromn_dp
                        qhat[n][n][j][t] = 1 - sum([qhat[n][mm][j][t] for mm in self.compartment_info[n]['tocpm'].keys()
                                                    if n != mm])
            else:
                for j in range(self.group):
                    for n in self.compartment_name.keys():
                        for m in self.compartment_info[n]['tocpm'].keys():
                            qhat[n][m][j][t] = self.functionq(state=(n, m, j, t), X=X, x=x, phat=phat, types=types) \
                                               * randomq[n][m]
                        qhat_fromn = sum([qhat[n][m][j][t] for m in self.compartment_info[n]['tocpm'].keys() if m != n])
                        if qhat_fromn > 1:
                            qhat_fromn_dp = sum([qhat[n][m][j][t] for m in self.compartment_info[n]['tocpm'].keys()
                                                  if self.compartment_info[n]['tocpm'][m]['q_dp'] != {}])
                            qhat_fromn_idp = qhat_fromn - qhat_fromn_dp
                            for m in self.compartment_info[n]['tocpm'].keys():
                                if self.compartment_info[n]['tocpm'][m]['q_dp'] != {}:
                                    qhat[n][m][j][t] = qhat[n][m][j][t] * max(0, 1-qhat_fromn_idp) / qhat_fromn_dp

                        qtotal = max(1, sum([qhat[n][m][j][t] * randomq[n][m]
                                        for m in self.compartment_info[n]['tocpm'].keys() if n!=m]))
                        for m in self.compartment_info[n]['tocpm'].keys():
                            if n != m:
                                qhat[n][m][j][t] = qhat[n][m][j][t] / qtotal
                        qhat[n][n][j][t] = 1 - sum([qhat[n][mm][j][t]
                                for mm in self.compartment_info[n]['tocpm'].keys() if n != mm])

            for n in self.compartment_name.keys():
                for j in range(self.group):
                    X[n][j][t + 1] = sum([min(X[m][j][t], x[m][n][j][t]) for m in self.compartment_name.keys()]) \
                                     + sum([max(X[m][k][t] - sum([x[m][mm][k][t] for mm in self.compartment_name.keys()]), 0) *
                                            qhat[m][n][j][t] * phat[m][k][j][t]
                                            for m in self.compartment_name.keys() for k in range(self.group)])

            if sum([X[2][j][t + 1] for j in range(self.group)]) > Xbound[2]:
                # print(t, sum([X[2][j][t + 1] for j in range(self.group)]), Xbound[2])
                a = Xbound[2] / sum([X[2][j][t + 1] for j in range(self.group)])
                for j in range(self.group):
                    X[1][j][t + 1] += X[2][j][t + 1] * (1 - a) * (1 - qhat[2][4][j][t] - qhat[2][5][j][t])
                    X[4][j][t + 1] += X[2][j][t + 1] * (1 - a) * qhat[2][4][j][t]
                    X[5][j][t + 1] += X[2][j][t + 1] * (1 - a) * qhat[2][5][j][t]
                    X[2][j][t + 1] = X[2][j][t + 1] * a
                # print(t, sum([X[2][j][t + 1] for j in range(self.group)]), Xbound[2])
            if sum([X[3][j][t + 1] for j in range(self.group)]) > Xbound[3]:
                # print(t, sum([X[3][j][t + 1] for j in range(self.group)]), Xbound[3])
                a = Xbound[3] / sum([X[3][j][t + 1] for j in range(self.group)])
                for j in range(self.group):
                    X[1][j][t + 1] += X[3][j][t + 1] * (1 - a)\
                                      * (1 - qhat[3][4][j][t] - qhat[3][5][j][t])
                    X[4][j][t + 1] += X[3][j][t + 1] * (1 - a) * qhat[3][4][j][t]
                    X[5][j][t + 1] += X[3][j][t + 1] * (1 - a) * qhat[3][5][j][t]
                    X[3][j][t + 1] = a * X[3][j][t + 1]
        if types == 'stair':
            return X, phat, qhat
        else:
            return X, x, qhat

    def DeterPrediction(self, types=None, randomq=None):
        '''Deterministic Epidemiological Prediction Model'''
        # population.insert(0,[sum([self.population[n][j] for n in range(self.compartment)]) for j in range(self.group)])
        X = [[[0 for t in range(self.Time)] for j in range(self.group)]
             for n in range(max(self.compartment_name.keys())+1)]
        Xdagger = [[[0 for t in range(self.Time)] for j in range(self.group)]
             for n in range(max(self.compartment_name.keys())+1)]
        randomp = None
        if self.x != None:
            x = copy.deepcopy(self.x)
        else:
            x = [[[[0 for t in range(self.Time)] for j in range(self.group)]
                  for m in range(max(self.compartment_name.keys())+1)]
                 for n in range(max(self.compartment_name.keys())+1)]
        qhat = [[[[0 for t in range(self.Time)] for j in range(self.group)]
              for m in range(max(self.compartment_name.keys())+1)]
             for n in range(max(self.compartment_name.keys())+1)]
        phat = [[[[0 for t in range(self.Time)] for k in range(self.group)] for j in range(self.group)]
             for n in range(max(self.compartment_name.keys())+1)]


        for n in self.compartment_name.keys():
            for j in range(self.group):
                X[n][j][0] = self.compartment_info[n]['local']['population'][j]

        for t in range(self.Time - 1):
            for m in self.compartment_name.keys():
                for k in range(self.group):
                    if sum([x[m][mm][k][t] for mm in self.compartment_name.keys()]) > X[m][k][t]:
                        tempx = sum([x[m][mm][k][t] for mm in self.compartment_name.keys()])
                        for mm in self.compartment_name.keys():
                            x[m][mm][k][t] = x[m][mm][k][t] * X[m][k][t] / tempx

            if randomp == None:
                for n in self.compartment_name.keys():
                    for j in range(self.group):
                        for k in range(self.group):
                            phat[n][j][k][t] = self.functionp(state=(n, j, k, t), X=X, x=x, types=types)
            else:
                for n in self.compartment_name.keys():
                    for j in range(self.group):
                        for k in range(self.group):
                            phat[n][j][k][t] = self.functionp(state=(n, j, k, t), X=X, x=x, types=types) * randomp[n]
                        ptotal = max(1, sum([phat[n][j][k][t] for k in range(self.group) if j != k]))
                        for k in range(self.group):
                            if j != k:
                                phat[n][j][k][t] = phat[n][j][k][t] / ptotal
                        phat[n][j][j][t] = 1 - sum([phat[n][j][kk][t] for kk in range(self.group) if j != kk])
            if randomq == None:
                for j in range(self.group):
                    for n in self.compartment_name.keys():
                        for m in self.compartment_info[n]['tocpm'].keys():
                            qhat[n][m][j][t] = self.functionq(state=(n, m, j, t), X=X, x=x, phat=phat, types=types)
                        qhat_fromn = sum([qhat[n][m][j][t] for m in self.compartment_info[n]['tocpm'].keys() if m != n])
                        if qhat_fromn > 1:
                            qhat_fromn_dp = sum([qhat[n][m][j][t] for m in self.compartment_info[n]['tocpm'].keys()
                                                 if self.compartment_info[n]['tocpm'][m]['q_dp'] != {}])
                            qhat_fromn_idp = qhat_fromn - qhat_fromn_dp
                            for m in self.compartment_info[n]['tocpm'].keys():
                                if self.compartment_info[n]['tocpm'][m]['q_dp'] != {}:
                                    qhat[n][m][j][t] = qhat[n][m][j][t] * max(0, 1 - qhat_fromn_idp) / qhat_fromn_dp
                        qhat[n][n][j][t] = 1 - sum([qhat[n][mm][j][t] for mm in self.compartment_info[n]['tocpm'].keys()
                                                    if n != mm])
            else:
                for j in range(self.group):
                    for n in self.compartment_name.keys():
                        for m in self.compartment_info[n]['tocpm'].keys():
                            qhat[n][m][j][t] = self.functionq(state=(n, m, j, t), X=X, x=x, phat=phat, types=types) \
                                               * randomq[n][m]
                        qhat_fromn = sum([qhat[n][m][j][t] for m in self.compartment_info[n]['tocpm'].keys() if m != n])
                        if qhat_fromn > 1:
                            qhat_fromn_dp = sum([qhat[n][m][j][t] for m in self.compartment_info[n]['tocpm'].keys()
                                                  if self.compartment_info[n]['tocpm'][m]['q_dp'] != {}])
                            qhat_fromn_idp = qhat_fromn - qhat_fromn_dp
                            for m in self.compartment_info[n]['tocpm'].keys():
                                if self.compartment_info[n]['tocpm'][m]['q_dp'] != {}:
                                    qhat[n][m][j][t] = qhat[n][m][j][t] * max(0, 1-qhat_fromn_idp) / qhat_fromn_dp

                        qtotal = max(1, sum([qhat[n][m][j][t] * randomq[n][m]
                                        for m in self.compartment_info[n]['tocpm'].keys() if n!=m]))
                        for m in self.compartment_info[n]['tocpm'].keys():
                            if n != m:
                                qhat[n][m][j][t] = qhat[n][m][j][t] / qtotal
                        qhat[n][n][j][t] = 1 - sum([qhat[n][mm][j][t]
                                for mm in self.compartment_info[n]['tocpm'].keys() if n != mm])

            for n in self.compartment_name.keys():
                for j in range(self.group):
                    X[n][j][t + 1] = sum([min(X[m][j][t], x[m][n][j][t]) for m in self.compartment_name.keys()]) \
                                     + sum([max(X[m][k][t] - sum([x[m][mm][k][t] for mm in self.compartment_name.keys()]), 0) *
                                            qhat[m][n][j][t] * phat[m][k][j][t]
                                            for m in self.compartment_name.keys() for k in range(self.group)])
        if types == 'stair':
            return X, phat, qhat
        else:
            return X, x, qhat

    def Sample_generation(self, n=None, m=None, lb=1, ub=2, size=1000):
        if isinstance(n, str) and isinstance(lb, (int, float)):
            n = [n]
            lb = [lb]
        if isinstance(m, str) and isinstance(ub, (int, float)):
            m = [m]
            ub = [ub]

        if isinstance(n, list) and isinstance(m, list) and len(n)==len(m)==len(ub)==len(lb):
            randomq = [[[1 for i in range(self.compartment)] for ii in range(self.compartment)] for s in range(size)]
            for i in range(len(n)):
                n1 = get_keys(self.compartment_name, n[i])[0]
                m1 = get_keys(self.compartment_name, m[i])[0]
                for s in range(size):
                    randomq[s][n1][m1] = random.uniform(lb[i], ub[i])
            return randomq
        else:
            ValueError('n or m should be a string or a list of string with same dimensionality')

    def StochaPrediction(self, randomq=None):
        '''Stochastic Epidemiological Prediction Model'''
        # population.insert(0,[sum([self.population[n][j] for n in range(self.compartment)]) for j in range(self.group)])
        X = [[[0 for t in range(self.Time)] for j in range(self.group)]
             for n in range(max(self.compartment_name.keys()) + 1)]
        randomp = None
        # Y = [[[0 for t in range(self.Time)] for j in range(self.group)]
        #      for n in range(max(self.compartment_name.keys()) + 1)]
        # Z = [[0 for t in range(self.Time)] for n in range(max(self.compartment_name.keys()) + 1)]
        if self.x != None:
            x = copy.deepcopy(self.x)
        else:
            x = [[[[0 for t in range(self.Time)] for j in range(self.group)]
                  for m in range(max(self.compartment_name.keys())+1)]
                 for n in range(max(self.compartment_name.keys())+1)]

        qhat = [[[[0 for t in range(self.Time)] for j in range(self.group)]
                 for m in range(max(self.compartment_name.keys()) + 1)]
                for n in range(max(self.compartment_name.keys()) + 1)]
        phat = [[[[0 for t in range(self.Time)] for k in range(self.group)] for j in range(self.group)]
                    for n in range(max(self.compartment_name.keys()) + 1)]
        population = [self.compartment_info[n]['local']['population'] if n in self.compartment_name.keys() else []
                      for n in range(max(self.compartment_name.keys()) + 1)]

        for n in self.compartment_name.keys():
            # Z[n][0] = sum([self.compartment_info[n]['local']['population'][j] for j in range(self.group)])
            for j in range(self.group):
                X[n][j][0] = self.compartment_info[n]['local']['population'][j]
                # Y[n][j][0] = self.compartment_info[n]['local']['population'][j]

        # if max([sum(population[n]) for n in range(max(self.compartment_name.keys()) + 1)]) < 10000:
        X = [[[X[n][j][t] * 1000 for t in range(self.Time)] for j in range(self.group)]
             for n in range(max(self.compartment_name.keys()) + 1)]
        # Y = [[[Y[n][j][t] * 1000 for t in range(self.Time)] for j in range(self.group)]
        #      for n in range(max(self.compartment_name.keys()) + 1)]
        # Z = [[Z[n][t] * 1000 for t in range(self.Time)]
        #      for n in range(max(self.compartment_name.keys()) + 1)]
        x = [[[[x[n][m][j][t]*1000 for t in range(self.Time)] for j in range(self.group)]
              for m in range(max(self.compartment_name.keys()) + 1)]
             for n in range(max(self.compartment_name.keys()) + 1)]

        for t in range(self.Time - 1):

            for m in self.compartment_name.keys():
                for k in range(self.group):
                    if sum([x[m][mm][k][t] for mm in self.compartment_name.keys()]) > X[m][k][t]:
                        for mm in self.compartment_name.keys():
                            x[m][mm][k][t] = x[m][mm][k][t] * X[m][k][t] / \
                                             sum([x[m][mm][k][t] for mm in self.compartment_name.keys()])

            if randomp == None:
                for n in self.compartment_name.keys():
                    for j in range(self.group):
                        for k in range(self.group):
                            phat[n][j][k][t] = self.functionp(state=(n, j, k, t), X=X, x=x)
            else:
                for n in self.compartment_name.keys():
                    for j in range(self.group):
                        for k in range(self.group):
                            phat[n][j][k][t] = self.functionp(state=(n, j, k, t), X=X, x=x) * randomp[n]
                        ptotal = max(1, sum([phat[n][j][k][t] for k in range(self.group) if j != k]))
                        for k in range(self.group):
                            if j != k:
                                phat[n][j][k][t] = phat[n][j][k][t] / ptotal
                        phat[n][j][j][t] = 1 - sum([phat[n][j][kk][t] for kk in range(self.group) if j != kk])

            if randomq == None:
                for j in range(self.group):
                    for n in self.compartment_name.keys():
                        for m in self.compartment_info[n]['tocpm'].keys():
                            qhat[n][m][j][t] = self.functionq(state=(n, m, j, t), X=X, x=x, phat=phat, scale=1000)
                        qhat_fromn = sum([qhat[n][m][j][t] for m in self.compartment_info[n]['tocpm'].keys() if m != n])
                        if qhat_fromn > 1:
                            qhat_fromn_dp = sum([qhat[n][m][j][t] for m in self.compartment_info[n]['tocpm'].keys()
                                                 if self.compartment_info[n]['tocpm'][m]['q_dp'] != {}])
                            qhat_fromn_idp = qhat_fromn - qhat_fromn_dp
                            for m in self.compartment_info[n]['tocpm'].keys():
                                if self.compartment_info[n]['tocpm'][m]['q_dp'] != {}:
                                    qhat[n][m][j][t] = qhat[n][m][j][t] * max(0, 1 - qhat_fromn_idp) / qhat_fromn_dp
                        qhat[n][n][j][t] = 1 - sum([qhat[n][mm][j][t] for mm in self.compartment_info[n]['tocpm'].keys()
                                                    if n != mm])
            else:
                for j in range(self.group):
                    for n in self.compartment_name.keys():
                        for m in self.compartment_info[n]['tocpm'].keys():
                            qhat[n][m][j][t] = self.functionq(state=(n, m, j, t), X=X, x=x, phat=phat, scale=1000) \
                                               * randomq[n][m]
                        qhat_fromn = sum([qhat[n][m][j][t] for m in self.compartment_info[n]['tocpm'].keys() if m != n])
                        if qhat_fromn > 1:
                            qhat_fromn_dp = sum([qhat[n][m][j][t] for m in self.compartment_info[n]['tocpm'].keys()
                                                 if self.compartment_info[n]['tocpm'][m]['q_dp'] != {}])
                            qhat_fromn_idp = qhat_fromn - qhat_fromn_dp
                            for m in self.compartment_info[n]['tocpm'].keys():
                                if self.compartment_info[n]['tocpm'][m]['q_dp'] != {}:
                                    qhat[n][m][j][t] = qhat[n][m][j][t] * max(0, 1 - qhat_fromn_idp) / qhat_fromn_dp
                        qtotal = max(1, sum([qhat[n][m][j][t] * randomq[n][m]
                                        for m in self.compartment_info[n]['tocpm'].keys() if n!=m]))
                        for m in self.compartment_info[n]['tocpm'].keys():
                            if n != m:
                                qhat[n][m][j][t] = qhat[n][m][j][t] / qtotal
                        qhat[n][n][j][t] = 1 - sum([qhat[n][mm][j][t]
                                for mm in self.compartment_info[n]['tocpm'].keys() if n!=mm])

            for n in self.compartment_name.keys():
                for j in range(self.group):
                    X[n][j][t + 1] = sum([min(X[m][j][t],x[m][n][j][t]) for m in self.compartment_name.keys()]) \
                                     + sum([numpy.random.binomial(max(X[m][k][t] - sum([x[m][mm][k][t]
                                        for mm in self.compartment_name.keys()]), 0), qhat[m][n][j][t] * phat[m][k][j][t])
                                    for m in self.compartment_name.keys() for k in range(self.group)])
        X = [[[X[n][j][t] / 1000.0 for t in range(self.Time)] for j in range(self.group)]
             for n in range(max(self.compartment_name.keys()) + 1)]

        return X

    def RobustPrediction_get_upperbound(self, theta=10, coef = 1):
        '''Robust Epidemiological Prediction Model For Upper bound'''
        muupper = [[[0 for t in range(self.Time)] for j in range(self.group)]
                   for n in range(max(self.compartment_name.keys()) + 1)]
        X = [[[0 for t in range(self.Time)] for j in range(self.group)]
             for n in range(max(self.compartment_name.keys()) + 1)]
        x = copy.deepcopy(self.x)
        qhat = [[[[0 for t in range(self.Time)] for j in range(self.group)]
                 for m in range(max(self.compartment_name.keys()) + 1)]
                for n in range(max(self.compartment_name.keys()) + 1)]
        phat = [[[[0 for t in range(self.Time)] for k in range(self.group)] for j in range(self.group)]
                    for n in range(max(self.compartment_name.keys()) + 1)]

        for n in self.compartment_name.keys():
            for j in range(self.group):
                X[n][j][0] = self.compartment_info[n]['local']['population'][j]
                muupper[n][j][0] = self.compartment_info[n]['local']['population'][j]
        for t in range(self.Time - 1):

            for m in self.compartment_name.keys():
                for k in range(self.group):
                    if sum([x[m][mm][k][t] for mm in self.compartment_name.keys()]) > X[m][k][t]:
                        for mm in self.compartment_name.keys():
                            x[m][mm][k][t] = x[m][mm][k][t] * X[m][k][t] / \
                                             sum([x[m][mm][k][t] for mm in self.compartment_name.keys()])

            for n in self.compartment_name.keys():
                for j in range(self.group):
                    for k in range(self.group):
                        phat[n][j][k][t] = self.functionp(state=(n, j, k, t), X=X, x=x)

            for j in range(self.group):
                for n in self.compartment_name.keys():
                    for m in self.compartment_info[n]['tocpm'].keys():
                        qhat[n][m][j][t] = self.functionq(state=(n, m, j, t), X=X, x=x, phat=phat)
                    qhat_fromn = sum([qhat[n][m][j][t] for m in self.compartment_info[n]['tocpm'].keys() if m != n])
                    if qhat_fromn > 1:
                        qhat_fromn_dp = sum([qhat[n][m][j][t] for m in self.compartment_info[n]['tocpm'].keys()
                                             if self.compartment_info[n]['tocpm'][m]['q_dp'] != {}])
                        qhat_fromn_idp = qhat_fromn - qhat_fromn_dp
                        for m in self.compartment_info[n]['tocpm'].keys():
                            if self.compartment_info[n]['tocpm'][m]['q_dp'] != {}:
                                qhat[n][m][j][t] = qhat[n][m][j][t] * max(0, 1 - qhat_fromn_idp) / qhat_fromn_dp
                    qhat[n][n][j][t] = 1 - sum([qhat[n][mm][j][t] for mm in self.compartment_info[n]['tocpm'].keys()
                                                if n != mm])

            for n in self.compartment_name.keys():
                for j in range(self.group):
                    X[n][j][t + 1] = sum([x[m][n][j][t] for m in self.compartment_name.keys()]) \
                                     + sum([max(X[m][k][t] - sum([x[m][mm][k][t] for mm in self.compartment_name.keys()]), 0) *
                         qhat[m][n][j][t] * phat[m][k][j][t]
                         for m in self.compartment_name.keys() for k in range(self.group)])

            for n in self.compartment_name.keys():
                for j in range(self.group):
                    muupper[n][j][t + 1] = coef * sum([x[m][n][j][t] for m in self.compartment_name.keys()]) \
                                           + sum([max(X[m][k][t] - sum([x[m][mm][k][t] for mm in self.compartment_name.keys()]), 0)
                                                  * theta * log(
                        1 - qhat[m][n][j][t] * phat[m][k][j][t] + qhat[m][n][j][t] * phat[m][k][j][t] * exp(
                            coef * 1.0 / theta)) for m in self.compartment_name.keys() for k in range(self.group)])
        return muupper

    def RobustPrediction_get_lowerbound(self, theta=10, coef=1):
        '''Robust Epidemiological Prediction Model For Lower bound'''
        mulower = [[[0 for t in range(self.Time)] for j in range(self.group)]
                   for n in range(max(self.compartment_name.keys()) + 1)]
        X = [[[0 for t in range(self.Time)] for j in range(self.group)]
             for n in range(max(self.compartment_name.keys()) + 1)]
        x = copy.deepcopy(self.x)
        qhat = [[[[0 for t in range(self.Time)] for j in range(self.group)]
                 for m in range(max(self.compartment_name.keys()) + 1)]
                for n in range(max(self.compartment_name.keys()) + 1)]
        phat = [[[[0 for t in range(self.Time)] for k in range(self.group)] for j in range(self.group)]
                    for n in range(max(self.compartment_name.keys()) + 1)]

        for n in self.compartment_name.keys():
            for j in range(self.group):
                X[n][j][0] = self.compartment_info[n]['local']['population'][j]
                mulower[n][j][0] = self.compartment_info[n]['local']['population'][j]

        for t in range(self.Time - 1):

            for m in self.compartment_name.keys():
                for k in range(self.group):
                    if sum([x[m][mm][k][t] for mm in self.compartment_name.keys()]) > X[m][k][t]:
                        for mm in self.compartment_name.keys():
                            x[m][mm][k][t] = x[m][mm][k][t] * X[m][k][t] / \
                                             sum([x[m][mm][k][t] for mm in self.compartment_name.keys()])

            for n in self.compartment_name.keys():
                for j in range(self.group):
                    for k in range(self.group):
                        phat[n][j][k][t] = self.functionp(state=(n, j, k, t), X=X, x=x)

            for j in range(self.group):
                for n in self.compartment_name.keys():
                    for m in self.compartment_info[n]['tocpm'].keys():
                        qhat[n][m][j][t] = self.functionq(state=(n, m, j, t), X=X, x=x, phat=phat)
                    qhat_fromn = sum([qhat[n][m][j][t] for m in self.compartment_info[n]['tocpm'].keys() if m != n])
                    if qhat_fromn > 1:
                        qhat_fromn_dp = sum([qhat[n][m][j][t] for m in self.compartment_info[n]['tocpm'].keys()
                                             if self.compartment_info[n]['tocpm'][m]['q_dp'] != {}])
                        qhat_fromn_idp = qhat_fromn - qhat_fromn_dp
                        for m in self.compartment_info[n]['tocpm'].keys():
                            if self.compartment_info[n]['tocpm'][m]['q_dp'] != {}:
                                qhat[n][m][j][t] = qhat[n][m][j][t] * max(0, 1 - qhat_fromn_idp) / qhat_fromn_dp
                    qhat[n][n][j][t] = 1 - sum([qhat[n][mm][j][t] for mm in self.compartment_info[n]['tocpm'].keys()
                                                if n != mm])

            for n in self.compartment_name.keys():
                for j in range(self.group):
                    X[n][j][t + 1] = sum([x[m][n][j][t] for m in self.compartment_name.keys()]) \
                                     + sum(
                        [max(X[m][k][t] - sum([x[m][mm][k][t] for mm in self.compartment_name.keys()]), 0) *
                         qhat[m][n][j][t] * phat[m][k][j][t]
                         for m in self.compartment_name.keys() for k in range(self.group)])
            for n in self.compartment_name.keys():
                for j in range(self.group):
                    mulower[n][j][t + 1] = coef * sum([x[m][n][j][t] for m in self.compartment_name.keys()]) \
                                           - sum(
                        [max(X[m][k][t] - sum([x[m][mm][k][t] for mm in self.compartment_name.keys()]), 0)
                         * theta * log(
                            1 - qhat[m][n][j][t] * phat[m][k][j][t] + qhat[m][n][j][t] * phat[m][k][j][t] * exp(
                                -coef * 1.0 / theta)) for m in self.compartment_name.keys() for k in range(self.group)])
        return mulower

    def Prediction(self, opt:list):
        self.set_transition_compartment_self()
        for i in self.compartment_name:
            self.set_transition_group(compartment=i, prob=[[[1 if k == j else 0 for t in range(self.Time)]
                                                            for k in range(self.group)] for j in range(self.group)])
        if opt[0] == 'D':
            if len(opt) == 1:
                X, x, q = self.DeterPrediction()
            elif len(opt) == 2:
                X, x, q = self.DeterPrediction(types=opt[1])
            elif len(opt) == 3:
                X, x, q = self.DeterPrediction(types=opt[1],randomq=opt[2])
            return X
        elif opt[0] == 'S':
            if len(opt) == 1:
                X = self.StochaPrediction()
            elif len(opt) == 2:
                X = self.StochaPrediction(randomq=opt[1])
            return X
        elif opt[0] == 'RL':
            if len(opt) == 1:
                X = self.RobustPrediction_get_lowerbound()
            elif len(opt) == 2:
                X = self.RobustPrediction_get_lowerbound(theta=opt[1])
            elif len(opt) == 3:
                X = self.RobustPrediction_get_lowerbound(theta=opt[1], coef=opt[2])
            return X
        elif opt[0] == 'RU':
            if len(opt) == 1:
                X = self.RobustPrediction_get_upperbound()
            elif len(opt) == 2:
                X = self.RobustPrediction_get_upperbound(theta=opt[1])
            elif len(opt) == 3:
                X = self.RobustPrediction_get_upperbound(theta=opt[1], coef=opt[2])
            return X
        else:
            raise ValueError('opt should be a list and the first element should belong to {D, S, RL, RU}')

    def save_result(self, filename=None, compartment=None, dvar=None, custom:tuple=None, lhs=None):
        try:
            os.remove(filename+'.xlsx')
        except:
            pass
        outwb = openpyxl.Workbook()  # 打开一个将写的文件
        if compartment != None:
            outws = outwb.create_sheet('Simulated compartment')  # 在将写的文件创建sheet
            head = ['Sample no.', 'Compartment name', 'Group'] \
                   + ['T = ' + str(i) for i in range(len(compartment[0][0][0]))]
            for i in range(len(head)):
                outws.cell(row=1, column=1+i).value = head[i]
            p = 0
            for i in range(len(compartment)):
                for n in range(len(compartment[i])):
                    for j in range(len(compartment[i][n])):
                        outws.cell(row=2 + p, column=1).value = i + 1
                        outws.cell(row=2 + p, column=2).value = self.compartment_name[n]
                        outws.cell(row=2 + p, column=3).value = j
                        for t in range(len(compartment[i][n][j])):
                            outws.cell(row=2 + p, column=4 + t).value = compartment[i][n][j][t]
                        p += 1
            outws2 = outwb.create_sheet('Simulated compartment (total)')  # 在将写的文件创建sheet
            head = ['Sample no.', 'Compartment name'] \
                   + ['T = ' + str(i) for i in range(len(compartment[0][0][0]))]
            for i in range(len(head)):
                outws2.cell(row=1, column=1+i).value=head[i]
            p = 0
            for i in range(len(compartment)):
                for n in range(len(self.compartment_name)):
                    outws2.cell(row=2 + p, column=1).value = i+1
                    outws2.cell(row=2 + p, column=2).value = self.compartment_name[n]
                    for t in range(len(compartment[i][n][0])):
                        outws2.cell(row=2 + p, column=3 + t).value = \
                            sum([compartment[i][n][j][t] for j in range(len(compartment[i][n]))])
                    p += 1
            for n in range(len(self.compartment_name)):
                outws2.cell(row=2 + p, column=1).value = 'avg'
                outws2.cell(row=2 + p, column=2).value = self.compartment_name[n]
                for t in range(len(compartment[0][n][0])):
                    outws2.cell(row=2 + p, column=3 + t).value = \
                        sum([compartment[i][n][j][t] for i in range(len(compartment))
                             for j in range(len(compartment[i][n]))])/ len(compartment)
                p += 1
        if dvar != None:
            outws3 = outwb.create_sheet('Varables')
            head = ['Compartment (From)', 'Compartment (To)', 'Groups'] \
                   + ['T = ' + str(i) for i in range(len(dvar[0][0][0]))]
            for i in range(len(head)):
                outws3.cell(row=1, column=1+i).value=head[i]
            ii = 0
            for n in range(len(dvar)):
                for m in range(len(dvar[n])):
                    if sum([dvar[n][m][j][t] for j in range(len(dvar[n][m])) for t in range(len(dvar[n][m][j]))]) != 0:
                        for j in range(len(dvar[n][m])):
                            outws3.cell(row=2 + ii, column=1).value = self.compartment_name[n]
                            outws3.cell(row=2 + ii, column=2).value = self.compartment_name[m]
                            outws3.cell(row=2 + ii, column=3).value = j
                            for t in range(len(dvar[n][m][j])):
                                outws3.cell(row=2 + ii, column=4 + t).value = dvar[n][m][j][t]
                            ii += 1
        if lhs != None:
            outws4 = outwb.create_sheet('LHS')
            outws4.cell(row=1, column=1).value = 'Value of the LHS of constraints and the objective function'
            outws4.cell(row=2, column=2).value = 'Obj/Constraint name'
            outws4.cell(row=2, column=3).value = 'LHS (Simulation)'
            ii = 0
            for i in range(len(lhs)):
                outws4.cell(row=3 + ii, column=1).value = lhs[i][0]
                for s in range(len(lhs[i][1])):
                    outws4.cell(row=3 + ii, column=2 + s).value = lhs[i][1][s]
                ii += 1

        if custom != None:
            outws5 = outwb.create_sheet('Customized varables')
            head = ['Name', 'value']
            for i in range(len(head)):
                outws5.cell(row=1, column=1 + i).value = head[i]
            ii = 0
            for n in custom.keys():
                outws5.cell(row=2 + ii, column=1).value = n
                outws5.cell(row=2 + ii, column=2).value = custom[n]
                ii += 1

        # workbook = xlwt.Workbook(encoding='utf-8')
        # if compartment != None:
        #     worksheet1 = workbook.add_sheet('Simulated compartment')
        #     worksheet1.write(0, 0, 'Compartment')
        #     worksheet1.write(1, 0, 'Sample no.')
        #     worksheet1.write(1, 1, 'Compartment name')
        #     worksheet1.write(1, 2, 'Group')
        #     worksheet1.write(1, 3, 'Time Sequence')
        #     p = 0
        #     for i in range(len(compartment)):
        #         for n in range(len(compartment[i])):
        #             for j in range(len(compartment[i][n])):
        #                 worksheet1.write(2 + p, 0, i+1)
        #                 worksheet1.write(2 + p, 1, self.compartment_name[n])
        #                 worksheet1.write(2 + p, 2, j)
        #                 for t in range(len(compartment[i][n][j])):
        #                     worksheet1.write(2 + p, 3 + t, compartment[i][n][j][t])
        #                 p += 1
        #     worksheet2 = workbook.add_sheet('Simulated compartment (total)')
        #     worksheet2.write(0, 0, 'Compartment')
        #     worksheet2.write(1, 0, 'Sample no.')
        #     worksheet2.write(1, 1, 'Compartment name')
        #     worksheet2.write(1, 2, 'Time Sequence')
        #     p = 0
        #     for i in range(len(compartment)):
        #         for n in range(len(self.compartment_name)):
        #             worksheet2.write(2 + p, 0, i+1)
        #             worksheet2.write(2 + p, 1, self.compartment_name[n])
        #             for t in range(self.Time):
        #                 worksheet2.write(2 + p, 2 + t, sum([compartment[i][n][j][t] for j in range(len(compartment[i][n]))]))
        #             p += 1
        #     for n in range(len(self.compartment_name)):
        #         worksheet2.write(2 + p, 0, 'avg')
        #         worksheet2.write(2 + p, 1, self.compartment_name[n])
        #         for t in range(self.Time):
        #             worksheet2.write(2 + p, 2 + t,
        #                              sum([compartment[i][n][j][t] for i in range(len(compartment))
        #                                   for j in range(len(compartment[i][n]))])/ len(compartment) )
        #         p += 1
        # if dvar != None:
        #     worksheet4 = workbook.add_sheet('Varables')
        #     worksheet4.write(0, 0, 'Variables')
        #     worksheet4.write(1, 0, 'Compartment (From)')
        #     worksheet4.write(1, 1, 'Compartment (To)')
        #     worksheet4.write(1, 2, 'Groups')
        #     worksheet4.write(1, 3, 'Time Sequence')
        #     ii = 0
        #     for n in range(len(dvar)):
        #         for m in range(len(dvar[n])):
        #             if sum([dvar[n][m][j][t] for j in range(len(dvar[n][m])) for t in range(len(dvar[n][m][j]))]) != 0:
        #                 for j in range(len(dvar[n][m])):
        #                     worksheet4.write(2 + ii, 0, self.compartment_name[n])
        #                     worksheet4.write(2 + ii, 1, self.compartment_name[m])
        #                     worksheet4.write(2 + ii, 2, j)
        #                     for t in range(len(dvar[n][m][j])):
        #                         worksheet4.write(2 + ii, 3 + t, dvar[n][m][j][t])
        #                     ii += 1
        # if lhs != None:
        #     worksheet5 = workbook.add_sheet('LHS')
        #     worksheet5.write(0, 0, 'Value of the LHS of constraints and the objective function')
        #     worksheet5.write(1, 0, 'Obj/Constraint name')
        #     worksheet5.write(1, 1, 'LHS (Simulation)')
        #     ii = 0
        #     for i in range(len(lhs)):
        #         worksheet5.write(2 + ii, 0, lhs[i][0])
        #         for s in range(len(lhs[i][1])):
        #             worksheet5.write(2 + ii, 1 + s, lhs[i][1][s])
        #         ii += 1
        # workbook.save(filename + '.xls')

        outwb.save(filename + '.xlsx')
        outwb.close()

    def get_solution_compartment(self, compartment=None):
        if isinstance(compartment, str):
            compartment = [compartment]
        n = [0 for i in range(len(compartment))]
        X = {}
        for i in range(len(compartment)):
            n[i] = get_keys(self.compartment_name, compartment[i])[0]
            try:
                X.update({compartment[i]: [
                    [round(self.bestsol[2][n[i]][j][t], 6) for t in range(self.Time)] for j in range(self.group)]})
            except:
                if self.solver == 'cplex':
                    X.update({compartment[i]: [
                        [round(self.model.solution.get_values('X.'+str(n[i])+'.'+str(j)+'.'+str(t)), 6)
                         for t in range(self.Time)] for j in range(self.group)]})
                elif self.solver == 'gurobi':
                    X.update({compartment[i]: [
                        [round(self.model.getVarByName('X.' + str(n[i]) + '.' + str(j) + '.' + str(t)).x, 6)
                         for t in range(self.Time)] for j in range(self.group)]})

        return X

    def get_solution_flow_variable(self, compartmentfrom=None, compartmentto=None):
        if isinstance(compartmentfrom, str):
            compartmentfrom = [compartmentfrom]
        if isinstance(compartmentto, str):
            compartmentto = [compartmentto]
        if len(compartmentfrom) == len(compartmentto):
            n = [0 for i in range(len(compartmentfrom))]
            m = [0 for i in range(len(compartmentto))]
            x = {}
            for i in range(len(compartmentfrom)):
                n[i] = get_keys(self.compartment_name, compartmentfrom[i])[0]
                m[i] = get_keys(self.compartment_name, compartmentto[i])[0]
                try:
                    if self.solver == 'cplex':
                        x.update({(compartmentfrom[i], compartmentto[i]): [
                            [round(self.bestsol[3][n[i]][m[i]][j][t], 6)
                             for t in range(self.Time)] for j in range(self.group)]})
                except:
                    if self.solver == 'cplex':
                        x.update({(compartmentfrom[i],compartmentto[i]): [
                            [round(self.model.solution.get_values('x.' + str(n[i]) + '.' + str(m[i]) + '.'
                                                                  + str(j) + '.' + str(t)), 6)
                             for t in range(self.Time)] for j in range(self.group)]})
                    elif self.solver == 'gurobi':
                        x.update({(compartmentfrom[i], compartmentto[i]): [
                            [round(self.model.getVarByName('x.' + str(n[i]) + '.' + str(m[i]) + '.'
                            + str(j) + '.' + str(t)).x, 6) for t in range(self.Time)] for j in range(self.group)]})
            return x
        else:
            raise ValueError('lens of compartmentfrom and compartmentto are not same')

    def get_solution_custom_variable(self,name=None):
        if not isinstance(name, list):
            name = [name]
        else:
            X = {}
            for i in range(len(name)):
                if name[i] in self.custom_var.keys():
                    try:
                        if self.solver == 'cplex':
                            X.update({name[i]: self.bestsol[-1][name[i]]})
                    except:
                        if self.solver == 'cplex':
                            X.update({name[i]: self.model.solution.get_values(name[i])})
                        elif self.solver == 'gurobi':
                            X.update({name[i]: self.model.getVarByName(name[i]).x})
                else:
                    raise ValueError('name is not found')
            return X

    def Solution_print(self,X:tuple=None, x:tuple=None, xc:tuple=None):
        if self.status != 0:
            if X != None:
                print('============== Compartment ===============')
                for n in X.keys():
                    print(n)
                    for j in range(self.group):
                        print(X[n][j])
                    print('Total = ', [sum([X[n][j][t] for j in range(self.group)]) for t in range(self.Time)])
            if x != None:
                print('============= Flow decision variable ================')
                for n in x.keys():
                    print('From compartment ', n[0], 'to compartment', n[1])
                    for j in range(self.group):
                        print(x[n][j])

            if xc != None:
                print('============= Customized decision variable ================')
                for n in xc.keys():
                    for j in range(self.group):
                        print(n, xc[n])
        else:
            raise ValueError('No solution')
    def set_estimated_parameter_bound(self, ub:dict, lb:dict, latin=1000):
        for n in ub.keys():
            if len(n) == 1:
                temp = [[random.uniform((ub[n][j] - lb[n][j]) * i / latin, (ub[n][j] - lb[n][j]) * (i + 1) / latin)
                    for i in range(latin - 1)] for j in range(self.group)]
                temp = random.shuffle(temp)
                self.latin_params1.update({n: temp})
            elif len(n) == 2:
                temp = [[random.uniform((ub[n][j] - lb[n][j]) * i / latin, (ub[n][j] - lb[n][j]) * (i + 1) / latin)
                         for i in range(latin - 1)] for j in range(self.group)]
                temp = random.shuffle(temp)
                self.latin_params2.update({n: temp})
            elif len(n) == 3:
                temp = [[random.uniform((ub[n][j] - lb[n][j]) * i / latin, (ub[n][j] - lb[n][j]) * (i + 1) / latin)
                         for i in range(latin - 1)] for j in range(self.group)]
                temp = random.shuffle(temp)
                self.latin_params3.update({n: temp})
    def set_log_stream_SO(self, label=0, file:str='model_log_stream'):
        self.__log_stream = label
        self.__log_file = file